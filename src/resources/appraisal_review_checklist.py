"""
Appraisal Review Checklist Generator
------------------------------------
This script creates an Excel workbook with a comprehensive appraisal review
checklist for real estate lending, including compliance checks, valuation 
methodology assessment, and data verification sections.

Dependencies:
- openpyxl
- pandas

Install with: pip install openpyxl pandas
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
import datetime
import os

def create_workbook():
    """Creates the Excel workbook with all sheets and formatting."""
    wb = openpyxl.Workbook()
    
    # Create sheets
    dashboard = wb.active
    dashboard.title = "Dashboard"
    general_info = wb.create_sheet("General Information")
    compliance = wb.create_sheet("Compliance Review")
    methodology = wb.create_sheet("Valuation Methodology")
    market_analysis = wb.create_sheet("Market Analysis")
    property_analysis = wb.create_sheet("Property Analysis")
    
    # Set up each sheet
    setup_dashboard(dashboard)
    setup_general_info(general_info)
    setup_compliance(compliance)
    setup_methodology(methodology)
    setup_market_analysis(market_analysis)
    setup_property_analysis(property_analysis)
    
    # Save the workbook
    filename = "Real_Estate_Appraisal_Review_Checklist.xlsx"
    wb.save(filename)
    
    print(f"Created {filename} successfully!")
    
    return filename

def setup_dashboard(ws):
    """Sets up the Dashboard sheet with summary information."""
    # Set column widths
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Title and Introduction
    ws.merge_cells('A1:I1')
    ws['A1'] = "REAL ESTATE APPRAISAL REVIEW CHECKLIST"
    ws['A1'].font = Font(size=18, bold=True, color="1B4620")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Add Clarity Impact Finance branding
    ws.merge_cells('A2:I2')
    ws['A2'] = "Clarity Impact Finance - Supporting Mission-Driven Lenders"
    ws['A2'].font = Font(size=12, italic=True, color="E67E45")
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Instructions
    ws.merge_cells('A4:I4')
    ws['A4'] = "DASHBOARD"
    ws['A4'].font = Font(size=14, bold=True)
    ws['A4'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A5:I6')
    ws['A5'] = "This workbook contains a comprehensive appraisal review checklist to help ensure compliance and proper valuation assessment for mission-driven lenders and CDFIs."
    ws['A5'].alignment = Alignment(wrap_text=True)
    
    # Summary section
    ws['A8'] = "APPRAISAL REVIEW SUMMARY"
    ws['A8'].font = Font(size=12, bold=True)
    
    headers = [
        "Review Date", "Property Type", "Appraised Value", 
        "Compliance Score", "Methodology Score", "Market Analysis Score",
        "Property Analysis Score", "Overall Score"
    ]
    
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=i)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # Example data row
    example_data = [
        "03/20/2025", "Commercial Office", "$2,450,000", 
        "92%", "88%", "95%", 
        "90%", "91%"
    ]
    
    for i, data in enumerate(example_data, 1):
        cell = ws.cell(row=10, column=i)
        cell.value = data
        cell.alignment = Alignment(horizontal='center')
    
    # Key findings section
    ws['A12'] = "KEY FINDINGS AND RECOMMENDATIONS"
    ws['A12'].font = Font(size=12, bold=True)
    
    headers = ["Finding Category", "Description", "Risk Level", "Action Required"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=13, column=i)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # Example findings
    findings = [
        ("Compliance", "Appraisal is missing required USPAP certification.", "High", "Request updated appraisal with proper certification."),
        ("Methodology", "Income approach relies on optimistic rent growth assumptions.", "Medium", "Recalculate value using more conservative assumptions."),
        ("Market Analysis", "Comparable sales are from different submarket.", "Medium", "Request additional comparables from subject's submarket."),
        ("Property", "Deferred maintenance may be underestimated.", "Low", "Consider adjustment to value or additional property inspection.")
    ]
    
    for row_idx, finding in enumerate(findings, 14):
        for col_idx, data in enumerate(finding, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = data
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            # Add color coding for risk level
            if col_idx == 3:
                if data == "High":
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif data == "Medium":
                    cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                elif data == "Low":
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Add instructions
    ws['A18'] = "INSTRUCTIONS"
    ws['A18'].font = Font(size=12, bold=True)
    
    ws.merge_cells('A19:I22')
    ws['A19'] = """1. Fill out the General Information sheet with details about the property and appraisal
2. Complete each review checklist by marking items as Y (Yes), N (No), or NA (Not Applicable)
3. Add comments to explain any "No" responses or to provide additional context
4. Identify high-risk issues and document recommended actions
5. Review the Dashboard summary to assess overall appraisal quality"""
    ws['A19'].alignment = Alignment(wrap_text=True)

def setup_general_info(ws):
    """Sets up the General Information sheet."""
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    
    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = "GENERAL APPRAISAL INFORMATION"
    ws['A1'].font = Font(size=16, bold=True, color="1B4620")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Instructions
    ws.merge_cells('A2:B3')
    ws['A2'] = "Complete the fields below with information about the property and the appraisal being reviewed."
    ws['A2'].alignment = Alignment(wrap_text=True)
    
    # General information section
    sections = [
        # Appraisal Information
        ("Appraisal Information", [
            ("Appraisal Date", "01/15/2025"),
            ("Effective Date of Value", "01/15/2025"),
            ("Date of Review", "03/20/2025"),
            ("Appraiser Name", "John Smith, MAI"),
            ("Appraiser License #", "AG123456"),
            ("Appraisal Firm", "Smith Valuation Services"),
            ("Report Type", "Full Narrative"),
        ]),
        # Subject Property Information
        ("Subject Property Information", [
            ("Property Address", "123 Main Street"),
            ("City", "Springfield"),
            ("State", "IL"),
            ("Zip Code", "62701"),
            ("Property Type", "Commercial Office"),
            ("Property Size/Units", "25,000 SF"),
            ("Year Built", "2005"),
            ("Parcel Number(s)", "123-456-789"),
            ("Zoning", "C-2 Commercial"),
        ]),
        # Value Information
        ("Value Information", [
            ("Appraised Value", "$2,450,000"),
            ("Value per SF/Unit", "$98.00/SF"),
            ("Value Approach Used", "Income & Sales Comparison"),
            ("Cap Rate (if applicable)", "7.5%"),
            ("Requested Loan Amount", "$1,837,500"),
            ("Loan-to-Value Ratio", "75%"),
        ]),
        # Client Information
        ("Client Information", [
            ("Client/Lender Name", "Community First CDFI"),
            ("Loan Officer", "Jane Wilson"),
            ("Loan/File Number", "L-2025-0123"),
            ("Borrower Name", "Main Street Holdings LLC"),
        ]),
    ]
    
    row = 5
    for section_title, fields in sections:
        # Section header
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = section_title
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        row += 1
        
        # Fields
        for field_name, example_value in fields:
            ws[f'A{row}'] = field_name
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = example_value
            
            # Add borders
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            row += 1
        
        # Add spacing between sections
        row += 1

def setup_compliance(ws):
    """Sets up the Compliance Review sheet."""
    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    
    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = "COMPLIANCE REVIEW"
    ws['A1'].font = Font(size=16, bold=True, color="1B4620")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Instructions
    ws.merge_cells('A2:C2')
    ws['A2'] = "Review the appraisal for compliance with standards, regulations, and requirements. Mark each item as 'Y' (Yes), 'N' (No), or 'NA' (Not Applicable)."
    ws['A2'].alignment = Alignment(wrap_text=True)
    
    # Headers
    headers = ["Review Item", "Compliant? (Y/N/NA)", "Comments"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col).value = header
        ws.cell(row=4, column=col).font = Font(bold=True)
        ws.cell(row=4, column=col).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        ws.cell(row=4, column=col).alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Compliance checklist items with example data
    checklist_items = [
        # USPAP Compliance
        ("USPAP Compliance", [
            ("Appraisal includes USPAP certification statement", "N", "Missing required certification on page 2"),
            ("Appraisal includes required USPAP limiting conditions", "Y", ""),
            ("Appraiser is competent for this assignment", "Y", ""),
            ("Appraiser has disclosed any prior services for this property", "Y", ""),
            ("Report complies with Ethics Rule requirements", "Y", ""),
            ("Report complies with Record Keeping Rule requirements", "Y", ""),
            ("Report complies with Competency Rule requirements", "Y", ""),
            ("Report complies with Scope of Work Rule requirements", "Y", ""),
        ]),
        
        # Regulatory Compliance
        ("Regulatory Compliance", [
            ("Appraisal complies with federal financial institutions regulations", "Y", ""),
            ("Appraisal meets Interagency Appraisal Guidelines", "Y", ""),
            ("Appraisal is FIRREA compliant (if applicable)", "Y", ""),
            ("Appraisal meets Fannie Mae/Freddie Mac requirements (if applicable)", "NA", "Not applicable for commercial property"),
            ("Appraisal complies with applicable state regulations", "Y", ""),
        ]),
        
        # Report Requirements
        ("Report Requirements", [
            ("Report includes property identification", "Y", ""),
            ("Report states intended use and users", "Y", ""),
            ("Report has appropriate effective date of value", "Y", ""),
            ("Report includes appropriate scope of work", "Y", ""),
            ("Report correctly identifies real property interest appraised", "Y", ""),
            ("Report includes all required extraordinary assumptions", "Y", ""),
            ("Report includes all required hypothetical conditions", "Y", ""),
            ("Report is in the agreed-upon format (narrative, form, etc.)", "Y", ""),
        ]),
        
        # Appraiser Qualifications
        ("Appraiser Qualifications", [
            ("Appraiser has appropriate license/certification for assignment", "Y", ""),
            ("Appraiser license is current and in good standing", "Y", ""),
            ("Appraiser has geographic competence in subject market", "Y", ""),
            ("Appraiser has demonstrated property type competence", "Y", ""),
        ]),
    ]
    
    row = 5
    for section_title, items in checklist_items:
        # Section header
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = section_title
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        row += 1
        
        # Items
        for item_name, example_value, comment in items:
            ws[f'A{row}'] = item_name
            ws[f'B{row}'] = example_value
            ws[f'C{row}'] = comment
            
            # Add borders and color coding for compliance status
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Color code based on compliance status
            if example_value == "N":
                ws[f'B{row}'].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            elif example_value == "Y":
                ws[f'B{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            
            row += 1
        
        # Add spacing between sections
        row += 1

def setup_methodology(ws):
    """Sets up the Valuation Methodology sheet."""
    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    
    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = "VALUATION METHODOLOGY REVIEW"
    ws['A1'].font = Font(size=16, bold=True, color="1B4620")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Instructions
    ws.merge_cells('A2:C2')
    ws['A2'] = "Review the appraisal's valuation methodology and approaches. Mark each item as 'Y' (Yes), 'N' (No), or 'NA' (Not Applicable)."
    ws['A2'].alignment = Alignment(wrap_text=True)
    
    # Headers
    headers = ["Review Item", "Acceptable? (Y/N/NA)", "Comments"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col).value = header
        ws.cell(row=4, column=col).font = Font(bold=True)
        ws.cell(row=4, column=col).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        ws.cell(row=4, column=col).alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Methodology checklist items with example data
    methodology_items = [
        ("Appropriate approaches to value were used", "Y", ""),
        ("Sufficient explanation for omitted approaches", "Y", ""),
        ("Sales Comparison Approach properly applied", "Y", ""),
        ("Comparable properties are truly comparable", "N", "Comps 2 and 3 are in higher-value submarkets"),
        ("Appropriate adjustments made to comparable sales", "Y", ""),
        ("Income Approach properly applied", "Y", ""),
        ("Market rents are well-supported", "Y", ""),
        ("Vacancy and collection loss estimate is reasonable", "Y", ""),
        ("Operating expense estimates are well-supported", "Y", ""),
        ("Capitalization rate is appropriate and well-supported", "N", "Cap rate seems aggressive - area average is 8.0%"),
        ("Cost Approach properly applied (if used)", "NA", "Not used"),
        ("Land value is well-supported", "NA", "Not applicable"),
        ("Cost estimates are from reliable sources", "NA", "Not applicable"),
        ("Depreciation estimates are reasonable", "NA", "Not applicable"),
        ("Reconciliation of value approaches is logical", "Y", ""),
        ("Final value conclusion is well-supported", "N", "Overweights income approach with aggressive assumptions")
    ]
    
    row = 5
    for item_name, example_value, comment in methodology_items:
        ws[f'A{row}'] = item_name
        ws[f'B{row}'] = example_value
        ws[f'C{row}'] = comment
        
        # Add borders and color coding
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Color code based on status
        if example_value == "N":
            ws[f'B{row}'].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        elif example_value == "Y":
            ws[f'B{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        row += 1

def setup_market_analysis(ws):
    """Sets up the Market Analysis sheet."""
    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    
    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = "MARKET ANALYSIS REVIEW"
    ws['A1'].font = Font(size=16, bold=True, color="1B4620")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Instructions
    ws.merge_cells('A2:C2')
    ws['A2'] = "Review the market analysis section of the appraisal. Mark each item as 'Y' (Yes), 'N' (No), or 'NA' (Not Applicable)."
    ws['A2'].alignment = Alignment(wrap_text=True)
    
    # Headers
    headers = ["Review Item", "Acceptable? (Y/N/NA)", "Comments"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col).value = header
        ws.cell(row=4, column=col).font = Font(bold=True)
        ws.cell(row=4, column=col).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        ws.cell(row=4, column=col).alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Market analysis checklist items with example data
    market_items = [
        ("National economic trends properly analyzed", "Y", ""),
        ("Regional economic trends properly analyzed", "Y", ""),
        ("Local economic trends properly analyzed", "Y", ""),
        ("Neighborhood description is accurate and complete", "Y", ""),
        ("Market area delineation is appropriate", "Y", ""),
        ("Supply and demand factors are well-researched", "Y", ""),
        ("Competitive properties properly identified", "N", "Missing two new office properties within 1 mile"),
        ("Market trends are accurately identified", "Y", ""),
        ("Vacancy rates are well-researched and supported", "Y", ""),
        ("Absorption rates are well-researched and supported", "Y", ""),
        ("Rental rates are well-researched and supported", "Y", ""),
        ("Market forecast is reasonable and supported", "N", "Rent growth projections above market consensus")
    ]
    
    row = 5
    for item_name, example_value, comment in market_items:
        ws[f'A{row}'] = item_name
        ws[f'B{row}'] = example_value
        ws[f'C{row}'] = comment
        
        # Add borders and color coding
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Color code based on status
        if example_value == "N":
            ws[f'B{row}'].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        elif example_value == "Y":
            ws[f'B{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        row += 1

def setup_property_analysis(ws):
    """Sets up the Property Analysis sheet."""
    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    
    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = "PROPERTY ANALYSIS REVIEW"
    ws['A1'].font = Font(size=16, bold=True, color="1B4620")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Instructions
    ws.merge_cells('A2:C2')
    ws['A2'] = "Review the property analysis section of the appraisal. Mark each item as 'Y' (Yes), 'N' (No), or 'NA' (Not Applicable)."
    ws['A2'].alignment = Alignment(wrap_text=True)
    
    # Headers
    headers = ["Review Item", "Acceptable? (Y/N/NA)", "Comments"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col).value = header
        ws.cell(row=4, column=col).font = Font(bold=True)
        ws.cell(row=4, column=col).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        ws.cell(row=4, column=col).alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Property analysis checklist items with example data
    property_items = [
        ("Site description is accurate and complete", "Y", ""),
        ("Site improvements accurately described", "Y", ""),
        ("Building description is accurate and complete", "Y", ""),
        ("Floor plans/layout accurately described", "Y", ""),
        ("Building measurements/square footage verified", "Y", ""),
        ("Construction quality accurately described", "Y", ""),
        ("Building condition accurately described", "N", "Downplays needed roof repairs"),
        ("Deferred maintenance properly identified", "N", "HVAC system issues not mentioned"),
        ("Functional utility appropriately analyzed", "Y", ""),
        ("External obsolescence appropriately analyzed", "Y", ""),
        ("Zoning analysis is accurate and complete", "Y", ""),
        ("Flood zone information is accurate", "Y", ""),
        ("Environmental issues properly addressed", "Y", ""),
        ("ADA compliance issues addressed (if applicable)", "Y", ""),
        ("Property photographs are adequate and recent", "Y", "")
    ]
    
    row = 5
    for item_name, example_value, comment in property_items:
        ws[f'A{row}'] = item_name
        ws[f'B{row}'] = example_value
        ws[f'C{row}'] = comment
        
        # Add borders and color coding
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Color code based on status
        if example_value == "N":
            ws[f'B{row}'].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        elif example_value == "Y":
            ws[f'B{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        row += 1

if __name__ == "__main__":
    create_workbook()
