"""
Cash Flow Projection Template with Dashboard and Guidance Notes

This script creates a comprehensive Excel template for cash flow projections,
including a dashboard, guidance notes, and detailed projection worksheets.
The template helps businesses forecast their cash position over time.

Created for Clarity Impact Finance
"""

import os
import sys
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Font, 
    PatternFill, 
    Border, 
    Side, 
    Alignment, 
    Color, 
    Protection,
    colors
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.chart import LineChart, BarChart, PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.drawing.image import Image
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation

# Constants for styling
GREEN_FILL = PatternFill(start_color="1B4620", end_color="1B4620", fill_type="solid")  # Dark Green
LIGHT_GREEN_FILL = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")  # Light Green
BLUE_FILL = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")  # Blue
LIGHT_BLUE_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")  # Light Blue for data entry
GREY_FILL = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")  # Light Grey for headers
RED_FILL = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")  # Red for negative cash flow
ORANGE_FILL = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")  # Orange for warnings

# Company brand colors
CIF_GREEN = PatternFill(start_color="00A651", end_color="00A651", fill_type="solid")  # Clarity Impact Finance green
CIF_ORANGE = PatternFill(start_color="F7941D", end_color="F7941D", fill_type="solid")  # Clarity Impact Finance orange

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

THICK_BORDER = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

# Title font
TITLE_FONT = Font(name='Arial', size=14, bold=True, color="FFFFFF")
HEADER_FONT = Font(name='Arial', size=12, bold=True)
NORMAL_FONT = Font(name='Arial', size=10)
HIGHLIGHT_FONT = Font(name='Arial', size=10, bold=True, color="1B4620")

def create_workbook(output_path):
    """Create the Cash Flow Projection Template workbook with all worksheets."""
    wb = Workbook()
    
    # Rename the default sheet to "Dashboard"
    dashboard_sheet = wb.active
    dashboard_sheet.title = "Dashboard"
    
    # Create all worksheets first
    input_sheet = wb.create_sheet("Input")
    monthly_sheet = wb.create_sheet("Monthly Projection")
    annual_sheet = wb.create_sheet("Annual Summary")
    assumptions_sheet = wb.create_sheet("Assumptions")
    guidance_sheet = wb.create_sheet("Guidance")
    
    # Set up all worksheets
    setup_input_tab(wb, input_sheet)
    setup_monthly_projection(wb, monthly_sheet)
    setup_annual_summary(wb, annual_sheet)
    setup_assumptions(wb, assumptions_sheet)
    setup_guidance(wb, guidance_sheet)
    
    # Setup dashboard last since it references other sheets
    setup_dashboard(wb, dashboard_sheet)
    
    # Save the workbook
    wb.save(output_path)
    print(f"Cash Flow Projection Template created successfully at {output_path}")
    
    return wb

def setup_dashboard(wb, sheet):
    """Set up the Dashboard worksheet with key metrics and visualizations"""
    # Set column widths
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15
    
    # Page Title
    sheet.merge_cells('A1:F1')
    sheet['A1'] = "CASH FLOW PROJECTION DASHBOARD"
    sheet['A1'].font = TITLE_FONT
    sheet['A1'].fill = CIF_GREEN
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Add introduction text
    sheet.merge_cells('A2:F3')
    sheet['A2'] = "This dashboard provides a visual overview of your cash flow projections. Use the charts and key metrics below to monitor your business's financial health and make informed decisions."
    sheet['A2'].font = Font(name='Arial', size=10)
    sheet['A2'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Cash Flow Status section
    row = 5
    sheet.merge_cells(f'A{row}:F{row}')
    sheet[f'A{row}'] = "CASH FLOW STATUS"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = BLUE_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Key metrics header
    row += 1
    for col, header in zip(['A', 'B', 'C', 'D', 'E', 'F'], 
                          ['Metric', 'Starting', 'Peak', 'Low', 'Ending', 'Net Change']):
        sheet[f'{col}{row}'] = header
        sheet[f'{col}{row}'].font = Font(name='Arial', size=10, bold=True)
        sheet[f'{col}{row}'].fill = GREY_FILL
        sheet[f'{col}{row}'].alignment = Alignment(horizontal='center')
    
    # Metrics
    metrics = [
        "Cash Balance",
        "Monthly Cash Flow",
        "Cumulative Cash Flow",
        "Revenue",
        "Expenses",
        "Cash Flow Ratio"
    ]
    
    # Add metrics rows
    for i, metric in enumerate(metrics):
        row += 1
        sheet[f'A{row}'] = metric
        sheet[f'A{row}'].font = NORMAL_FONT
        sheet[f'A{row}'].alignment = Alignment(horizontal='left')
        
        if metric == "Cash Balance":
            sheet[f'B{row}'] = "=Input!D13"  # Starting balance from Input sheet
            sheet[f'C{row}'] = "=MAX('Monthly Projection'!O6:O17)"  # Peak balance from Monthly Projection
            sheet[f'D{row}'] = "=MIN('Monthly Projection'!O6:O17)"  # Low balance from Monthly Projection
            sheet[f'E{row}'] = "='Monthly Projection'!O17"  # Ending balance (December)
            sheet[f'F{row}'] = "=E7-B7"  # Net change in cash
        elif metric == "Monthly Cash Flow":
            sheet[f'B{row}'] = "='Monthly Projection'!N6"  # January net cash flow
            sheet[f'C{row}'] = "=MAX('Monthly Projection'!N6:N17)"  # Peak monthly cash flow
            sheet[f'D{row}'] = "=MIN('Monthly Projection'!N6:N17)"  # Low monthly cash flow
            sheet[f'E{row}'] = "='Monthly Projection'!N17"  # December net cash flow
            sheet[f'F{row}'] = "=SUM('Monthly Projection'!N6:N17)"  # Sum of all monthly cash flows
        elif metric == "Cumulative Cash Flow":
            sheet[f'B{row}'] = "='Monthly Projection'!N6"  # January (same as monthly for first month)
            sheet[f'C{row}'] = "=MAX('Monthly Projection'!N6,'Monthly Projection'!N6+'Monthly Projection'!N7,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8+'Monthly Projection'!N9,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8+'Monthly Projection'!N9+'Monthly Projection'!N10,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8+'Monthly Projection'!N9+'Monthly Projection'!N10+'Monthly Projection'!N11,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8+'Monthly Projection'!N9+'Monthly Projection'!N10+'Monthly Projection'!N11+'Monthly Projection'!N12,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8+'Monthly Projection'!N9+'Monthly Projection'!N10+'Monthly Projection'!N11+'Monthly Projection'!N12+'Monthly Projection'!N13,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8+'Monthly Projection'!N9+'Monthly Projection'!N10+'Monthly Projection'!N11+'Monthly Projection'!N12+'Monthly Projection'!N13+'Monthly Projection'!N14,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8+'Monthly Projection'!N9+'Monthly Projection'!N10+'Monthly Projection'!N11+'Monthly Projection'!N12+'Monthly Projection'!N13+'Monthly Projection'!N14+'Monthly Projection'!N15,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8+'Monthly Projection'!N9+'Monthly Projection'!N10+'Monthly Projection'!N11+'Monthly Projection'!N12+'Monthly Projection'!N13+'Monthly Projection'!N14+'Monthly Projection'!N15+'Monthly Projection'!N16,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8+'Monthly Projection'!N9+'Monthly Projection'!N10+'Monthly Projection'!N11+'Monthly Projection'!N12+'Monthly Projection'!N13+'Monthly Projection'!N14+'Monthly Projection'!N15+'Monthly Projection'!N16+'Monthly Projection'!N17)"
            sheet[f'D{row}'] = "=MIN('Monthly Projection'!N6,'Monthly Projection'!N6+'Monthly Projection'!N7,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8+'Monthly Projection'!N9,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8+'Monthly Projection'!N9+'Monthly Projection'!N10,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8+'Monthly Projection'!N9+'Monthly Projection'!N10+'Monthly Projection'!N11,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8+'Monthly Projection'!N9+'Monthly Projection'!N10+'Monthly Projection'!N11+'Monthly Projection'!N12,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8+'Monthly Projection'!N9+'Monthly Projection'!N10+'Monthly Projection'!N11+'Monthly Projection'!N12+'Monthly Projection'!N13,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8+'Monthly Projection'!N9+'Monthly Projection'!N10+'Monthly Projection'!N11+'Monthly Projection'!N12+'Monthly Projection'!N13+'Monthly Projection'!N14,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8+'Monthly Projection'!N9+'Monthly Projection'!N10+'Monthly Projection'!N11+'Monthly Projection'!N12+'Monthly Projection'!N13+'Monthly Projection'!N14+'Monthly Projection'!N15,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8+'Monthly Projection'!N9+'Monthly Projection'!N10+'Monthly Projection'!N11+'Monthly Projection'!N12+'Monthly Projection'!N13+'Monthly Projection'!N14+'Monthly Projection'!N15+'Monthly Projection'!N16,'Monthly Projection'!N6+'Monthly Projection'!N7+'Monthly Projection'!N8+'Monthly Projection'!N9+'Monthly Projection'!N10+'Monthly Projection'!N11+'Monthly Projection'!N12+'Monthly Projection'!N13+'Monthly Projection'!N14+'Monthly Projection'!N15+'Monthly Projection'!N16+'Monthly Projection'!N17)"
            sheet[f'E{row}'] = "=SUM('Monthly Projection'!N6:N17)"  # Sum of all monthly cash flows
            sheet[f'F{row}'] = "=E9"  # Same as ending cumulative cash flow
        elif metric == "Revenue":
            sheet[f'B{row}'] = "='Monthly Projection'!B6"  # January revenue
            sheet[f'C{row}'] = "=MAX('Monthly Projection'!B6:B17)"  # Peak monthly revenue
            sheet[f'D{row}'] = "=MIN('Monthly Projection'!B6:B17)"  # Low monthly revenue
            sheet[f'E{row}'] = "='Monthly Projection'!B17"  # December revenue
            sheet[f'F{row}'] = "=SUM('Monthly Projection'!B6:B17)"  # Total annual revenue
        elif metric == "Expenses":
            sheet[f'B{row}'] = "='Monthly Projection'!H6"  # January expenses
            sheet[f'C{row}'] = "=MAX('Monthly Projection'!H6:H17)"  # Peak monthly expenses
            sheet[f'D{row}'] = "=MIN('Monthly Projection'!H6:H17)"  # Low monthly expenses
            sheet[f'E{row}'] = "='Monthly Projection'!H17"  # December expenses
            sheet[f'F{row}'] = "=SUM('Monthly Projection'!H6:H17)"  # Total annual expenses
        elif metric == "Cash Flow Ratio":
            sheet[f'B{row}'] = "=IF(B10=0,0,B8/B10)"  # January cash flow ratio
            sheet[f'C{row}'] = "=MAX('Monthly Projection'!B6/'Monthly Projection'!H6,'Monthly Projection'!B7/'Monthly Projection'!H7,'Monthly Projection'!B8/'Monthly Projection'!H8,'Monthly Projection'!B9/'Monthly Projection'!H9,'Monthly Projection'!B10/'Monthly Projection'!H10,'Monthly Projection'!B11/'Monthly Projection'!H11,'Monthly Projection'!B12/'Monthly Projection'!H12,'Monthly Projection'!B13/'Monthly Projection'!H13,'Monthly Projection'!B14/'Monthly Projection'!H14,'Monthly Projection'!B15/'Monthly Projection'!H15,'Monthly Projection'!B16/'Monthly Projection'!H16,'Monthly Projection'!B17/'Monthly Projection'!H17)"
            sheet[f'D{row}'] = "=MIN('Monthly Projection'!B6/'Monthly Projection'!H6,'Monthly Projection'!B7/'Monthly Projection'!H7,'Monthly Projection'!B8/'Monthly Projection'!H8,'Monthly Projection'!B9/'Monthly Projection'!H9,'Monthly Projection'!B10/'Monthly Projection'!H10,'Monthly Projection'!B11/'Monthly Projection'!H11,'Monthly Projection'!B12/'Monthly Projection'!H12,'Monthly Projection'!B13/'Monthly Projection'!H13,'Monthly Projection'!B14/'Monthly Projection'!H14,'Monthly Projection'!B15/'Monthly Projection'!H15,'Monthly Projection'!B16/'Monthly Projection'!H16,'Monthly Projection'!B17/'Monthly Projection'!H17)"
            sheet[f'E{row}'] = "=IF(E10=0,0,E8/E10)"  # December cash flow ratio
            sheet[f'F{row}'] = "=IF(F10=0,0,F8/F10)"  # Annual cash flow ratio
            
        # Format cells based on metric type
        for col in ['B', 'C', 'D', 'E', 'F']:
            if metric in ["Cash Balance", "Monthly Cash Flow", "Cumulative Cash Flow", "Revenue", "Expenses"]:
                sheet[f'{col}{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
            elif metric == "Cash Flow Ratio":
                sheet[f'{col}{row}'].number_format = '0.00'
    
    # Cash Flow Chart
    chart_title = "Monthly Cash Flow"
    row = 14
    sheet.merge_cells(f'A{row}:F{row}')
    sheet[f'A{row}'] = chart_title
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = BLUE_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Create monthly cash flow chart
    chart1 = BarChart()
    chart1.title = chart_title
    chart1.style = 10
    chart1.y_axis.title = "Amount"
    chart1.x_axis.title = "Month"
    
    # Create data references
    months = Reference(wb["Monthly Projection"], min_col=2, min_row=4, max_row=4, max_col=14)
    data = Reference(wb["Monthly Projection"], min_col=14, min_row=5, max_row=17, max_col=14)
    
    chart1.add_data(data)
    chart1.set_categories(months)
    chart1.shape = 4
    sheet.add_chart(chart1, "A15")
    
    # Income vs Expenses chart
    chart_title = "Income vs Expenses"
    row = 30
    sheet.merge_cells(f'A{row}:F{row}')
    sheet[f'A{row}'] = chart_title
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = BLUE_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Create income vs expenses chart
    chart2 = LineChart()
    chart2.title = chart_title
    chart2.style = 10
    chart2.y_axis.title = "Amount"
    chart2.x_axis.title = "Month"
    
    # Add hidden cells to the sheet for series names
    sheet['X1'] = "Revenue"
    sheet['X2'] = "Expenses"
    sheet['X3'] = "Cash Balance"
    
    # Create data references for revenue
    revenue_data = Reference(wb["Monthly Projection"], min_col=2, min_row=5, max_row=17, max_col=2)
    expense_data = Reference(wb["Monthly Projection"], min_col=8, min_row=5, max_row=17, max_col=8)
    
    # Add the series with titles from data
    sheet['Y1'] = "='Monthly Projection'!B5"
    sheet['Y2'] = "='Monthly Projection'!H5"
    revenue_title = Reference(sheet, min_col=24, min_row=1, max_row=1)
    expense_title = Reference(sheet, min_col=24, min_row=2, max_row=2)
    
    chart2.add_data(revenue_data, titles_from_data=False)
    chart2.add_data(expense_data, titles_from_data=False)
    chart2.set_categories(months)
    
    # Style the lines
    chart2.series[0].graphicalProperties.line.solidFill = "00B050"  # Green for revenue
    chart2.series[0].graphicalProperties.line.width = 20000  # Width of line
    
    chart2.series[1].graphicalProperties.line.solidFill = "FF0000"  # Red for expenses
    chart2.series[1].graphicalProperties.line.width = 20000
    
    chart2.shape = 4
    sheet.add_chart(chart2, "A31")
    
    # Cash Balance Trend chart
    chart_title = "Cash Balance Trend"
    row = 46
    sheet.merge_cells(f'A{row}:F{row}')
    sheet[f'A{row}'] = chart_title
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = BLUE_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Create cash balance trend chart
    chart3 = LineChart()
    chart3.title = chart_title
    chart3.style = 10
    chart3.y_axis.title = "Balance"
    chart3.x_axis.title = "Month"
    
    # Create data reference for cash balance
    balance_data = Reference(wb["Monthly Projection"], min_col=15, min_row=5, max_row=17, max_col=15)
    
    # Add data series
    chart3.add_data(balance_data, titles_from_data=False)
    chart3.set_categories(months)
    
    # Style the line
    chart3.series[0].graphicalProperties.line.solidFill = "0070C0"  # Blue for cash balance
    chart3.series[0].graphicalProperties.line.width = 30000  # Width of line
    
    chart3.shape = 4
    sheet.add_chart(chart3, "A47")
    
    # Add company branding
    row = 62
    sheet.merge_cells(f'A{row}:F{row}')
    sheet[f'A{row}'] = "Prepared by Clarity Impact Finance"
    sheet[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    row += 1
    sheet.merge_cells(f'A{row}:F{row}')
    sheet[f'A{row}'] = "contact@clarityimpactfinance.com"
    sheet[f'A{row}'].font = Font(name='Arial', size=9)
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')

def setup_input_tab(wb, sheet):
    """Set up the Input tab for entering business information and assumptions"""
    # Set column widths
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 40
    
    # Business Information Section
    sheet.merge_cells('A1:C1')
    sheet['A1'] = "BUSINESS INFORMATION"
    sheet['A1'].font = TITLE_FONT
    sheet['A1'].fill = BLUE_FILL
    sheet['A1'].alignment = Alignment(horizontal='center')
    
    # Input fields for business info
    business_info = [
        ("Business Name", "", "Enter your business name"),
        ("Industry", "", "Select or enter your primary industry"),
        ("Projection Start Date", "", "MM/DD/YYYY format"),
        ("Projection Period", "12", "Number of months to project (default: 12)"),
        ("Starting Cash Balance", 0, "Current cash balance to start projection"),
        ("Currency", "USD", "Default currency for all values")
    ]
    
    row = 3
    for info in business_info:
        label, default, hint = info
        
        sheet[f'A{row}'] = label
        sheet[f'A{row}'].font = NORMAL_FONT
        sheet[f'A{row}'].alignment = Alignment(horizontal='left')
        
        sheet[f'B{row}'] = default
        if label == "Starting Cash Balance":
            sheet[f'B{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        
        sheet[f'C{row}'] = hint
        sheet[f'C{row}'].font = Font(name='Arial', size=9, italic=True)
        sheet[f'C{row}'].alignment = Alignment(horizontal='left')
        
        row += 1
    
    # Revenue Assumptions Section
    row += 2
    sheet.merge_cells(f'A{row}:C{row}')
    sheet[f'A{row}'] = "REVENUE ASSUMPTIONS"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = LIGHT_GREEN_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    row += 1
    sheet[f'A{row}'] = "Revenue Stream"
    sheet[f'B{row}'] = "Monthly Amount"
    sheet[f'C{row}'] = "Growth/Change Assumptions"
    
    for col in ['A', 'B', 'C']:
        sheet[f'{col}{row}'].font = Font(name='Arial', size=10, bold=True)
        sheet[f'{col}{row}'].fill = GREY_FILL
    
    # Revenue streams (5 rows)
    revenue_streams = [
        "Product Sales",
        "Service Revenue",
        "Subscription Income",
        "Other Revenue 1",
        "Other Revenue 2"
    ]
    
    for i, stream in enumerate(revenue_streams):
        row += 1
        sheet[f'A{row}'] = stream
        sheet[f'B{row}'] = 0
        sheet[f'B{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        sheet[f'C{row}'] = ""
    
    # Store the last revenue row for reference
    last_revenue_row = row
    
    # Expense Assumptions Section
    row += 2
    expense_start_row = row
    
    sheet.merge_cells(f'A{row}:C{row}')
    sheet[f'A{row}'] = "EXPENSE ASSUMPTIONS"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = LIGHT_BLUE_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    row += 1
    sheet[f'A{row}'] = "Expense Category"
    sheet[f'B{row}'] = "Monthly Amount"
    sheet[f'C{row}'] = "Growth/Change Assumptions"
    
    for col in ['A', 'B', 'C']:
        sheet[f'{col}{row}'].font = Font(name='Arial', size=10, bold=True)
        sheet[f'{col}{row}'].fill = GREY_FILL
    
    # Expense categories
    expense_categories = [
        "Salaries & Wages",
        "Rent/Mortgage",
        "Utilities",
        "Insurance",
        "Supplies",
        "Marketing & Advertising",
        "Professional Services",
        "Equipment & Maintenance",
        "Loan Payments",
        "Taxes",
        "Other Expenses"
    ]
    
    for i, category in enumerate(expense_categories):
        row += 1
        sheet[f'A{row}'] = category
        sheet[f'B{row}'] = 0
        sheet[f'B{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        sheet[f'C{row}'] = ""
    
    # Capital Expenditures Section
    row += 2
    capex_start_row = row
    
    sheet.merge_cells(f'A{row}:C{row}')
    sheet[f'A{row}'] = "CAPITAL EXPENDITURES & INVESTMENTS"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = ORANGE_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    row += 1
    sheet[f'A{row}'] = "Expenditure Description"
    sheet[f'B{row}'] = "Amount"
    sheet[f'C{row}'] = "Expected Month"
    
    for col in ['A', 'B', 'C']:
        sheet[f'{col}{row}'].font = Font(name='Arial', size=10, bold=True)
        sheet[f'{col}{row}'].fill = GREY_FILL
    
    # Capital expenditure rows (3 rows)
    for i in range(3):
        row += 1
        sheet[f'A{row}'] = ""
        sheet[f'B{row}'] = 0
        sheet[f'B{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        sheet[f'C{row}'] = ""
    
    # Funding & Financing Section
    row += 2
    funding_start_row = row
    
    sheet.merge_cells(f'A{row}:C{row}')
    sheet[f'A{row}'] = "FUNDING & FINANCING"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = GREEN_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    row += 1
    sheet[f'A{row}'] = "Source/Description"
    sheet[f'B{row}'] = "Amount"
    sheet[f'C{row}'] = "Expected Month"
    
    for col in ['A', 'B', 'C']:
        sheet[f'{col}{row}'].font = Font(name='Arial', size=10, bold=True)
        sheet[f'{col}{row}'].fill = GREY_FILL
    
    # Funding sources (3 rows)
    funding_sources = [
        "Bank Loan",
        "Investor Funding",
        "Business Line of Credit"
    ]
    
    for i, source in enumerate(funding_sources):
        row += 1
        sheet[f'A{row}'] = source
        sheet[f'B{row}'] = 0
        sheet[f'B{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        sheet[f'C{row}'] = ""
    
    # Add borders to all filled cells
    max_row = row
    for row in range(1, max_row + 1):
        for col in range(1, 4):  # Columns A-C
            cell = sheet.cell(row=row, column=col)
            cell.border = THIN_BORDER
    
    # Add note about how data will be used
    row += 2
    sheet.merge_cells(f'A{row}:C{row}')
    sheet[f'A{row}'] = "Note: Data entered on this sheet will automatically populate the monthly projection worksheet."
    sheet[f'A{row}'].font = Font(name='Arial', size=9, italic=True)
    sheet[f'A{row}'].alignment = Alignment(horizontal='left')
    
    # Add company branding
    row += 2
    sheet.merge_cells(f'A{row}:C{row}')
    sheet[f'A{row}'] = "Prepared by Clarity Impact Finance"
    sheet[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    row += 1
    sheet.merge_cells(f'A{row}:C{row}')
    sheet[f'A{row}'] = "contact@clarityimpactfinance.com"
    sheet[f'A{row}'].font = Font(name='Arial', size=9)
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')

def setup_monthly_projection(wb, sheet):
    """Set up the Monthly Projection worksheet with a detailed cash flow projection"""
    # Set column widths
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 20
    for col in range(3, 15):  # Columns C through N (12 months)
        sheet.column_dimensions[get_column_letter(col)].width = 15
    
    # Projection Title
    sheet.merge_cells('A1:N1')
    sheet['A1'] = "MONTHLY CASH FLOW PROJECTION"
    sheet['A1'].font = TITLE_FONT
    sheet['A1'].fill = BLUE_FILL
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Business info line
    sheet.merge_cells('A2:N2')
    sheet['A2'] = "=CONCATENATE(\"Business: \", Input!B3, \" - \", \"Industry: \", Input!B4)"
    sheet['A2'].font = Font(name='Arial', size=10, italic=True)
    sheet['A2'].alignment = Alignment(horizontal='center')
    
    # Month headers
    sheet['B4'] = "CATEGORY"
    sheet['B4'].font = HEADER_FONT
    sheet['B4'].fill = GREY_FILL
    sheet['B4'].alignment = Alignment(horizontal='center')
    
    # Generate month headers using formulas based on start date
    for i in range(12):
        col_letter = get_column_letter(i + 3)  # Start at column C
        # Formula to calculate month based on start date
        sheet[f'{col_letter}4'] = f'=TEXT(EDATE(Input!B5,{i}), "mmm-yy")'
        sheet[f'{col_letter}4'].font = HEADER_FONT
        sheet[f'{col_letter}4'].fill = GREY_FILL
        sheet[f'{col_letter}4'].alignment = Alignment(horizontal='center')
    
    # Cash Inflows Section
    row = 6
    sheet.merge_cells(f'A{row}:N{row}')
    sheet[f'A{row}'] = "CASH INFLOWS"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = LIGHT_GREEN_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Category label for cash inflows summary
    row += 1
    sheet[f'A{row}'] = "Total Cash Inflows"
    sheet[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
    sheet[f'B{row}'].font = Font(name='Arial', size=10, bold=True)
    
    # Formulas for Total Cash Inflows will be added later
    
    # Revenue categories
    row += 2
    sheet[f'A{row}'] = "REVENUE"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = GREY_FILL
    
    # Revenue streams from input tab
    revenue_rows = []
    revenue_streams = [
        "Product Sales",
        "Service Revenue", 
        "Subscription Income",
        "Other Revenue 1",
        "Other Revenue 2"
    ]
    
    for stream in revenue_streams:
        row += 1
        revenue_rows.append(row)
        sheet[f'A{row}'] = stream
        sheet[f'B{row}'] = "Monthly"
        
        # Add formula for each month (copy from input with growth assumptions)
        for i in range(12):
            col_letter = get_column_letter(i + 3)  # Start at column C
            
            # Look up the value from Input tab
            input_row = None
            for r in range(11, 17):  # Approximate range where revenue streams are defined
                if sheet.cell(row=r, column=1).value == stream:
                    input_row = r
                    break
            
            if input_row:
                # Basic formula that can be enhanced with growth logic
                sheet[f'{col_letter}{row}'] = f'=Input!B{input_row}'
                sheet[f'{col_letter}{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    # Funding & Financing
    row += 2
    sheet[f'A{row}'] = "FUNDING & FINANCING"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = GREY_FILL
    
    # Funding sources from input tab
    funding_rows = []
    funding_sources = [
        "Bank Loan",
        "Investor Funding",
        "Business Line of Credit"
    ]
    
    for source in funding_sources:
        row += 1
        funding_rows.append(row)
        sheet[f'A{row}'] = source
        sheet[f'B{row}'] = "One-time"
        
        # For each funding source, amount goes in specified month only
        for i in range(12):
            col_letter = get_column_letter(i + 3)  # Start at column C
            
            # Find corresponding input row
            input_row = None
            for r in range(30, 40):  # Approximate range for funding sources
                if wb["Input"].cell(row=r, column=1).value == source:
                    input_row = r
                    break
            
            if input_row:
                # Only put the amount in the specified month
                month_cell = wb["Input"].cell(row=input_row, column=3).value
                if month_cell and isinstance(month_cell, int) and month_cell == i + 1:
                    sheet[f'{col_letter}{row}'] = f'=Input!B{input_row}'
                else:
                    sheet[f'{col_letter}{row}'] = 0
                
                sheet[f'{col_letter}{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    # Other Inflows
    row += 2
    sheet[f'A{row}'] = "OTHER INFLOWS"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = GREY_FILL
    
    # Add a couple of blank rows for other inflows
    other_inflow_rows = []
    for i in range(2):
        row += 1
        other_inflow_rows.append(row)
        sheet[f'A{row}'] = f"Other Inflow {i+1}"
        sheet[f'B{row}'] = ""
        
        for j in range(12):
            col_letter = get_column_letter(j + 3)
            sheet[f'{col_letter}{row}'] = 0
            sheet[f'{col_letter}{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    # Now add formulas for Total Cash Inflows
    first_revenue_row = revenue_rows[0]
    last_other_inflow_row = other_inflow_rows[-1]
    
    # Go back to the Total Cash Inflows row
    inflow_total_row = 7
    for i in range(12):
        col_letter = get_column_letter(i + 3)
        # Sum all inflows
        sheet[f'{col_letter}{inflow_total_row}'] = f'=SUM({col_letter}{first_revenue_row}:{col_letter}{last_other_inflow_row})'
        sheet[f'{col_letter}{inflow_total_row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        sheet[f'{col_letter}{inflow_total_row}'].font = Font(name='Arial', size=10, bold=True)
    
    # Cash Outflows Section
    row = last_other_inflow_row + 2
    outflow_section_start = row
    
    sheet.merge_cells(f'A{row}:N{row}')
    sheet[f'A{row}'] = "CASH OUTFLOWS"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = LIGHT_BLUE_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Category label for cash outflows summary
    row += 1
    outflow_total_row = row
    sheet[f'A{row}'] = "Total Cash Outflows"
    sheet[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
    sheet[f'B{row}'].font = Font(name='Arial', size=10, bold=True)
    
    # Operating Expenses
    row += 2
    sheet[f'A{row}'] = "OPERATING EXPENSES"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = GREY_FILL
    
    # Expense categories from input tab
    expense_rows = []
    expense_categories = [
        "Salaries & Wages",
        "Rent/Mortgage",
        "Utilities",
        "Insurance",
        "Supplies",
        "Marketing & Advertising",
        "Professional Services",
        "Equipment & Maintenance",
        "Loan Payments",
        "Taxes",
        "Other Expenses"
    ]
    
    for category in expense_categories:
        row += 1
        expense_rows.append(row)
        sheet[f'A{row}'] = category
        sheet[f'B{row}'] = "Monthly"
        
        # Add formula for each month
        for i in range(12):
            col_letter = get_column_letter(i + 3)  # Start at column C
            
            # Find corresponding input row
            input_row = None
            for r in range(17, 30):  # Approximate range for expense categories
                if wb["Input"].cell(row=r, column=1).value == category:
                    input_row = r
                    break
            
            if input_row:
                # Basic formula that can be enhanced with growth logic
                sheet[f'{col_letter}{row}'] = f'=Input!B{input_row}'
                sheet[f'{col_letter}{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    # Capital Expenditures
    row += 2
    sheet[f'A{row}'] = "CAPITAL EXPENDITURES"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = GREY_FILL
    
    # Capital expenditure rows from input
    capex_rows = []
    for i in range(3):  # Assuming 3 capex rows in input
        row += 1
        capex_rows.append(row)
        sheet[f'A{row}'] = f"Capital Expenditure {i+1}"
        sheet[f'B{row}'] = "One-time"
        
        # For each capex, amount goes in specified month only
        for j in range(12):
            col_letter = get_column_letter(j + 3)
            
            # Find corresponding input row (since we don't have exact label matches)
            input_row = 24 + i  # Approximate position in input
            
            # Only put the amount in the specified month
            month_cell = wb["Input"].cell(row=input_row, column=3).value
            if month_cell and isinstance(month_cell, int) and month_cell == j + 1:
                sheet[f'{col_letter}{row}'] = f'=Input!B{input_row}'
            else:
                sheet[f'{col_letter}{row}'] = 0
            
            sheet[f'{col_letter}{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    # Now add formulas for Total Cash Outflows
    first_expense_row = expense_rows[0]
    last_capex_row = capex_rows[-1]
    
    # Go back to the Total Cash Outflows row
    for i in range(12):
        col_letter = get_column_letter(i + 3)
        # Sum all outflows
        sheet[f'{col_letter}{outflow_total_row}'] = f'=SUM({col_letter}{first_expense_row}:{col_letter}{last_capex_row})'
        sheet[f'{col_letter}{outflow_total_row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        sheet[f'{col_letter}{outflow_total_row}'].font = Font(name='Arial', size=10, bold=True)
    
    # Net Cash Flow Section
    row = last_capex_row + 2
    
    sheet.merge_cells(f'A{row}:N{row}')
    sheet[f'A{row}'] = "CASH FLOW SUMMARY"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = CIF_GREEN
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Net Monthly Cash Flow
    row += 1
    net_flow_row = row
    sheet[f'A{row}'] = "Net Monthly Cash Flow"
    sheet[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
    sheet[f'B{row}'].font = Font(name='Arial', size=10, bold=True)
    
    for i in range(12):
        col_letter = get_column_letter(i + 3)
        # Inflows minus outflows
        sheet[f'{col_letter}{row}'] = f'={col_letter}{inflow_total_row}-{col_letter}{outflow_total_row}'
        sheet[f'{col_letter}{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        
        # Conditional formatting for negative cash flow
        red_font = Font(color="FF0000", bold=True)
        green_font = Font(color="00A651", bold=True)
        
        rule = Rule(type="cellIs", operator="lessThan", formula=["0"], dxf=DifferentialStyle(font=red_font))
        sheet.conditional_formatting.add(f'{col_letter}{row}', rule)
        
        rule = Rule(type="cellIs", operator="greaterThanOrEqual", formula=["0"], dxf=DifferentialStyle(font=green_font))
        sheet.conditional_formatting.add(f'{col_letter}{row}', rule)
    
    # Running Cash Balance
    row += 1
    sheet[f'A{row}'] = "Running Cash Balance"
    sheet[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
    sheet[f'B{row}'] = "Cumulative"
    sheet[f'B{row}'].font = Font(name='Arial', size=10, bold=True)
    
    # First month is starting balance plus first month's net flow
    sheet[f'C{row}'] = f'=Input!B7+C{net_flow_row}'
    sheet[f'C{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    # Subsequent months add previous balance to current month's net flow
    for i in range(1, 12):
        prev_col = get_column_letter(i + 2)
        curr_col = get_column_letter(i + 3)
        sheet[f'{curr_col}{row}'] = f'={prev_col}{row}+{curr_col}{net_flow_row}'
        sheet[f'{curr_col}{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        
        # Add conditional formatting for negative balance
        red_font = Font(color="FF0000", bold=True)
        green_font = Font(color="00A651", bold=True)
        
        rule = Rule(type="cellIs", operator="lessThan", formula=["0"], dxf=DifferentialStyle(font=red_font))
        sheet.conditional_formatting.add(f'{curr_col}{row}', rule)
        
        rule = Rule(type="cellIs", operator="greaterThanOrEqual", formula=["0"], dxf=DifferentialStyle(font=green_font))
        sheet.conditional_formatting.add(f'{curr_col}{row}', rule)
    
    # Add borders to all filled cells
    max_row = row
    for row in range(1, max_row + 1):
        for col in range(1, 15):  # Columns A-N
            cell = sheet.cell(row=row, column=col)
            cell.border = THIN_BORDER
    
    # Add notes and warnings
    row += 2
    sheet.merge_cells(f'A{row}:N{row}')
    sheet[f'A{row}'] = "Note: Negative cash flow or balance is highlighted in red. Review your assumptions or adjust your business plan accordingly."
    sheet[f'A{row}'].font = Font(name='Arial', size=9, italic=True)
    sheet[f'A{row}'].alignment = Alignment(horizontal='left')
    
    # Add company branding
    row += 2
    sheet.merge_cells(f'A{row}:N{row}')
    sheet[f'A{row}'] = "Prepared by Clarity Impact Finance"
    sheet[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    row += 1
    sheet.merge_cells(f'A{row}:N{row}')
    sheet[f'A{row}'] = "contact@clarityimpactfinance.com"
    sheet[f'A{row}'].font = Font(name='Arial', size=9)
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')

def setup_annual_summary(wb, sheet):
    """Set up the Annual Summary worksheet with yearly totals and key metrics"""
    # Set column widths
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    
    # Page Title
    sheet.merge_cells('A1:C1')
    sheet['A1'] = "ANNUAL CASH FLOW SUMMARY"
    sheet['A1'].font = TITLE_FONT
    sheet['A1'].fill = BLUE_FILL
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Business info line
    sheet.merge_cells('A2:C2')
    sheet['A2'] = "=CONCATENATE(\"Business: \", Input!B3, \" - \", \"Industry: \", Input!B4)"
    sheet['A2'].font = Font(name='Arial', size=10, italic=True)
    sheet['A2'].alignment = Alignment(horizontal='center')
    
    # Date range
    sheet.merge_cells('A3:C3')
    sheet['A3'] = "=CONCATENATE(\"Period: \", TEXT(Input!B5, \"mmm yyyy\"), \" to \", TEXT(EDATE(Input!B5, 11), \"mmm yyyy\"))"
    sheet['A3'].font = Font(name='Arial', size=10, italic=True)
    sheet['A3'].alignment = Alignment(horizontal='center')
    
    # Key Financial Metrics Section
    row = 5
    sheet.merge_cells(f'A{row}:C{row}')
    sheet[f'A{row}'] = "KEY FINANCIAL METRICS"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = CIF_GREEN
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Headers
    row += 1
    sheet[f'A{row}'] = "Metric"
    sheet[f'B{row}'] = "Amount"
    sheet[f'C{row}'] = "Notes"
    
    for col in ['A', 'B', 'C']:
        sheet[f'{col}{row}'].font = Font(name='Arial', size=10, bold=True)
        sheet[f'{col}{row}'].fill = GREY_FILL
    
    # Key metrics
    metrics = [
        ("Starting Cash Balance", "=Input!B7", "Beginning cash on hand"),
        ("Ending Cash Balance", "='Monthly Projection'!N47", "Projected final cash position"),
        ("Total Cash Inflows", "=SUM('Monthly Projection'!C7:N7)", "Sum of all revenue and funding"),
        ("Total Cash Outflows", "=SUM('Monthly Projection'!C32:N32)", "Sum of all expenses and capital expenditures"),
        ("Net Annual Cash Flow", "=B9-B7", "Change in cash position over the period"),
        ("Average Monthly Cash Flow", "=B11/12", "Average monthly net cash flow"),
        ("Months with Positive Cash Flow", "=COUNTIF('Monthly Projection'!C46:N46,\">0\")", "Number of months with positive net flow"),
        ("Months with Negative Cash Flow", "=COUNTIF('Monthly Projection'!C46:N46,\"<0\")", "Number of months with negative net flow"),
        ("Lowest Monthly Cash Balance", "=MIN('Monthly Projection'!C47:N47)", "Lowest point in cash reserves"),
        ("Peak Monthly Cash Balance", "=MAX('Monthly Projection'!C47:N47)", "Highest point in cash reserves")
    ]
    
    for metric in metrics:
        row += 1
        title, formula, note = metric
        sheet[f'A{row}'] = title
        sheet[f'A{row}'].font = NORMAL_FONT
        sheet[f'A{row}'].alignment = Alignment(horizontal='left')
        
        sheet[f'B{row}'] = formula
        sheet[f'B{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        
        sheet[f'C{row}'] = note
        sheet[f'C{row}'].font = Font(name='Arial', size=9, italic=True)
    
    # Cash Flow Breakdown Section
    row += 3
    sheet.merge_cells(f'A{row}:C{row}')
    sheet[f'A{row}'] = "CASH FLOW BREAKDOWN"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = LIGHT_GREEN_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Income Breakdown
    row += 1
    sheet[f'A{row}'] = "INCOME SUMMARY"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = GREY_FILL
    
    # Headers
    row += 1
    sheet[f'A{row}'] = "Revenue Stream"
    sheet[f'B{row}'] = "Annual Total"
    sheet[f'C{row}'] = "% of Revenue"
    
    for col in ['A', 'B', 'C']:
        sheet[f'{col}{row}'].font = Font(name='Arial', size=10, bold=True)
    
    # Revenue streams
    revenue_streams = [
        "Product Sales",
        "Service Revenue", 
        "Subscription Income",
        "Other Revenue 1",
        "Other Revenue 2"
    ]
    
    first_revenue_row = row + 1
    for stream in revenue_streams:
        row += 1
        sheet[f'A{row}'] = stream
        
        # Find the corresponding rows in the monthly projection
        for proj_row in range(10, 20):
            if wb["Monthly Projection"].cell(row=proj_row, column=1).value == stream:
                sheet[f'B{row}'] = f'=SUM(\'Monthly Projection\'!C{proj_row}:N{proj_row})'
                break
        
        sheet[f'B{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    # Total Revenue row
    last_revenue_row = row
    row += 1
    sheet[f'A{row}'] = "Total Revenue"
    sheet[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
    sheet[f'B{row}'] = f'=SUM(B{first_revenue_row}:B{last_revenue_row})'
    sheet[f'B{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    sheet[f'B{row}'].font = Font(name='Arial', size=10, bold=True)
    
    # Calculate percentages of revenue
    revenue_total_row = row
    for r in range(first_revenue_row, last_revenue_row + 1):
        sheet[f'C{r}'] = f'=IF(${revenue_total_row}=0,0,B{r}/B${revenue_total_row})'
        sheet[f'C{r}'].number_format = '0.00%'
    
    # Expense Breakdown
    row += 2
    sheet[f'A{row}'] = "EXPENSE SUMMARY"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = GREY_FILL
    
    # Headers
    row += 1
    sheet[f'A{row}'] = "Expense Category"
    sheet[f'B{row}'] = "Annual Total"
    sheet[f'C{row}'] = "% of Expenses"
    
    for col in ['A', 'B', 'C']:
        sheet[f'{col}{row}'].font = Font(name='Arial', size=10, bold=True)
    
    # Expense categories
    expense_categories = [
        "Salaries & Wages",
        "Rent/Mortgage",
        "Utilities",
        "Insurance",
        "Supplies",
        "Marketing & Advertising",
        "Professional Services",
        "Equipment & Maintenance",
        "Loan Payments",
        "Taxes",
        "Other Expenses"
    ]
    
    first_expense_row = row + 1
    for category in expense_categories:
        row += 1
        sheet[f'A{row}'] = category
        
        # Find the corresponding rows in the monthly projection
        for proj_row in range(30, 45):
            if wb["Monthly Projection"].cell(row=proj_row, column=1).value == category:
                sheet[f'B{row}'] = f'=SUM(\'Monthly Projection\'!C{proj_row}:N{proj_row})'
                break
        
        sheet[f'B{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    # Total Expenses row
    last_expense_row = row
    row += 1
    sheet[f'A{row}'] = "Total Expenses"
    sheet[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
    sheet[f'B{row}'] = f'=SUM(B{first_expense_row}:B{last_expense_row})'
    sheet[f'B{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    sheet[f'B{row}'].font = Font(name='Arial', size=10, bold=True)
    
    # Calculate percentages of expenses
    expense_total_row = row
    for r in range(first_expense_row, last_expense_row + 1):
        sheet[f'C{r}'] = f'=IF(${expense_total_row}=0,0,B{r}/B${expense_total_row})'
        sheet[f'C{r}'].number_format = '0.00%'
    
    # Cash Flow Summary Section
    row += 2
    sheet.merge_cells(f'A{row}:C{row}')
    sheet[f'A{row}'] = "CASH FLOW RATIO ANALYSIS"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = LIGHT_BLUE_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Headers
    row += 1
    sheet[f'A{row}'] = "Ratio"
    sheet[f'B{row}'] = "Value"
    sheet[f'C{row}'] = "Interpretation"
    
    for col in ['A', 'B', 'C']:
        sheet[f'{col}{row}'].font = Font(name='Arial', size=10, bold=True)
        sheet[f'{col}{row}'].fill = GREY_FILL
    
    # Cash flow ratios
    ratios = [
        ("Revenue to Expense Ratio", f"=B{revenue_total_row}/B{expense_total_row}", 
         "=IF(B{row}<1,\"Spending exceeds income\",\"Income covers expenses\")"),
        ("Operating Cash Flow Margin", f"=(B{revenue_total_row}-B{expense_total_row})/B{revenue_total_row}", 
         "=IF(B{row}<0.1,\"Low margin - review costs\",\"Healthy operating margin\")"),
        ("Cash Flow to Debt Ratio", "=IF(SUM('Monthly Projection'!C42:N42)=0,\"N/A\",B11/SUM('Monthly Projection'!C42:N42))", 
         "=IF(B{row}<1,\"May struggle with debt service\",\"Can service debt comfortably\")")
    ]
    
    for ratio in ratios:
        row += 1
        title, formula, interpretation = ratio
        sheet[f'A{row}'] = title
        sheet[f'A{row}'].font = NORMAL_FONT
        
        sheet[f'B{row}'] = formula.format(row=row+1)  # Adjust row reference for interpretation formula
        if "Margin" in title or "Ratio" in title and "N/A" not in formula:
            sheet[f'B{row}'].number_format = '0.00'
        
        # Replace {row} with actual row number in interpretation formula
        sheet[f'C{row}'] = interpretation.format(row=row)
        sheet[f'C{row}'].font = Font(name='Arial', size=9, italic=True)
    
    # Add borders to all filled cells
    max_row = row
    for row in range(1, max_row + 1):
        for col in range(1, 4):  # Columns A-C
            cell = sheet.cell(row=row, column=col)
            cell.border = THIN_BORDER
    
    # Add notes section
    row += 2
    sheet.merge_cells(f'A{row}:C{row}')
    sheet[f'A{row}'] = "NOTES AND RECOMMENDATIONS"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = ORANGE_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Add some automatic recommendations based on the cash flow
    row += 1
    sheet.merge_cells(f'A{row}:C{row}')
    sheet[f'A{row}'] = "=IF(B9<0,\"WARNING: Your annual cash flow is negative. Consider reducing expenses or increasing revenue sources.\",\"Your annual cash flow is positive.\")"
    
    row += 1
    sheet.merge_cells(f'A{row}:C{row}')
    sheet[f'A{row}'] = "=IF(B13<6,\"Your business has negative cash flow in multiple months. Review your monthly projection for problem areas.\",\"Your business maintains positive cash flow in most months.\")"
    
    row += 1
    sheet.merge_cells(f'A{row}:C{row}')
    sheet[f'A{row}'] = "=IF(B17<B7*0.5,\"Your lowest cash balance drops significantly below your starting position. Consider maintaining higher reserves or restructuring expenses.\",\"Your cash reserves remain at healthy levels throughout the projection period.\")"
    
    for r in range(row-2, row+1):
        sheet[f'A{r}'].font = Font(name='Arial', size=10)
        sheet[f'A{r}'].alignment = Alignment(horizontal='left', wrap_text=True)
    
    # Add company branding
    row += 2
    sheet.merge_cells(f'A{row}:C{row}')
    sheet[f'A{row}'] = "Prepared by Clarity Impact Finance"
    sheet[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    row += 1
    sheet.merge_cells(f'A{row}:C{row}')
    sheet[f'A{row}'] = "contact@clarityimpactfinance.com"
    sheet[f'A{row}'].font = Font(name='Arial', size=9)
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')

def setup_assumptions(wb, sheet):
    """Set up the Assumptions worksheet with detailed explanations about financial projections"""
    # Set column widths
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 60
    
    # Page Title
    sheet.merge_cells('A1:B1')
    sheet['A1'] = "CASH FLOW PROJECTION ASSUMPTIONS"
    sheet['A1'].font = TITLE_FONT
    sheet['A1'].fill = BLUE_FILL
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Introduction text
    sheet.merge_cells('A2:B3')
    sheet['A2'] = "This worksheet explains the key assumptions underlying the cash flow projections. Understanding these assumptions is critical for accurate financial planning."
    sheet['A2'].font = Font(name='Arial', size=10)
    sheet['A2'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Revenue Assumptions
    row = 5
    sheet.merge_cells(f'A{row}:B{row}')
    sheet[f'A{row}'] = "REVENUE ASSUMPTIONS"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = LIGHT_GREEN_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Headers
    row += 1
    sheet[f'A{row}'] = "Category"
    sheet[f'B{row}'] = "Explanation"
    
    for col in ['A', 'B']:
        sheet[f'{col}{row}'].font = Font(name='Arial', size=10, bold=True)
        sheet[f'{col}{row}'].fill = GREY_FILL
    
    # Revenue assumptions
    revenue_assumptions = [
        ("Growth Rate", "The monthly growth rate applied to revenue streams. Enter growth assumptions in the Input tab, such as '5% monthly' or 'seasonal: +10% in Q4'."),
        ("Seasonality", "If your business experiences seasonal fluctuations, note this in the growth assumptions for each revenue stream."),
        ("New Products/Services", "For new offerings, consider a ramp-up period rather than immediate full revenue."),
        ("Customer Acquisition", "Consider how new customer acquisition affects revenue growth and whether there are delays between marketing efforts and revenue."),
        ("Payment Timing", "Account for the timing of payments. If customers typically pay 30 days after service, adjust your cash flow accordingly.")
    ]
    
    for assumption in revenue_assumptions:
        row += 1
        category, explanation = assumption
        sheet[f'A{row}'] = category
        sheet[f'A{row}'].font = NORMAL_FONT
        sheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        sheet[f'B{row}'] = explanation
        sheet[f'B{row}'].font = NORMAL_FONT
        sheet[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Expense Assumptions
    row += 2
    sheet.merge_cells(f'A{row}:B{row}')
    sheet[f'A{row}'] = "EXPENSE ASSUMPTIONS"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = LIGHT_BLUE_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Headers
    row += 1
    sheet[f'A{row}'] = "Category"
    sheet[f'B{row}'] = "Explanation"
    
    for col in ['A', 'B']:
        sheet[f'{col}{row}'].font = Font(name='Arial', size=10, bold=True)
        sheet[f'{col}{row}'].fill = GREY_FILL
    
    # Expense assumptions
    expense_assumptions = [
        ("Fixed vs. Variable Costs", "Identify which expenses are fixed (rent, insurance) versus variable (commissions, material costs). Variable costs should scale with revenue."),
        ("Inflation", "Consider the impact of inflation on expenses, typically 2-3% annually for stable economies."),
        ("Staffing Plans", "Account for planned hires, raises, and bonuses in your salary projections."),
        ("Capital Expenditures", "Major purchases should be placed in the month they will occur, not spread across months."),
        ("Tax Payments", "Include quarterly or annual tax payments in the appropriate months."),
        ("Debt Service", "Ensure loan payments reflect the correct amortization schedule."),
        ("One-time Expenses", "Identify and place one-time expenses in the correct month rather than averaging across the year.")
    ]
    
    for assumption in expense_assumptions:
        row += 1
        category, explanation = assumption
        sheet[f'A{row}'] = category
        sheet[f'A{row}'].font = NORMAL_FONT
        sheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        sheet[f'B{row}'] = explanation
        sheet[f'B{row}'].font = NORMAL_FONT
        sheet[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Cash Flow Timing Assumptions
    row += 2
    sheet.merge_cells(f'A{row}:B{row}')
    sheet[f'A{row}'] = "CASH FLOW TIMING ASSUMPTIONS"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = ORANGE_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Headers
    row += 1
    sheet[f'A{row}'] = "Category"
    sheet[f'B{row}'] = "Explanation"
    
    for col in ['A', 'B']:
        sheet[f'{col}{row}'].font = Font(name='Arial', size=10, bold=True)
        sheet[f'{col}{row}'].fill = GREY_FILL
    
    # Timing assumptions
    timing_assumptions = [
        ("Accounts Receivable", "If you offer credit terms to customers, factor in delayed receipt of payments (typically 30-90 days)."),
        ("Accounts Payable", "If you pay suppliers on delayed terms, factor in the timing of these outflows."),
        ("Inventory Purchases", "Cash outflows for inventory typically occur before the corresponding sales revenue."),
        ("Funding Timing", "Loans or investments should be placed in the specific month they are expected to be received."),
        ("Cash Reserves", "Consider how much cash buffer you need to maintain for unexpected expenses or opportunities.")
    ]
    
    for assumption in timing_assumptions:
        row += 1
        category, explanation = assumption
        sheet[f'A{row}'] = category
        sheet[f'A{row}'].font = NORMAL_FONT
        sheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        sheet[f'B{row}'] = explanation
        sheet[f'B{row}'].font = NORMAL_FONT
        sheet[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Sensitivity Considerations
    row += 2
    sheet.merge_cells(f'A{row}:B{row}')
    sheet[f'A{row}'] = "SENSITIVITY CONSIDERATIONS"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = GREEN_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Sensitivity analysis text
    row += 1
    sheet.merge_cells(f'A{row}:B{row+3}')
    sheet[f'A{row}'] = (
        "Consider how changes to key assumptions might impact your cash flow projections. Best practice is to create "
        "multiple scenarios:\n\n"
        "1. Base Case: Your most likely projection based on reasonable assumptions\n"
        "2. Conservative Case: Reduced revenue growth and/or increased expenses\n"
        "3. Optimistic Case: Higher revenue growth and/or better cost control\n\n"
        "This allows you to understand the range of possible outcomes and prepare contingency plans."
    )
    sheet[f'A{row}'].font = NORMAL_FONT
    sheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Add borders to all filled cells
    max_row = row + 3
    for row in range(1, max_row + 1):
        for col in range(1, 3):  # Columns A-B
            cell = sheet.cell(row=row, column=col)
            cell.border = THIN_BORDER
    
    # Add company branding
    row = max_row + 2
    sheet.merge_cells(f'A{row}:B{row}')
    sheet[f'A{row}'] = "Prepared by Clarity Impact Finance"
    sheet[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    row += 1
    sheet.merge_cells(f'A{row}:B{row}')
    sheet[f'A{row}'] = "contact@clarityimpactfinance.com"
    sheet[f'A{row}'].font = Font(name='Arial', size=9)
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')

def setup_guidance(wb, sheet):
    """Set up the Guidance worksheet with step-by-step instructions for using the template"""
    # Set column widths
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 60
    
    # Page Title
    sheet.merge_cells('A1:B1')
    sheet['A1'] = "CASH FLOW PROJECTION GUIDANCE"
    sheet['A1'].font = TITLE_FONT
    sheet['A1'].fill = BLUE_FILL
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Introduction text
    sheet.merge_cells('A2:B3')
    sheet['A2'] = "This worksheet provides step-by-step guidance on how to use this cash flow projection template effectively. Follow these instructions to create accurate and useful cash flow projections for your business."
    sheet['A2'].font = Font(name='Arial', size=10)
    sheet['A2'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Getting Started Section
    row = 5
    sheet.merge_cells(f'A{row}:B{row}')
    sheet[f'A{row}'] = "GETTING STARTED"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = CIF_GREEN
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Headers
    row += 1
    sheet[f'A{row}'] = "Step"
    sheet[f'B{row}'] = "Instructions"
    
    for col in ['A', 'B']:
        sheet[f'{col}{row}'].font = Font(name='Arial', size=10, bold=True)
        sheet[f'{col}{row}'].fill = GREY_FILL
    
    # Getting started steps
    getting_started = [
        ("1. Navigate the Template", "This template contains multiple worksheets: Dashboard, Input, Monthly Projection, Annual Summary, Assumptions, and Guidance (this sheet). Use the tabs at the bottom to navigate between sheets."),
        ("2. Enter Business Information", "Begin by entering your business information and starting cash balance in the Input tab. This includes your business name, industry, projection start date, and projection period."),
        ("3. Input Revenue Streams", "In the Input tab, enter your expected monthly revenue for each revenue stream. Add notes about growth or seasonality in the Growth/Change Assumptions column."),
        ("4. Input Expenses", "Enter your monthly expenses for each category in the Input tab. Be as accurate as possible, consulting historical data if available."),
        ("5. Enter Capital Expenditures", "If you plan to make major purchases, enter them in the Capital Expenditures section. Specify which month (1-12) they will occur in."),
        ("6. Enter Funding Sources", "If you expect loans, investments, or other funding, enter these in the Funding & Financing section. Specify which month (1-12) they will be received.")
    ]
    
    for step in getting_started:
        row += 1
        step_num, instructions = step
        sheet[f'A{row}'] = step_num
        sheet[f'A{row}'].font = NORMAL_FONT
        sheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        sheet[f'B{row}'] = instructions
        sheet[f'B{row}'].font = NORMAL_FONT
        sheet[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Analyzing Results Section
    row += 2
    sheet.merge_cells(f'A{row}:B{row}')
    sheet[f'A{row}'] = "ANALYZING YOUR PROJECTIONS"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = LIGHT_BLUE_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Headers
    row += 1
    sheet[f'A{row}'] = "Area of Focus"
    sheet[f'B{row}'] = "What to Look For"
    
    for col in ['A', 'B']:
        sheet[f'{col}{row}'].font = Font(name='Arial', size=10, bold=True)
        sheet[f'{col}{row}'].fill = GREY_FILL
    
    # Analysis guidance
    analysis_guidance = [
        ("Monthly Cash Flow", "Review the Monthly Projection tab to identify months with negative cash flow. These are highlighted in red and represent periods where expenses exceed income."),
        ("Cash Balance Trend", "Observe how your cash balance changes over time. A consistently declining balance indicates an unsustainable situation."),
        ("Key Metrics", "Check the Key Financial Metrics in the Annual Summary tab to understand your overall cash position, including total inflows, outflows, and net cash flow."),
        ("Revenue and Expense Breakdown", "Review the percentage breakdown of your revenue streams and expenses in the Annual Summary tab to identify your main sources of income and largest expense categories."),
        ("Ratio Analysis", "Examine the Cash Flow Ratio Analysis in the Annual Summary tab for insights on your business's financial health."),
        ("Dashboard Visuals", "Use the Dashboard tab for a visual overview of your cash flow, including charts that show trends and potential issues at a glance.")
    ]
    
    for guidance in analysis_guidance:
        row += 1
        area, instructions = guidance
        sheet[f'A{row}'] = area
        sheet[f'A{row}'].font = NORMAL_FONT
        sheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        sheet[f'B{row}'] = instructions
        sheet[f'B{row}'].font = NORMAL_FONT
        sheet[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Common Issues Section
    row += 2
    sheet.merge_cells(f'A{row}:B{row}')
    sheet[f'A{row}'] = "COMMON ISSUES AND SOLUTIONS"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = ORANGE_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Headers
    row += 1
    sheet[f'A{row}'] = "Issue"
    sheet[f'B{row}'] = "Potential Solution"
    
    for col in ['A', 'B']:
        sheet[f'{col}{row}'].font = Font(name='Arial', size=10, bold=True)
        sheet[f'{col}{row}'].fill = GREY_FILL
    
    # Common issues and solutions
    issues_solutions = [
        ("Negative Cash Flow", "Identify months with negative cash flow and determine if they are caused by one-time expenses or ongoing operational deficits. Address by reducing expenses, increasing revenue, or securing additional funding."),
        ("Cash Balance Drops Too Low", "If your cash balance approaches zero, review your timing of large expenses and consider delaying non-essential purchases or seeking short-term financing."),
        ("Seasonal Fluctuations", "If your business experiences seasonal patterns, ensure you build sufficient cash reserves during peak periods to cover expenses during slower periods."),
        ("Overdependence on Single Revenue Stream", "If one revenue stream dominates your income (>70%), consider diversification strategies to reduce business risk."),
        ("High Fixed Costs", "If fixed costs consume a large percentage of revenue, look for opportunities to convert fixed costs to variable costs or negotiate better terms with suppliers."),
        ("Growth Requires More Cash", "Rapid growth often requires more working capital. Plan for this by securing financing before cash shortfalls occur.")
    ]
    
    for issue in issues_solutions:
        row += 1
        problem, solution = issue
        sheet[f'A{row}'] = problem
        sheet[f'A{row}'].font = NORMAL_FONT
        sheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top')
        
        sheet[f'B{row}'] = solution
        sheet[f'B{row}'].font = NORMAL_FONT
        sheet[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Tips for Accurate Projections
    row += 2
    sheet.merge_cells(f'A{row}:B{row}')
    sheet[f'A{row}'] = "TIPS FOR ACCURATE PROJECTIONS"
    sheet[f'A{row}'].font = HEADER_FONT
    sheet[f'A{row}'].fill = LIGHT_GREEN_FILL
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Tips text
    row += 1
    sheet.merge_cells(f'A{row}:B{row+5}')
    sheet[f'A{row}'] = (
        "1. Be Conservative: It's better to underestimate revenue and overestimate expenses than the reverse.\n\n"
        "2. Use Historical Data: Base projections on actual historical performance when available.\n\n"
        "3. Consider Timing: Pay special attention to when cash actually enters and leaves your business, not just when sales or expenses are recorded.\n\n"
        "4. Update Regularly: Cash flow projections should be living documents. Update them monthly with actual results and adjust future projections accordingly.\n\n"
        "5. Involve Team Members: Get input from different departments to ensure projections capture all relevant factors.\n\n"
        "6. Document Assumptions: Keep notes about why you made specific assumptions for future reference."
    )
    sheet[f'A{row}'].font = NORMAL_FONT
    sheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Add borders to all filled cells
    max_row = row + 5
    for row in range(1, max_row + 1):
        for col in range(1, 3):  # Columns A-B
            cell = sheet.cell(row=row, column=col)
            cell.border = THIN_BORDER
    
    # Add contact information
    row = max_row + 2
    sheet.merge_cells(f'A{row}:B{row}')
    sheet[f'A{row}'] = "Need Additional Assistance?"
    sheet[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    row += 1
    sheet.merge_cells(f'A{row}:B{row}')
    sheet[f'A{row}'] = "Clarity Impact Finance offers custom financial advisory services for businesses of all sizes."
    sheet[f'A{row}'].font = Font(name='Arial', size=10)
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    row += 1
    sheet.merge_cells(f'A{row}:B{row}')
    sheet[f'A{row}'] = "Contact us at contact@clarityimpactfinance.com"
    sheet[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')

if __name__ == "__main__":
    output_file = "cash_flow_projection_template.xlsx"
    create_workbook(output_file)
