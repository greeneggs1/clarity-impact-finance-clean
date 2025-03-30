"""
Common styles and utilities for the CDFI Financial Literacy Excel Module.
This module helps avoid circular imports by centralizing all shared styles.
"""

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection

# Configuration
COMPANY_NAME = "Clarity Impact Finance"
OUTPUT_DIR = "excel_tools_output"

# Common styles for consistent appearance
GREEN_FILL = PatternFill(start_color="00A776", end_color="00A776", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="E3F4F1", end_color="E3F4F1", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="F26522", end_color="F26522", fill_type="solid")
LIGHT_ORANGE_FILL = PatternFill(start_color="FCE6DA", end_color="FCE6DA", fill_type="solid")
HEADER_FONT = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
TITLE_FONT = Font(name='Calibri', size=14, bold=True, color="00A776")
SUBTITLE_FONT = Font(name='Calibri', size=12, bold=True, color="F26522")
NOTES_FONT = Font(name='Calibri', size=10, italic=True)

# Create border styles
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Create common cell styles functions
def header_style(cell):
    """Apply header styling to a cell."""
    cell.font = HEADER_FONT
    cell.fill = GREEN_FILL
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def title_style(cell):
    """Apply title styling to a cell."""
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center')

def subtitle_style(cell):
    """Apply subtitle styling to a cell."""
    cell.font = SUBTITLE_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center')

def input_style(cell):
    """Apply input cell styling to a cell."""
    cell.fill = LIGHT_GREEN_FILL
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.protection = Protection(locked=False)

def output_style(cell):
    """Apply output cell styling to a cell."""
    cell.fill = LIGHT_ORANGE_FILL
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
