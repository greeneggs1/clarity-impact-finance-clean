"""
CDFI Financial Literacy Excel Budget Templates

This module contains functions to create the business budget
templates for the Excel-based financial literacy toolkit.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

# Import common styles
from common_styles import (
    GREEN_FILL, LIGHT_GREEN_FILL, ORANGE_FILL, LIGHT_ORANGE_FILL,
    HEADER_FONT, TITLE_FONT, SUBTITLE_FONT, NOTES_FONT,
    thin_border, header_style, title_style, subtitle_style, input_style, output_style
)

# Create budget template worksheet.
def create_budget_template(sheet):
    """Create a business budget template worksheet."""
    # Add title
    sheet['B2'] = "Business Budget Template"
    sheet['B2'].font = Font(name='Calibri', size=16, bold=True, color="00A776")
    
    sheet['B3'] = "Track and manage your business income and expenses"
    sheet['B3'].font = Font(name='Calibri', size=12, italic=True)
    
    # Add instructions
    sheet['B5'] = "Instructions:"
    sheet['B5'].font = SUBTITLE_FONT
    
    instructions = """
    1. Enter your projected income and expenses in the green cells
    2. Enter actual figures as they occur 
    3. Review variance to monitor your budget performance
    4. Use the charts to visualize your budget
    """
    
    sheet['B6'] = instructions.strip()
    sheet['B6'].alignment = Alignment(wrap_text=True)
    
    # Add business information section
    sheet['B8'] = "Business Information"
    sheet['B8'].font = TITLE_FONT
    
    info_labels = ["Business Name", "Budget Period", "Prepared By", "Last Updated"]
    for i, label in enumerate(info_labels, start=9):
        sheet[f'B{i}'] = label
        input_cell = sheet[f'C{i}']
        input_style(input_cell)
    
    # Set default values
    sheet['C10'] = "YYYY-MM to YYYY-MM"
    
    # Set up Income section
    sheet['B13'] = "Income"
    sheet['B13'].font = TITLE_FONT
    
    # Add income headers
    income_headers = ["Category", "Projected", "Actual", "Variance", "% of Total Income"]
    for col, header in enumerate(income_headers, start=2):
        cell = sheet.cell(row=14, column=col)
        cell.value = header
        header_style(cell)
    
    # Add common income categories
    income_categories = [
        "Product Sales",
        "Service Revenue",
        "Consulting Fees",
        "Contract Work",
        "Rental Income",
        "Investment Income",
        "Grants",
        "Other Income"
    ]
    
    for i, category in enumerate(income_categories, start=15):
        sheet[f'B{i}'] = category
        
        # Projected and Actual cells are input cells
        input_style(sheet[f'C{i}'])
        input_style(sheet[f'D{i}'])
        
        # Variance formula
        variance_cell = sheet[f'E{i}']
        variance_cell.value = f'=D{i}-C{i}'
        output_style(variance_cell)
        
        # Percentage formula (only for actual values)
        percent_cell = sheet[f'F{i}']
        percent_cell.value = f'=IF(D{i}=0,0,D{i}/D{i+len(income_categories)+1})'
        percent_cell.number_format = '0.0%'
        output_style(percent_cell)
    
    # Add Total Income row
    total_row = 15 + len(income_categories)
    sheet[f'B{total_row}'] = "Total Income"
    sheet[f'B{total_row}'].font = Font(bold=True)
    
    # Total formulas
    sheet[f'C{total_row}'] = f'=SUM(C15:C{total_row-1})'
    sheet[f'D{total_row}'] = f'=SUM(D15:D{total_row-1})'
    sheet[f'E{total_row}'] = f'=D{total_row}-C{total_row}'
    sheet[f'F{total_row}'] = "100%"
    
    # Format total row
    for col in range(3, 7):
        output_style(sheet.cell(row=total_row, column=col))
    
    # Set up Expenses section
    expense_start_row = total_row + 2
    sheet[f'B{expense_start_row}'] = "Expenses"
    sheet[f'B{expense_start_row}'].font = TITLE_FONT
    
    # Add expense headers
    expense_headers = ["Category", "Projected", "Actual", "Variance", "% of Total Expenses"]
    for col, header in enumerate(expense_headers, start=2):
        cell = sheet.cell(row=expense_start_row+1, column=col)
        cell.value = header
        header_style(cell)
    
    # Add common expense categories
    expense_categories = [
        "Rent/Mortgage",
        "Utilities",
        "Insurance",
        "Payroll",
        "Taxes",
        "Office Supplies",
        "Marketing/Advertising",
        "Professional Services",
        "Loan Repayments",
        "Software/Subscriptions",
        "Travel",
        "Meals/Entertainment",
        "Vehicle Expenses",
        "Equipment Purchases",
        "Licenses/Permits",
        "Training/Education",
        "Miscellaneous"
    ]
    
    for i, category in enumerate(expense_categories):
        row = expense_start_row + 2 + i
        sheet[f'B{row}'] = category
        
        # Projected and Actual cells are input cells
        input_style(sheet[f'C{row}'])
        input_style(sheet[f'D{row}'])
        
        # Variance formula
        variance_cell = sheet[f'E{row}']
        variance_cell.value = f'=D{row}-C{row}'
        output_style(variance_cell)
        
        # Percentage formula (only for actual values)
        percent_cell = sheet[f'F{row}']
        percent_cell.value = f'=IF(D{row}=0,0,D{row}/D{row+len(expense_categories)+1})'
        percent_cell.number_format = '0.0%'
        output_style(percent_cell)
    
    # Add Total Expenses row
    total_expense_row = expense_start_row+2+len(expense_categories)
    sheet[f'B{total_expense_row}'] = "Total Expenses"
    sheet[f'B{total_expense_row}'].font = Font(bold=True)
    
    # Total formulas
    first_expense_row = expense_start_row+2
    sheet[f'C{total_expense_row}'] = f'=SUM(C{first_expense_row}:C{total_expense_row-1})'
    sheet[f'D{total_expense_row}'] = f'=SUM(D{first_expense_row}:D{total_expense_row-1})'
    sheet[f'E{total_expense_row}'] = f'=D{total_expense_row}-C{total_expense_row}'
    sheet[f'F{total_expense_row}'] = "100%"
    
    # Format total expenses row
    for col in range(3, 7):
        output_style(sheet.cell(row=total_expense_row, column=col))
    
    # Add Net Income section
    net_row = total_expense_row + 2
    sheet[f'B{net_row}'] = "Net Income (Income - Expenses)"
    sheet[f'B{net_row}'].font = TITLE_FONT
    
    # Net Income headers
    for col, header in enumerate(["Category", "Projected", "Actual", "Variance"], start=2):
        cell = sheet.cell(row=net_row+1, column=col)
        cell.value = header
        header_style(cell)
    
    # Net Income calculations
    net_calc_row = net_row + 2
    sheet[f'B{net_calc_row}'] = "Net Income"
    sheet[f'B{net_calc_row}'].font = Font(bold=True)
    
    sheet[f'C{net_calc_row}'] = f'=C{total_row}-C{total_expense_row}'
    sheet[f'D{net_calc_row}'] = f'=D{total_row}-D{total_expense_row}'
    sheet[f'E{net_calc_row}'] = f'=D{net_calc_row}-C{net_calc_row}'
    
    # Format Net Income row and add conditional formatting
    for col in range(3, 6):
        cell = sheet.cell(row=net_calc_row, column=col)
        output_style(cell)
        
        # Add special formatting to show positive net income in green, negative in red
        rule = CellIsRule(
            operator='lessThan',
            formula=['0'],
            stopIfTrue=True,
            fill=PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        )
        
        sheet.conditional_formatting.add(f'{get_column_letter(col)}{net_calc_row}', rule)
    
    # Add charts for visualization
    chart_row = net_calc_row + 4
    sheet[f'B{chart_row}'] = "Budget Visualization"
    sheet[f'B{chart_row}'].font = TITLE_FONT
    
    # Create Income vs Expenses Bar Chart
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 10
    bar_chart.title = "Income vs Expenses"
    bar_chart.y_axis.title = "Amount ($)"
    
    # Define data for bar chart
    labels = Reference(sheet, min_col=2, min_row=total_row, max_row=total_expense_row, max_col=2)
    data = Reference(sheet, min_col=4, min_row=total_row-1, max_row=total_expense_row, max_col=4)
    
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(labels)
    
    sheet.add_chart(bar_chart, f"B{chart_row+1}")
    
    # Create Expense Breakdown Pie Chart
    pie_chart = PieChart()
    pie_chart.title = "Expense Breakdown"
    
    # Get top 5 expenses for pie chart (by actual values)
    # In a real implementation, we would dynamically identify top expenses
    # For simplicity, we'll use first 5 categories
    
    expense_labels = Reference(sheet, min_col=2, min_row=expense_start_row+2, max_row=expense_start_row+7)
    expense_data = Reference(sheet, min_col=4, min_row=expense_start_row+1, max_row=expense_start_row+7)
    
    pie_chart.add_data(expense_data, titles_from_data=True)
    pie_chart.set_categories(expense_labels)
    
    sheet.add_chart(pie_chart, f"E{chart_row+1}")
    
    # Set column widths
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15
    
    # Add budget tips
    tips_row = chart_row + 15
    sheet[f'B{tips_row}'] = "Budget Management Tips for CDFI Borrowers"
    sheet[f'B{tips_row}'].font = TITLE_FONT
    
    budget_tips = [
        "Review your budget monthly and adjust projections as needed",
        "Keep a minimum 10% buffer for unexpected expenses",
        "Track seasonal patterns to better predict future needs",
        "Separate fixed costs from variable costs to identify potential savings",
        "Consider creating separate budgets for specific projects or business lines"
    ]
    
    for i, tip in enumerate(budget_tips, start=1):
        sheet[f'B{tips_row+i}'] = f"{i}. {tip}"
