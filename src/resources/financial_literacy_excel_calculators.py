"""
CDFI Financial Literacy Excel Calculators

This module contains functions to create various financial calculators and tools
for the Excel-based financial literacy toolkit.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

# Import common styles
from common_styles import (
    GREEN_FILL, LIGHT_GREEN_FILL, ORANGE_FILL, LIGHT_ORANGE_FILL,
    HEADER_FONT, TITLE_FONT, SUBTITLE_FONT, NOTES_FONT,
    thin_border, header_style, title_style, subtitle_style, input_style, output_style
)

def create_amortization_calculator(sheet):
    """Create a loan amortization calculator worksheet."""
    # Add title
    sheet['B2'] = "Loan Amortization Calculator"
    sheet['B2'].font = Font(name='Calibri', size=16, bold=True, color="00A776")
    
    sheet['B3'] = "Calculate monthly payments and view payment schedules for your loan"
    sheet['B3'].font = Font(name='Calibri', size=12, italic=True)
    
    # Add instructions
    sheet['B5'] = "Instructions:"
    sheet['B5'].font = SUBTITLE_FONT
    
    instructions = """
    1. Enter your loan details in the green input cells below
    2. View your monthly payment and loan summary in the orange cells
    3. Scroll down to see your complete amortization schedule
    """
    
    sheet['B6'] = instructions.strip()
    sheet['B6'].alignment = Alignment(wrap_text=True)
    
    # Set up input section
    sheet['B8'] = "Loan Details"
    sheet['B8'].font = TITLE_FONT
    
    input_labels = [
        "Loan Amount ($)", 
        "Annual Interest Rate (%)", 
        "Loan Term (Years)",
        "Payments Per Year",
        "Start Date"
    ]
    
    for i, label in enumerate(input_labels, start=10):
        sheet[f'B{i}'] = label
        input_cell = sheet[f'C{i}']
        input_style(input_cell)
    
    # Set default values
    sheet['C10'] = 100000  # Loan Amount
    sheet['C11'] = 5.5     # Interest Rate
    sheet['C12'] = 5       # Loan Term
    sheet['C13'] = 12      # Payments Per Year
    sheet['C14'] = "TODAY()"  # Start Date - Excel formula
    
    # Set up output section
    sheet['E8'] = "Loan Summary"
    sheet['E8'].font = TITLE_FONT
    
    output_labels = [
        "Monthly Payment ($)",
        "Total Payments ($)",
        "Total Interest ($)",
        "Total Principal ($)",
        "Last Payment Date"
    ]
    
    for i, label in enumerate(output_labels, start=10):
        sheet[f'E{i}'] = label
        output_cell = sheet[f'F{i}']
        output_style(output_cell)
    
    # Add formulas for calculations
    sheet['F10'] = '=PMT(C11/C13,C12*C13,C10)'  # Monthly Payment
    sheet['F11'] = '=F10*C12*C13'  # Total Payments
    sheet['F12'] = '=F11-C10'  # Total Interest
    sheet['F13'] = '=C10'  # Total Principal
    sheet['F14'] = '=EDATE(C14,C12*12)'  # Last Payment Date
    
    # Set number formats
    sheet['F10'].number_format = '$#,##0.00'
    sheet['F11'].number_format = '$#,##0.00'
    sheet['F12'].number_format = '$#,##0.00'
    sheet['F13'].number_format = '$#,##0.00'
    sheet['F14'].number_format = 'mm/dd/yyyy'
    
    # Add Amortization Schedule section
    sheet['B17'] = "Amortization Schedule"
    sheet['B17'].font = TITLE_FONT
    
    # Add schedule headers
    schedule_headers = ["Payment #", "Payment Date", "Payment Amount", "Principal", "Interest", "Remaining Balance"]
    for col, header in enumerate(schedule_headers, start=2):
        cell = sheet.cell(row=19, column=col)
        cell.value = header
        header_style(cell)
    
    # Add formulas for the first row of the schedule
    sheet['B20'] = 1  # Payment #
    sheet['C20'] = '=EDATE(C14,1)'  # Payment Date
    sheet['D20'] = '=F10'  # Payment Amount
    sheet['E20'] = '=D20-F20'  # Principal
    sheet['F20'] = '=H20*C11/C13'  # Interest
    sheet['G20'] = '=C10-E20'  # Remaining Balance
    
    # Add a hidden helper column for interest calculation
    sheet['H20'] = '=C10'  # Starting Balance for interest calculation
    sheet.column_dimensions['H'].hidden = True
    
    # Add formulas for subsequent rows (up to 360 payments for 30 years)
    max_rows = 360  # Maximum number of payments
    for row in range(21, 21 + max_rows):
        sheet[f'B{row}'] = f'={row-19}'  # Payment #
        sheet[f'C{row}'] = f'=EDATE(C14,{row-19})'  # Payment Date
        sheet[f'D{row}'] = '=F10'  # Payment Amount
        sheet[f'E{row}'] = f'=D{row}-F{row}'  # Principal
        sheet[f'F{row}'] = f'=H{row}*C11/C13'  # Interest
        sheet[f'G{row}'] = f'=G{row-1}-E{row}'  # Remaining Balance
        sheet[f'H{row}'] = f'=G{row-1}'  # Balance for interest calculation
        
        # Add conditional formatting to stop when loan is paid off
        sheet[f'B{row}'].number_format = '0'
        sheet[f'C{row}'].number_format = 'mm/dd/yyyy'
        sheet[f'D{row}'].number_format = '$#,##0.00'
        sheet[f'E{row}'].number_format = '$#,##0.00'
        sheet[f'F{row}'].number_format = '$#,##0.00'
        sheet[f'G{row}'].number_format = '$#,##0.00'
    
    # Format all calculation cells in the schedule
    for row in range(20, 21 + max_rows):
        for col in range(2, 8):
            cell = sheet.cell(row=row, column=col)
            cell.border = thin_border
            if col == 7:  # Remaining Balance
                cell.fill = LIGHT_ORANGE_FILL
    
    # Add a chart showing principal vs interest over time
    chart_row = 21 + max_rows + 5
    sheet[f'B{chart_row}'] = "Principal vs. Interest Over Time"
    sheet[f'B{chart_row}'].font = TITLE_FONT
    
    chart = LineChart()
    chart.title = "Principal vs. Interest Over Loan Term"
    chart.style = 13
    chart.x_axis.title = "Payment Number"
    chart.y_axis.title = "Amount ($)"
    
    # Create data references for the chart
    payment_numbers = Reference(sheet, min_col=2, min_row=19, max_row=19+12*5)  # 5 years of payments
    principal_data = Reference(sheet, min_col=5, min_row=19, max_row=19+12*5) 
    interest_data = Reference(sheet, min_col=6, min_row=19, max_row=19+12*5)
    
    # Add the data series to the chart
    chart.add_data(principal_data, titles_from_data=True)
    chart.add_data(interest_data, titles_from_data=True)
    chart.set_categories(payment_numbers)
    
    # Add the chart to the worksheet
    sheet.add_chart(chart, f"B{chart_row+1}")
    
    # Set column widths
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 20
    
    # Add note about early payoff
    sheet[f'B{chart_row+15}'] = "Want to pay off your loan early? Try making extra payments or increasing your monthly payment amount."
    sheet[f'B{chart_row+15}'].font = NOTES_FONT

def create_affordability_analyzer(sheet):
    """Create a loan affordability calculator worksheet."""
    # Add title
    sheet['B2'] = "Loan Affordability Analyzer"
    sheet['B2'].font = Font(name='Calibri', size=16, bold=True, color="00A776")
    
    sheet['B3'] = "Determine how much financing you can afford based on your business income"
    sheet['B3'].font = Font(name='Calibri', size=12, italic=True)
    
    # Add instructions
    sheet['B5'] = "Instructions:"
    sheet['B5'].font = SUBTITLE_FONT
    
    instructions = """
    1. Enter your business's financial information in the green cells
    2. The calculator will show how much loan you can afford based on CDFI lending standards
    3. Adjust your inputs to see how changes affect your borrowing capacity
    """
    
    sheet['B6'] = instructions.strip()
    sheet['B6'].alignment = Alignment(wrap_text=True)
    
    # Set up input section
    sheet['B8'] = "Business Financial Information"
    sheet['B8'].font = TITLE_FONT
    
    input_labels = [
        "Monthly Business Revenue ($)",
        "Monthly Business Expenses ($)",
        "Existing Monthly Debt Payments ($)",
        "Expected Interest Rate (%)",
        "Desired Loan Term (Years)",
        "Target Debt Service Coverage Ratio"
    ]
    
    for i, label in enumerate(input_labels, start=10):
        sheet[f'B{i}'] = label
        input_cell = sheet[f'C{i}']
        input_style(input_cell)
    
    # Set default values
    sheet['C10'] = 20000   # Monthly Revenue
    sheet['C11'] = 15000   # Monthly Expenses
    sheet['C12'] = 1000    # Existing Debt
    sheet['C13'] = 6.5     # Interest Rate
    sheet['C14'] = 7       # Loan Term
    sheet['C15'] = 1.25    # DSCR
    
    # Add explanations for key inputs
    sheet['D11'] = "Exclude existing debt payments"
    sheet['D11'].font = NOTES_FONT
    
    sheet['D15'] = "Most CDFIs require 1.25 or higher"
    sheet['D15'].font = NOTES_FONT
    
    # Set up output section
    sheet['B17'] = "Loan Affordability Analysis"
    sheet['B17'].font = TITLE_FONT
    
    output_labels = [
        "Monthly Net Operating Income ($)",
        "Maximum Monthly Debt Payment ($)",
        "Available for New Loan Payment ($)",
        "Maximum Loan Amount ($)"
    ]
    
    for i, label in enumerate(output_labels, start=19):
        sheet[f'B{i}'] = label
        output_cell = sheet[f'C{i}']
        output_style(output_cell)
    
    # Add formulas for calculations
    sheet['C19'] = '=C10-C11'  # Monthly NOI
    sheet['C20'] = '=C19/C15'  # Max Monthly Debt Payment
    sheet['C21'] = '=C20-C12'  # Available for New Loan
    sheet['C22'] = '=PV(C13/12,C14*12,-C21,0,0)'  # Maximum Loan Amount
    
    # Add CDFI-specific guidance
    sheet['B24'] = "CDFI Lending Guidance"
    sheet['B24'].font = TITLE_FONT
    
    guidance = [
        ("Debt Service Coverage Ratio (DSCR)", "Most CDFIs require a DSCR of 1.15-1.25 or higher for business loans."),
        ("Loan-to-Value (LTV) for Real Estate", "Typically 75-80% for commercial properties; up to 90% for some mission-focused CDFIs."),
        ("Borrower Contribution", "Usually 10-25% of project costs, though some CDFIs have lower requirements."),
        ("Technical Assistance", "Many CDFIs offer free financial counseling to help improve your financial position.")
    ]
    
    for i, (term, description) in enumerate(guidance, start=26):
        sheet[f'B{i}'] = term
        sheet[f'B{i}'].font = Font(bold=True)
        sheet[f'C{i}'] = description
        sheet[f'C{i}'].alignment = Alignment(wrap_text=True)
    
    # Add a "What If" analysis section
    sheet['E8'] = "What If Analysis"
    sheet['E8'].font = TITLE_FONT
    
    what_if_description = """
    Use this section to explore how changes to your business finances could increase your borrowing capacity.
    """
    sheet['E9'] = what_if_description.strip()
    sheet['E9'].alignment = Alignment(wrap_text=True)
    
    # Set up scenario comparison
    scenario_headers = ["Scenario", "Current", "Revenue +10%", "Expenses -10%", "Longer Term"]
    for col, header in enumerate(scenario_headers, start=5):
        cell = sheet.cell(row=11, column=col)
        cell.value = header
        header_style(cell)
    
    scenario_rows = [
        "Monthly Revenue ($)",
        "Monthly Expenses ($)",
        "Net Operating Income ($)",
        "Maximum Loan Amount ($)"
    ]
    
    for i, label in enumerate(scenario_rows, start=12):
        sheet.cell(row=i, column=5, value=label).font = Font(bold=True)
    
    # Current scenario
    sheet['F12'] = '=C10'  # Revenue
    sheet['F13'] = '=C11'  # Expenses
    sheet['F14'] = '=F12-F13'  # NOI
    sheet['F15'] = '=PV(C13/12,C14*12,-(F14/C15-C12),0,0)'  # Max Loan
    
    # Revenue +10% scenario
    sheet['G12'] = '=C10*1.1'  # Revenue
    sheet['G13'] = '=C11'  # Expenses
    sheet['G14'] = '=G12-G13'  # NOI
    sheet['G15'] = '=PV(C13/12,C14*12,-(G14/C15-C12),0,0)'  # Max Loan
    
    # Expenses -10% scenario
    sheet['H12'] = '=C10'  # Revenue
    sheet['H13'] = '=C11*0.9'  # Expenses
    sheet['H14'] = '=H12-H13'  # NOI
    sheet['H15'] = '=PV(C13/12,C14*12,-(H14/C15-C12),0,0)'  # Max Loan
    
    # Longer term scenario
    sheet['I12'] = '=C10'  # Revenue
    sheet['I13'] = '=C11'  # Expenses
    sheet['I14'] = '=I12-I13'  # NOI
    sheet['I15'] = '=PV(C13/12,(C14+3)*12,-(I14/C15-C12),0,0)'  # Max Loan with 3 years longer term
    
    # Format numbers
    for col in range(6, 10):
        sheet.cell(row=12, column=col).number_format = '$#,##0.00'
        sheet.cell(row=13, column=col).number_format = '$#,##0.00'
        sheet.cell(row=14, column=col).number_format = '$#,##0.00'
        sheet.cell(row=15, column=col).number_format = '$#,##0.00'
    
    # Add chart comparing scenarios
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Maximum Loan Amount by Scenario"
    chart.y_axis.title = "Loan Amount ($)"
    
    data = Reference(sheet, min_col=6, max_col=9, min_row=11, max_row=11+4)
    cats = Reference(sheet, min_col=5, max_col=5, min_row=12, max_row=15)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    sheet.add_chart(chart, "E19")
    
    # Set column widths
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 25
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 15
    sheet.column_dimensions['I'].width = 15
    
    # Add note about improving affordability
    sheet['B31'] = "Note: Improving your Debt Service Coverage Ratio through increased revenue or decreased expenses is the most effective way to increase your borrowing capacity."
    sheet['B31'].font = NOTES_FONT
