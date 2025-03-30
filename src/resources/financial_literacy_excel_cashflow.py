"""
CDFI Financial Literacy Excel Cash Flow Forecasting Tool

This module contains functions to create the cash flow forecasting
tool for the Excel-based financial literacy toolkit.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

# Import common styles
from common_styles import (
    GREEN_FILL, LIGHT_GREEN_FILL, ORANGE_FILL, LIGHT_ORANGE_FILL,
    HEADER_FONT, TITLE_FONT, SUBTITLE_FONT, NOTES_FONT,
    thin_border, header_style, title_style, subtitle_style, input_style, output_style
)

def create_cash_flow_forecast(sheet):
    """Create a cash flow forecasting tool worksheet."""
    # Add title
    sheet['B2'] = "Cash Flow Forecasting Tool"
    sheet['B2'].font = Font(name='Calibri', size=16, bold=True, color="00A776")
    
    sheet['B3'] = "Project your business's cash flow for the next 12 months"
    sheet['B3'].font = Font(name='Calibri', size=12, italic=True)
    
    # Add instructions
    sheet['B5'] = "Instructions:"
    sheet['B5'].font = SUBTITLE_FONT
    
    instructions = """
    1. Enter your starting cash balance
    2. Fill in projected cash inflows and outflows for each month
    3. The tool will calculate your monthly and cumulative cash flow
    4. Watch for negative balances that may indicate cash flow problems
    """
    
    sheet['B6'] = instructions.strip()
    sheet['B6'].alignment = Alignment(wrap_text=True)
    
    # Add the importance of cash flow for CDFI borrowers
    sheet['B8'] = "Why Cash Flow Matters for CDFI Borrowers"
    sheet['B8'].font = TITLE_FONT
    
    importance = """
    Cash flow forecasting is critical for CDFI loan applicants because:
    
    • Lenders analyze cash flow to determine loan affordability
    • It helps identify months when you might need additional financing
    • It demonstrates financial management capability to lenders
    • It allows you to plan for debt repayment responsibilities
    """
    
    sheet['B9'] = importance.strip()
    sheet['B9'].alignment = Alignment(wrap_text=True)
    
    # Add starting cash balance input
    sheet['B12'] = "Starting Cash Balance ($)"
    sheet['B12'].font = Font(bold=True)
    
    starting_cash_cell = sheet['C12']
    starting_cash_cell.value = 10000
    input_style(starting_cash_cell)
    
    # Set up month headers
    months = [
        "January", "February", "March", "April", "May", "June", 
        "July", "August", "September", "October", "November", "December"
    ]
    
    # Create month headers
    sheet['B14'] = "Category"
    header_style(sheet['B14'])
    
    for i, month in enumerate(months, start=1):
        col = get_column_letter(i+2)
        sheet[f'{col}14'] = month
        header_style(sheet[f'{col}14'])
    
    # Set column widths
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 25
    
    for i in range(3, 15):  # Month columns
        sheet.column_dimensions[get_column_letter(i)].width = 12
    
    # Set up cash inflow categories
    sheet['B15'] = "Cash Inflows"
    sheet['B15'].font = SUBTITLE_FONT
    
    inflow_categories = [
        "Sales Revenue",
        "Accounts Receivable Collections",
        "Loan Proceeds",
        "Owner Investments",
        "Asset Sales",
        "Tax Refunds",
        "Grants/Subsidies",
        "Other Income"
    ]
    
    # Add inflow categories
    for i, category in enumerate(inflow_categories, start=16):
        sheet[f'B{i}'] = category
        
        # Add input cells for each month
        for m in range(1, 13):
            col = get_column_letter(m+2)
            cell = sheet[f'{col}{i}']
            input_style(cell)
            cell.number_format = '$#,##0.00'
    
    # Add Total Inflows row
    inflow_total_row = 16 + len(inflow_categories)
    sheet[f'B{inflow_total_row}'] = "Total Inflows"
    sheet[f'B{inflow_total_row}'].font = Font(bold=True)
    
    # Add inflow totals for each month
    for m in range(1, 13):
        col = get_column_letter(m+2)
        formula = f'=SUM({col}16:{col}{inflow_total_row-1})'
        cell = sheet[f'{col}{inflow_total_row}']
        cell.value = formula
        output_style(cell)
        cell.number_format = '$#,##0.00'
    
    # Set up cash outflow categories
    outflow_start_row = inflow_total_row + 2
    sheet[f'B{outflow_start_row}'] = "Cash Outflows"
    sheet[f'B{outflow_start_row}'].font = SUBTITLE_FONT
    
    outflow_categories = [
        "Inventory Purchases",
        "Payroll",
        "Rent/Mortgage",
        "Utilities",
        "Equipment Purchases",
        "Loan Payments",
        "Insurance",
        "Taxes",
        "Marketing/Advertising",
        "Professional Fees",
        "Supplies",
        "Repairs/Maintenance",
        "Travel",
        "Owner Draws",
        "Other Expenses"
    ]
    
    # Add outflow categories
    for i, category in enumerate(outflow_categories, start=outflow_start_row+1):
        sheet[f'B{i}'] = category
        
        # Add input cells for each month
        for m in range(1, 13):
            col = get_column_letter(m+2)
            cell = sheet[f'{col}{i}']
            input_style(cell)
            cell.number_format = '$#,##0.00'
    
    # Add Total Outflows row
    outflow_total_row = outflow_start_row + 1 + len(outflow_categories)
    sheet[f'B{outflow_total_row}'] = "Total Outflows"
    sheet[f'B{outflow_total_row}'].font = Font(bold=True)
    
    # Add outflow totals for each month
    for m in range(1, 13):
        col = get_column_letter(m+2)
        formula = f'=SUM({col}{outflow_start_row+1}:{col}{outflow_total_row-1})'
        cell = sheet[f'{col}{outflow_total_row}']
        cell.value = formula
        output_style(cell)
        cell.number_format = '$#,##0.00'
    
    # Add Net Cash Flow section
    net_start_row = outflow_total_row + 2
    sheet[f'B{net_start_row}'] = "Net Cash Flow (Inflows - Outflows)"
    sheet[f'B{net_start_row}'].font = SUBTITLE_FONT
    
    # Calculate net cash flow for each month
    for m in range(1, 13):
        col = get_column_letter(m+2)
        formula = f'={col}{inflow_total_row}-{col}{outflow_total_row}'
        cell = sheet[f'{col}{net_start_row+1}']
        cell.value = formula
        output_style(cell)
        cell.number_format = '$#,##0.00'
        
        # Add conditional formatting for negative cash flow
        red_rule = CellIsRule(
            operator='lessThan',
            formula=['0'],
            stopIfTrue=True,
            fill=PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        )
        
        sheet.conditional_formatting.add(f'{col}{net_start_row+1}', red_rule)
    
    # Add label for net cash flow row
    sheet[f'B{net_start_row+1}'] = "Net Monthly Cash Flow"
    sheet[f'B{net_start_row+1}'].font = Font(bold=True)
    
    # Add Running Cash Balance section
    balance_row = net_start_row + 3
    sheet[f'B{balance_row}'] = "Running Cash Balance"
    sheet[f'B{balance_row}'].font = SUBTITLE_FONT
    
    # Calculate running cash balance for each month
    for m in range(1, 13):
        col = get_column_letter(m+2)
        
        if m == 1:
            # First month uses starting balance + net cash flow
            formula = f'=$C$12+{col}{net_start_row+1}'
        else:
            # Subsequent months use previous month's balance + current net cash flow
            prev_col = get_column_letter(m+1)
            formula = f'={prev_col}{balance_row+1}+{col}{net_start_row+1}'
        
        cell = sheet[f'{col}{balance_row+1}']
        cell.value = formula
        output_style(cell)
        cell.number_format = '$#,##0.00'
        
        # Add conditional formatting for negative balance
        red_rule = CellIsRule(
            operator='lessThan',
            formula=['0'],
            stopIfTrue=True,
            fill=PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        )
        
        sheet.conditional_formatting.add(f'{col}{balance_row+1}', red_rule)
    
    # Add label for running balance row
    sheet[f'B{balance_row+1}'] = "End of Month Cash Balance"
    sheet[f'B{balance_row+1}'].font = Font(bold=True)
    
    # Add cash flow charts
    chart_row = balance_row + 4
    sheet[f'B{chart_row}'] = "Cash Flow Visualization"
    sheet[f'B{chart_row}'].font = TITLE_FONT
    
    # Create Monthly Net Cash Flow chart
    line_chart1 = LineChart()
    line_chart1.title = "Monthly Net Cash Flow"
    line_chart1.style = 12
    line_chart1.x_axis.title = "Month"
    line_chart1.y_axis.title = "Amount ($)"
    
    # Add data to chart
    months_ref = Reference(sheet, min_col=3, max_col=14, min_row=14, max_row=14)
    data_ref = Reference(sheet, min_col=3, max_col=14, min_row=net_start_row+1, max_row=net_start_row+1)
    
    line_chart1.add_data(data_ref)
    line_chart1.set_categories(months_ref)
    
    sheet.add_chart(line_chart1, f"B{chart_row+1}")
    
    # Create Running Cash Balance chart
    line_chart2 = LineChart()
    line_chart2.title = "Running Cash Balance"
    line_chart2.style = 13
    line_chart2.x_axis.title = "Month"
    line_chart2.y_axis.title = "Amount ($)"
    
    # Add data to chart
    balance_ref = Reference(sheet, min_col=3, max_col=14, min_row=balance_row+1, max_row=balance_row+1)
    
    line_chart2.add_data(balance_ref)
    line_chart2.set_categories(months_ref)
    
    sheet.add_chart(line_chart2, f"H{chart_row+1}")
    
    # Add cash flow tips
    tips_row = chart_row + 15
    sheet[f'B{tips_row}'] = "Cash Flow Management Tips for CDFI Borrowers"
    sheet[f'B{tips_row}'].font = TITLE_FONT
    
    tips = [
        "Identify cash flow gaps early to arrange financing before it becomes an emergency",
        "Negotiate longer payment terms with suppliers and shorter terms with customers",
        "Consider offering early payment discounts to accelerate cash inflows",
        "Build a cash reserve of 3-6 months' operating expenses for unexpected events",
        "Update your cash flow forecast regularly as conditions change",
        "For seasonal businesses, secure lines of credit during strong cash flow periods"
    ]
    
    for i, tip in enumerate(tips, start=1):
        sheet[f'B{tips_row+i}'] = f"{i}. {tip}"
        
    # Add warning about cash flow vs profit
    warning_row = tips_row + len(tips) + 2
    sheet[f'B{warning_row}'] = "Important Note: Cash Flow vs. Profit"
    sheet[f'B{warning_row}'].font = SUBTITLE_FONT
    
    warning = """
    Remember that cash flow is different from profit. A business can be profitable on paper but still face cash shortages if:
    
    • Customers are slow to pay
    • Inventory ties up cash before it can be sold
    • Growth requires investment before generating returns
    • Loan payments exceed net profit
    
    CDFIs assess both profitability and cash flow when evaluating loan applications.
    """
    
    sheet[f'B{warning_row+1}'] = warning.strip()
    sheet[f'B{warning_row+1}'].alignment = Alignment(wrap_text=True)
