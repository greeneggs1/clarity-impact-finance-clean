"""
CDFI Financial Literacy Excel Comparison Tool

This module contains functions to create the CDFI comparison
tool for the Excel-based financial literacy toolkit.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

# Import common styles
from common_styles import (
    GREEN_FILL, LIGHT_GREEN_FILL, ORANGE_FILL, LIGHT_ORANGE_FILL,
    HEADER_FONT, TITLE_FONT, SUBTITLE_FONT, NOTES_FONT,
    thin_border, header_style, title_style, subtitle_style, input_style, output_style
)

def create_cdfi_comparison_tool(sheet):
    """Create a CDFI financing comparison tool worksheet."""
    # Add title
    sheet['B2'] = "CDFI Financing Comparison Tool"
    sheet['B2'].font = Font(name='Calibri', size=16, bold=True, color="00A776")
    
    sheet['B3'] = "Compare financing options from different CDFIs"
    sheet['B3'].font = Font(name='Calibri', size=12, italic=True)
    
    # Add instructions
    sheet['B5'] = "Instructions:"
    sheet['B5'].font = SUBTITLE_FONT
    
    instructions = """
    1. Enter the details of each CDFI loan offer in the green cells
    2. The tool will calculate key metrics for each offer
    3. Compare the total cost, monthly payments, and other factors to make an informed decision
    4. Consider non-financial factors in the qualitative comparison section
    """
    
    sheet['B6'] = instructions.strip()
    sheet['B6'].alignment = Alignment(wrap_text=True)
    
    # Add information about CDFIs
    sheet['B8'] = "About Community Development Financial Institutions (CDFIs)"
    sheet['B8'].font = TITLE_FONT
    
    cdfi_info = """
    CDFIs are specialized financial institutions that work in markets underserved by traditional financial institutions. They provide:
    
    • Responsible, affordable lending to help low-income, low-wealth, and disadvantaged people and communities
    • Financial products and services that may not be available from traditional lenders
    • Technical assistance and financial education to borrowers
    
    Different CDFIs may offer varying terms, rates, and support services, so it's important to compare options.
    """
    
    sheet['B9'] = cdfi_info.strip()
    sheet['B9'].alignment = Alignment(wrap_text=True)
    
    # Set up the comparison table
    sheet['B12'] = "Loan Offer Comparison"
    sheet['B12'].font = TITLE_FONT
    
    # Add CDFI lender headers
    cdfi_headers = ["Loan Features", "CDFI Option 1", "CDFI Option 2", "CDFI Option 3", "Best Option"]
    for col, header in enumerate(cdfi_headers, start=2):
        cell = sheet.cell(row=13, column=col)
        cell.value = header
        header_style(cell)
    
    # Set column widths
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 20
    
    # Add loan feature categories
    loan_features = [
        "Basic Loan Information",
        "CDFI Name",
        "Loan Amount ($)",
        "Loan Term (Years)",
        "Annual Interest Rate (%)",
        "Origination Fee (%)",
        "Other Fees ($)",
        "Closing Costs ($)",
        "Monthly Payment ($)",
        
        "Additional Terms",
        "Prepayment Penalty",
        "Collateral Required",
        "Down Payment Required (%)",
        "Minimum Credit Score",
        "Min. Debt Service Coverage Ratio",
        
        "Cost Analysis",
        "Total Principal ($)",
        "Total Interest ($)",
        "Total Fees and Costs ($)",
        "Total Cost of Borrowing ($)",
        "Annual Percentage Rate (APR) (%)"
    ]
    
    # Add feature rows
    current_row = 14
    for feature in loan_features:
        if feature in ["Basic Loan Information", "Additional Terms", "Cost Analysis"]:
            # These are section headers
            cell = sheet.cell(row=current_row, column=2, value=feature)
            cell.font = SUBTITLE_FONT
            current_row += 1
        else:
            # These are input/calculation rows
            sheet.cell(row=current_row, column=2, value=feature)
            
            # Add input cells for each CDFI option
            for col in range(3, 6):
                cell = sheet.cell(row=current_row, column=col)
                
                # Set special calculation cells
                if feature == "Monthly Payment ($)":
                    # Formula to calculate monthly payment
                    formula = f'=IF(AND(C{current_row-3}>0,C{current_row-2}>0,C{current_row-1}>0),PMT(C{current_row-1}/12,C{current_row-2}*12,C{current_row-3}),"")'
                    if col > 3:
                        col_letter = get_column_letter(col)
                        row_refs = [current_row-3, current_row-2, current_row-1]
                        formula = f'=IF(AND({col_letter}{row_refs[0]}>0,{col_letter}{row_refs[1]}>0,{col_letter}{row_refs[2]}>0),PMT({col_letter}{row_refs[2]}/12,{col_letter}{row_refs[1]}*12,{col_letter}{row_refs[0]}),"")'
                    
                    cell.value = formula
                    output_style(cell)
                    cell.number_format = '$#,##0.00'
                
                elif feature == "Total Principal ($)":
                    # Just reference the loan amount
                    row_ref = current_row - 14  # Approximate position of loan amount row
                    col_letter = get_column_letter(col)
                    cell.value = f'={col_letter}{row_ref}'
                    output_style(cell)
                    cell.number_format = '$#,##0.00'
                
                elif feature == "Total Interest ($)":
                    # Calculate total interest (monthly payment * term * 12 - principal)
                    payment_row = current_row - 7  # Monthly payment row
                    term_row = current_row - 13  # Loan term row
                    amount_row = current_row - 14  # Loan amount row
                    
                    col_letter = get_column_letter(col)
                    formula = f'=IF({col_letter}{payment_row}>0,{col_letter}{payment_row}*{col_letter}{term_row}*12-{col_letter}{amount_row},"")'
                    cell.value = formula
                    output_style(cell)
                    cell.number_format = '$#,##0.00'
                
                elif feature == "Total Fees and Costs ($)":
                    # Sum of origination fee, other fees, and closing costs
                    orig_fee_row = current_row - 11  # Origination Fee row
                    other_fees_row = current_row - 10  # Other Fees row
                    closing_costs_row = current_row - 9  # Closing Costs row
                    amount_row = current_row - 14  # Loan Amount row
                    
                    col_letter = get_column_letter(col)
                    formula = f'={col_letter}{amount_row}*{col_letter}{orig_fee_row}/100+{col_letter}{other_fees_row}+{col_letter}{closing_costs_row}'
                    cell.value = formula
                    output_style(cell)
                    cell.number_format = '$#,##0.00'
                
                elif feature == "Total Cost of Borrowing ($)":
                    # Sum of principal, interest, and fees
                    principal_row = current_row - 2  # Total Principal row
                    interest_row = current_row - 1  # Total Interest row
                    fees_row = current_row  # Total Fees row (current row)
                    
                    col_letter = get_column_letter(col)
                    formula = f'={col_letter}{principal_row}+{col_letter}{interest_row}+{col_letter}{fees_row-1}'
                    cell.value = formula
                    output_style(cell)
                    cell.number_format = '$#,##0.00'
                
                elif feature == "Annual Percentage Rate (APR) (%)":
                    # This is a simplified APR calculation
                    # In reality, APR calculation is more complex
                    interest_row = current_row - 3  # Interest Rate row
                    orig_fee_row = current_row - 12  # Origination Fee row
                    
                    col_letter = get_column_letter(col)
                    formula = f'={col_letter}{interest_row}+{col_letter}{orig_fee_row}/({col_letter}16)'  # Divide by term for annual impact
                    cell.value = formula
                    output_style(cell)
                    cell.number_format = '0.00%'
                
                else:
                    # Regular input cells
                    input_style(cell)
                    
                    # Set specific formats for certain cells
                    if "($)" in feature:
                        cell.number_format = '$#,##0.00'
                    elif "(%)" in feature:
                        cell.number_format = '0.00%'
                    elif feature == "Loan Term (Years)":
                        cell.number_format = '0.0'
            
            # Best option column logic
            if feature in ["Monthly Payment ($)", "Total Interest ($)", "Total Fees and Costs ($)", "Total Cost of Borrowing ($)", "Annual Percentage Rate (APR) (%)"] :
                best_cell = sheet.cell(row=current_row, column=6)
                
                # For these features, lower is better
                col_range = f'C{current_row}:E{current_row}'
                formula = f'=IF(OR(COUNTBLANK({col_range})=3,COUNTIF({col_range},">0")=0),"",INDEX({{"CDFI Option 1","CDFI Option 2","CDFI Option 3"}},MATCH(MIN(IF(C{current_row}>0,C{current_row},99999999),IF(D{current_row}>0,D{current_row},99999999),IF(E{current_row}>0,E{current_row},99999999)),IF(C{current_row}>0,C{current_row},99999999),IF(D{current_row}>0,D{current_row},99999999),IF(E{current_row}>0,E{current_row},99999999)),0)))'
                
                best_cell.value = formula
                output_style(best_cell)
            
            current_row += 1
    
    # Add qualitative comparison section
    qual_row = current_row + 2
    sheet[f'B{qual_row}'] = "Qualitative Comparison"
    sheet[f'B{qual_row}'].font = TITLE_FONT
    
    qual_description = """
    Consider these non-financial factors when comparing CDFI options:
    """
    
    sheet[f'B{qual_row+1}'] = qual_description.strip()
    
    # Add qualitative factor headers
    qual_headers = ["Factor", "CDFI Option 1", "CDFI Option 2", "CDFI Option 3", "Notes"]
    for col, header in enumerate(qual_headers, start=2):
        cell = sheet.cell(row=qual_row+3, column=col)
        cell.value = header
        header_style(cell)
    
    # Add qualitative factors
    factors = [
        "Technical Assistance Available",
        "Industry Expertise",
        "Flexibility of Terms",
        "Speed of Approval Process",
        "Additional Services Offered",
        "Reputation/Reviews",
        "Location/Accessibility"
    ]
    
    for i, factor in enumerate(factors, start=qual_row+4):
        sheet.cell(row=i, column=2, value=factor)
        
        # Add input cells for ratings
        for col in range(3, 7):
            cell = sheet.cell(row=i, column=col)
            input_style(cell)
    
    # Add decision support section
    decision_row = qual_row + len(factors) + 6
    sheet[f'B{decision_row}'] = "Decision Support"
    sheet[f'B{decision_row}'].font = TITLE_FONT
    
    support_text = """
    Before making your final decision, consider:
    
    • Total cost is important, but also consider the CDFI's mission alignment with your business
    • Technical assistance and ongoing support may be worth a slightly higher cost
    • Flexibility in repayment terms may be valuable during business downturns
    • Building a relationship with a CDFI can lead to future financing opportunities
    
    Your final decision should balance quantitative factors (cost, terms) with qualitative factors (support, flexibility, mission).
    """
    
    sheet[f'B{decision_row+1}'] = support_text.strip()
    sheet[f'B{decision_row+1}'].alignment = Alignment(wrap_text=True)
    
    # Add a final recommendation section
    rec_row = decision_row + 10
    sheet[f'B{rec_row}'] = "Final Recommendation"
    sheet[f'B{rec_row}'].font = TITLE_FONT
    
    sheet[f'B{rec_row+1}'] = "Based on my analysis, I recommend:"
    recommendation_cell = sheet[f'B{rec_row+2}']
    input_style(recommendation_cell)
    recommendation_cell.alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[rec_row+2].height = 60
    
    # Add key cost comparison chart
    chart_row = rec_row + 5
    sheet[f'B{chart_row}'] = "Cost Comparison Chart"
    sheet[f'B{chart_row}'].font = TITLE_FONT
    
    # Create cost comparison chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Total Cost Comparison"
    chart.y_axis.title = "Amount ($)"
    
    # Use the total cost of borrowing for each option
    cost_row = current_row - 2  # Total Cost of Borrowing row
    
    # Reference the data and categories
    data = Reference(sheet, min_col=3, max_col=5, min_row=cost_row, max_row=cost_row)
    cats = Reference(sheet, min_col=3, max_col=5, min_row=13, max_row=13)
    
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    
    sheet.add_chart(chart, f"B{chart_row+1}")
    
    # Add tips for negotiating with CDFIs
    tips_row = chart_row + 15
    sheet[f'B{tips_row}'] = "Tips for Negotiating with CDFIs"
    sheet[f'B{tips_row}'].font = TITLE_FONT
    
    tips = [
        "Ask about rate discounts for automatic payments or maintaining deposits with the CDFI",
        "Inquire about different term options to find the right balance between monthly payment and total cost",
        "Request fee waivers, especially if you're participating in their technical assistance programs",
        "Ask about special programs for your industry, location, or business type",
        "If you have multiple offers, respectfully let the CDFI know you're comparing options"
    ]
    
    for i, tip in enumerate(tips, start=1):
        sheet[f'B{tips_row+i}'] = f"{i}. {tip}"
