"""
CDFI Financial Literacy Excel Module Generator

This script generates a comprehensive Excel workbook with multiple tools designed
specifically for users without finance backgrounds who are seeking financing from
CDFIs for business or real estate projects.

The workbook includes interactive calculators and tools for:
- Loan terminology explainer
- Loan amortization calculator
- Loan affordability analyzer
- Business budget template
- Cash flow forecasting tool
- CDFI financing comparison tool

Requirements:
- openpyxl
- pandas

Install with: pip install openpyxl pandas
"""

import os
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

# Import the calculator modules
from financial_literacy_excel_calculators import create_amortization_calculator, create_affordability_analyzer
from financial_literacy_excel_budget import create_budget_template
from financial_literacy_excel_cashflow import create_cash_flow_forecast
from financial_literacy_excel_comparison import create_cdfi_comparison_tool

# Import common styles
from common_styles import (
    COMPANY_NAME, OUTPUT_DIR,
    GREEN_FILL, LIGHT_GREEN_FILL, ORANGE_FILL, LIGHT_ORANGE_FILL,
    HEADER_FONT, TITLE_FONT, SUBTITLE_FONT, NOTES_FONT,
    thin_border, header_style, title_style, subtitle_style, input_style, output_style
)

def create_financial_literacy_workbook():
    """Create a new Excel workbook with multiple financial tools."""
    wb = openpyxl.Workbook()
    
    # Set workbook properties
    wb.properties.creator = COMPANY_NAME
    wb.properties.title = "CDFI Financial Literacy Toolkit"
    wb.properties.created = datetime.now()
    
    # Remove the default sheet and create our custom sheets
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create worksheets for each tool
    intro_sheet = wb.create_sheet(title="Introduction")
    loan_terms_sheet = wb.create_sheet(title="Loan Terminology")
    amortization_sheet = wb.create_sheet(title="Loan Amortization")
    affordability_sheet = wb.create_sheet(title="Loan Affordability")
    budget_sheet = wb.create_sheet(title="Business Budget")
    cash_flow_sheet = wb.create_sheet(title="Cash Flow Forecast")
    comparison_sheet = wb.create_sheet(title="CDFI Comparison")
    
    # Create the Introduction sheet
    create_introduction_sheet(intro_sheet)
    
    # Create the Loan Terminology sheet
    create_loan_terminology_sheet(loan_terms_sheet)
    
    # Create the Loan Amortization Calculator
    create_amortization_calculator(amortization_sheet)
    
    # Create the Loan Affordability Analyzer
    create_affordability_analyzer(affordability_sheet)
    
    # Create the Business Budget Template
    create_budget_template(budget_sheet)
    
    # Create the Cash Flow Forecasting Tool
    create_cash_flow_forecast(cash_flow_sheet)
    
    # Create the CDFI Comparison Tool
    create_cdfi_comparison_tool(comparison_sheet)
    
    # Set Introduction as the active sheet when opening
    wb.active = 0
    
    return wb

def create_introduction_sheet(sheet):
    """Create an introduction page explaining the toolkit."""
    # Add title and company info
    sheet['B2'] = "CDFI Financial Literacy Toolkit"
    sheet['B2'].font = Font(name='Calibri', size=18, bold=True, color="00A776")
    
    sheet['B3'] = f"Provided by {COMPANY_NAME}"
    sheet['B3'].font = Font(name='Calibri', size=14, italic=True)
    
    sheet['B4'] = f"Created: {datetime.now().strftime('%B %d, %Y')}"
    sheet['B4'].font = Font(name='Calibri', size=12)
    
    # Add toolkit description
    sheet['B6'] = "About This Toolkit"
    sheet['B6'].font = TITLE_FONT
    
    description = """
    This Excel workbook provides a collection of financial tools designed specifically for individuals and small business owners who are seeking financing from Community Development Financial Institutions (CDFIs).
    
    The tools included in this workbook will help you understand key financial concepts, plan your financing needs, and make informed decisions about your business or real estate project.
    
    No financial background is required to use these tools. Each sheet includes instructions and explanations of key terms.
    """
    
    sheet['B7'] = description.strip()
    sheet['B7'].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[7].height = 100
    
    # Add guide for using the toolkit
    sheet['B9'] = "How to Use This Toolkit"
    sheet['B9'].font = TITLE_FONT
    
    how_to_use = """
    1. Navigate between tools using the tabs at the bottom of the Excel window.
    2. Enter your information in the light green cells.
    3. Results and calculations will appear in the light orange cells.
    4. Don't change formulas in the orange cells or any cells that aren't light green.
    5. Save a copy of this workbook with your specific information.
    """
    
    sheet['B10'] = how_to_use.strip()
    sheet['B10'].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[10].height = 85
    
    # Add tools overview
    sheet['B12'] = "Tools Included in This Workbook"
    sheet['B12'].font = TITLE_FONT
    
    tools = [
        ("Loan Terminology", "Explanations of common loan terms and concepts used by CDFIs"),
        ("Loan Amortization", "Calculate monthly payments, total interest, and view payment schedules"),
        ("Loan Affordability", "Determine how much financing you can afford based on your business income"),
        ("Business Budget", "Create and manage a business budget with income and expense tracking"),
        ("Cash Flow Forecast", "Project your business's cash flow for the next 12-24 months"),
        ("CDFI Comparison", "Compare financing options from different CDFIs")
    ]
    
    for i, (tool_name, description) in enumerate(tools, start=13):
        sheet[f'B{i}'] = tool_name
        sheet[f'B{i}'].font = SUBTITLE_FONT
        
        sheet[f'C{i}'] = description
        sheet[f'C{i}'].alignment = Alignment(wrap_text=True)
    
    # Add access information for secure features - based on client login system
    sheet['B20'] = "Access to Additional Resources"
    sheet['B20'].font = TITLE_FONT
    
    access_info = """
    Clarity Impact Finance offers additional resources through our secure client portal:
    
    • To access these resources, you'll need an invitation code (format: CIF-XXXXX)
    • Invitation codes are valid for 30 days
    • Contact your CDFI representative for an invitation code
    • Once registered, you can log in with your email and password
    
    Visit our website for more information about accessing these resources.
    """
    
    sheet['B21'] = access_info.strip()
    sheet['B21'].alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[21].height = 100
    
    # Add help and contact information
    sheet['B24'] = "Need Help?"
    sheet['B24'].font = TITLE_FONT
    
    sheet['B25'] = f"Email: contact@clarityimpactfinance.com"
    sheet['B26'] = f"Website: www.clarityimpactfinance.com"
    sheet['B27'] = "Additional resources and guides are available on our website."
    
    # Set column widths
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 60
    sheet.column_dimensions['D'].width = 15
    
    # Lock the sheet to prevent modifications
    sheet.protection.sheet = True
    sheet.protection.password = "clarity"

def create_loan_terminology_sheet(sheet):
    """Create a sheet explaining common loan terminology for CDFI borrowers."""
    # Add title
    sheet['B2'] = "Loan Terminology Explainer"
    sheet['B2'].font = Font(name='Calibri', size=16, bold=True, color="00A776")
    
    sheet['B3'] = "Understanding key financial terms used in CDFI financing"
    sheet['B3'].font = Font(name='Calibri', size=12, italic=True)
    
    # Add search functionality hint
    sheet['B5'] = "Looking for a specific term? Press Ctrl+F to search this sheet."
    sheet['B5'].font = NOTES_FONT
    
    # Set up headers
    headers = ["Term", "Definition", "Example", "Why It Matters"]
    for col, header in enumerate(headers, start=2):
        cell = sheet.cell(row=7, column=col)
        cell.value = header
        header_style(cell)
    
    # Set column widths
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 25
    sheet.column_dimensions['E'].width = 30
    
    # Add loan terms
    terms = [
        (
            "Principal", 
            "The original amount of money borrowed in a loan.",
            "A $100,000 loan for business equipment has a principal of $100,000.", 
            "Determines your total debt obligation and affects the interest you'll pay."
        ),
        (
            "Interest Rate", 
            "The percentage of the principal charged by lenders for the use of their money. Can be fixed (stays the same) or variable (changes over time).",
            "A loan with a 5% fixed annual interest rate.", 
            "Directly impacts your monthly payment and total cost of the loan."
        ),
        (
            "APR (Annual Percentage Rate)", 
            "The yearly cost of a loan including interest and fees, expressed as a percentage.",
            "A loan with a 5% interest rate might have a 5.5% APR when fees are included.", 
            "Helps you compare different loan offers on a standardized basis."
        ),
        (
            "Term", 
            "The length of time to repay the loan in full.",
            "A 5-year term means you have 60 monthly payments to repay the loan.", 
            "Affects your monthly payment amount and total interest paid."
        ),
        (
            "Amortization", 
            "The process of paying off a loan with regular payments over time, where each payment includes both principal and interest.",
            "In a 30-year mortgage, early payments are mostly interest, while later payments are mostly principal.", 
            "Helps you understand how much of each payment goes to principal vs. interest."
        ),
        (
            "Amortization Schedule", 
            "A table showing the breakdown of each loan payment into principal and interest over the life of the loan.",
            "Month 1: $500 payment = $100 principal + $400 interest\nMonth 2: $500 payment = $102 principal + $398 interest", 
            "Helps track loan payoff progress and plan for early payoff strategies."
        ),
        (
            "Collateral", 
            "Assets pledged to secure a loan that can be seized by the lender if you fail to repay.",
            "Using business equipment or real estate as collateral for a loan.", 
            "Reduces lender risk, which can help you qualify for larger loans or better rates."
        ),
        (
            "Debt Service Coverage Ratio (DSCR)", 
            "A measure of cash flow available to pay current debt obligations. Calculated as Net Operating Income divided by Total Debt Service.",
            "DSCR of 1.25 means you have 25% more income than needed for loan payments.", 
            "Lenders use this to determine if your business generates enough cash to repay the loan."
        ),
        (
            "Loan-to-Value (LTV) Ratio", 
            "The ratio of a loan amount to the value of the asset being purchased.",
            "An $80,000 loan for a $100,000 property has an 80% LTV ratio.", 
            "Affects your interest rate, loan terms, and whether you need additional collateral."
        ),
        (
            "Origination Fee", 
            "A fee charged by lenders to process a new loan application.",
            "A 1% origination fee on a $100,000 loan would cost $1,000.", 
            "Adds to the upfront cost of obtaining a loan and affects the total cost of borrowing."
        ),
        (
            "Prepayment Penalty", 
            "A fee charged by some lenders if you pay off your loan before the end of the term.",
            "A 2% prepayment penalty on a $100,000 balance would cost $2,000 if you pay off early.", 
            "May limit your flexibility to refinance or pay off the loan early."
        ),
        (
            "Term Sheet", 
            "A non-binding document outlining the key terms and conditions of a proposed loan.",
            "A term sheet might specify loan amount, interest rate, term, collateral requirements, and key conditions.", 
            "Allows you to understand and compare loan offers before signing binding documents."
        ),
        (
            "Debt-to-Income (DTI) Ratio", 
            "The percentage of your monthly income that goes toward paying debts.",
            "If your monthly income is $10,000 and debt payments total $3,000, your DTI is 30%.", 
            "Lenders use this to assess your ability to take on additional debt."
        ),
        (
            "Grace Period", 
            "A set period after a payment due date during which a late payment will not result in penalties.",
            "A 15-day grace period means no late fees if you pay within 15 days after the due date.", 
            "Provides flexibility in timing your payments."
        ),
        (
            "Balloon Payment", 
            "A large, lump-sum payment due at the end of a loan term.",
            "A 5-year loan with a $50,000 balloon payment at the end.", 
            "Requires planning to either refinance or have funds available for the final payment."
        )
    ]
    
    # Add terms to sheet
    for i, (term, definition, example, importance) in enumerate(terms, start=8):
        sheet.cell(row=i, column=2, value=term).font = Font(bold=True)
        sheet.cell(row=i, column=3, value=definition).alignment = Alignment(wrap_text=True)
        sheet.cell(row=i, column=4, value=example).alignment = Alignment(wrap_text=True)
        sheet.cell(row=i, column=5, value=importance).alignment = Alignment(wrap_text=True)
        
        # Add alternating row colors for readability
        if i % 2 == 0:
            for col in range(2, 6):
                sheet.cell(row=i, column=col).fill = LIGHT_GREEN_FILL
    
    # Add note at bottom
    note_row = 8 + len(terms) + 2
    sheet.cell(row=note_row, column=2, value="Note: This is not an exhaustive list. Ask your CDFI loan officer about any terms you don't understand.")
    sheet.cell(row=note_row, column=2).font = NOTES_FONT

def main():
    """Main function to create and save the Excel workbook."""
    try:
        # Create output directory if it doesn't exist
        if not os.path.exists(OUTPUT_DIR):
            os.makedirs(OUTPUT_DIR)
        
        # Create the workbook
        wb = create_financial_literacy_workbook()
        
        # Save the workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(OUTPUT_DIR, f"CDFI_Financial_Toolkit_{timestamp}.xlsx")
        wb.save(output_file)
        
        print(f"Excel workbook created successfully: {output_file}")
        return True
    except Exception as e:
        print(f"Error creating Excel workbook: {str(e)}")
        return False

if __name__ == "__main__":
    main()
