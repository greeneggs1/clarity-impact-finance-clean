import os
import sys
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import range_boundaries

# Constants for styling
GREEN_FILL = PatternFill(start_color="1B4620", end_color="1B4620", fill_type="solid")  # Dark Green
LIGHT_GREEN_FILL = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")  # Light Green
ORANGE_FILL = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")  # Orange
GREY_FILL = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")  # Light Grey
LIGHT_BLUE_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")  # Light Blue for data entry
LIGHT_GREY_FILL = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")  # Light Grey

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

THICK_BORDER = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

def create_workbook():
    """Create and set up the Excel workbook with all worksheets"""
    # Create workbook
    wb = Workbook()
    
    # Define worksheet names
    ws_names = {
        'Dashboard': 'Dashboard',
        'Deal_Structure': 'Deal Structure',
        'Financial_Analysis': 'Financial Analysis',
        'Reference_Materials': 'Reference Materials',
        'Structure_Chart': 'NMTC Structure Chart',
        'Input': 'Input'
    }
    
    # Rename the default sheet
    sheet = wb.active
    sheet.title = 'Dashboard'
    
    # Create other worksheets
    for key, name in ws_names.items():
        if key != 'Dashboard':  # Dashboard already exists as the active sheet
            wb.create_sheet(key)
    
    # Set up each worksheet
    setup_dashboard(wb, wb['Dashboard'])
    setup_deal_structure(wb, wb['Deal_Structure'])
    setup_financial_analysis(wb, wb['Financial_Analysis'])
    setup_reference_materials(wb, wb['Reference_Materials'])
    setup_structure_chart(wb, wb['Structure_Chart'])
    setup_input_tab(wb, wb['Input'])
    
    # Add data validation and formulas
    add_validation_and_formulas(wb)
    
    # Add example data
    add_example_data(wb)
    
    # Add formulas for calculations and dynamic updates
    add_formulas(wb)
    
    return wb

def setup_dashboard(wb, sheet):
    """Set up the main dashboard with overview, metrics, and status tracking"""
    # Set column widths
    for col in range(1, 20):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 15
    
    # Title and branding
    sheet.merge_cells('A1:I1')
    sheet['A1'] = "NMTC SOURCE LEVERAGE LENDER UNDERWRITING CHECKLIST"
    sheet['A1'].font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['A1'].fill = GREEN_FILL
    
    # Clarity Impact Finance branding
    sheet.merge_cells('A2:I2')
    sheet['A2'] = "Clarity Impact Finance | contact@clarityimpactfinance.com | Admin Portal: /admin"
    sheet['A2'].font = Font(name='Arial', size=10, italic=True)
    sheet['A2'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['A2'].fill = LIGHT_GREEN_FILL
    
    # Toggle for Example Data
    sheet['A4'] = "Show Example Data:"
    sheet['A4'].font = Font(bold=True)
    sheet['B4'] = "YES"  # Default value
    sheet['B4'].font = Font(bold=True)
    
    # Project Information Section
    sheet.merge_cells('A6:E6')
    sheet['A6'] = "PROJECT INFORMATION"
    sheet['A6'].font = Font(bold=True, color="FFFFFF")
    sheet['A6'].fill = GREEN_FILL
    sheet['A6'].alignment = Alignment(horizontal='center')
    
    labels = [
        ("Project Name:", "B7", "C7:E7"),
        ("Sponsor/Borrower:", "B8", "C8:E8"),
        ("NMTC Allocation Amount:", "B9", "C9:E9"),
        ("Source Leverage Loan Amount:", "B10", "C10:E10"),
        ("QLICI Amount:", "B11", "C11:E11"),
        ("Project Location:", "B12", "C12:E12"),
        ("CDE Name:", "B13", "C13:E13"),
        ("Tax Credit Investor:", "B14", "C14:E14"),
        ("Closing Date (Target):", "B15", "C15:E15")
    ]
    
    # Apply styling to the labels and merge value cells
    for label, label_cell, value_range in labels:
        sheet[label_cell] = label
        sheet[label_cell].font = Font(bold=True)
        sheet[label_cell].alignment = Alignment(horizontal='right')
        sheet.merge_cells(value_range)
        value_cell = value_range.split(':')[0]
        sheet[value_cell].fill = LIGHT_BLUE_FILL
        sheet[value_cell].border = THIN_BORDER
    
    # Status Dashboard
    sheet.merge_cells('G6:I6')
    sheet['G6'] = "UNDERWRITING STATUS"
    sheet['G6'].font = Font(bold=True, color="FFFFFF")
    sheet['G6'].fill = GREEN_FILL
    sheet['G6'].alignment = Alignment(horizontal='center')
    
    status_items = [
        ("Deal Structure Review", "G7", "H7:I7"),
        ("Financial Analysis", "G8", "H8:I8"),
        ("Overall Recommendation", "G9", "H9:I9")
    ]
    
    # Apply styling to the status items
    for item, label_cell, value_range in status_items:
        sheet[label_cell] = item
        sheet[label_cell].font = Font(bold=True)
        sheet[label_cell].alignment = Alignment(horizontal='right')
        sheet.merge_cells(value_range)
        value_cell = value_range.split(':')[0]
        sheet[value_cell].fill = LIGHT_BLUE_FILL
        sheet[value_cell].border = THIN_BORDER
    
    # Key Risk Metrics
    sheet.merge_cells('A17:I17')
    sheet['A17'] = "KEY METRICS FOR SOURCE LEVERAGE LENDER"
    sheet['A17'].font = Font(bold=True, color="FFFFFF")
    sheet['A17'].fill = GREEN_FILL
    sheet['A17'].alignment = Alignment(horizontal='center')
    
    # Financial metrics table
    metrics = [
        ("Metric", "Description", "Value", "Threshold", "Status"),
        ("Loan-to-Value (LTV)", "Loan amount as % of collateral value", "", "≤ 80%", ""),
        ("Debt Service Coverage Ratio", "Project NOI / Debt Service", "", "≥ 1.25x", ""),
        ("Sponsor Net Worth", "Total assets minus total liabilities", "", "≥ 25% of loan", ""),
        ("Sponsor Liquidity", "Cash and cash equivalents", "", "≥ 10% of loan", ""),
        ("Recapture Risk Score", "Assessment of NMTC recapture risk", "", "Low-Medium", ""),
        ("Guaranty Coverage", "% of loan covered by guarantees", "", "≥ 100%", ""),
        ("Exit Strategy Viability", "Likelihood of successful exit at year 7", "", "High", "")
    ]
    
    # Create the metrics table
    for row_idx, (col1, col2, col3, col4, col5) in enumerate(metrics, start=18):
        sheet[f'A{row_idx}'] = col1
        sheet[f'C{row_idx}'] = col2
        sheet[f'F{row_idx}'] = col3
        sheet[f'G{row_idx}'] = col4
        sheet[f'I{row_idx}'] = col5
        
        # Style header row
        if row_idx == 18:
            for col in ['A', 'C', 'F', 'G', 'I']:
                sheet[f'{col}{row_idx}'].font = Font(bold=True)
                sheet[f'{col}{row_idx}'].fill = GREY_FILL
                sheet[f'{col}{row_idx}'].border = THIN_BORDER
                sheet[f'{col}{row_idx}'].alignment = Alignment(horizontal='center')
        else:
            # Style data rows
            sheet[f'A{row_idx}'].font = Font(bold=True)
            for col in ['A', 'C', 'F', 'G', 'I']:
                sheet[f'{col}{row_idx}'].border = THIN_BORDER
                
            # Make value column editable with light blue background
            sheet[f'F{row_idx}'].fill = LIGHT_BLUE_FILL
    
    # Underwriting Guidance
    sheet.merge_cells('A26:I26')
    sheet['A26'] = "SOURCE LEVERAGE LENDER GUIDANCE"
    sheet['A26'].font = Font(bold=True, color="FFFFFF")
    sheet['A26'].fill = GREEN_FILL
    sheet['A26'].alignment = Alignment(horizontal='center')
    
    guidance_text = """
    This checklist is designed specifically for financial institutions serving as Source Leverage Lenders in NMTC transactions. 
    As a Source Leverage Lender, your role is to provide debt to the Sponsor/Borrower who will then make a leverage loan to the Investment Fund.
    
    KEY RESPONSIBILITIES:
    1. Assess the Sponsor/Borrower's creditworthiness and capacity to repay the source leverage loan
    2. Understand the NMTC structure and how your loan fits within the overall transaction
    3. Evaluate project feasibility and community impact
    4. Analyze collateral and guarantees specific to your position
    5. Identify and mitigate risks particular to NMTC transactions
    
    USING THIS CHECKLIST:
    • Complete each section thoroughly, starting with the Deal Structure Review and working through each tab
    • Use the Dashboard for a high-level overview of the transaction and key metrics
    • Reference the Status indicators to track your underwriting process
    • Consult the Reference Materials tab for NMTC-specific terminology and structure guidance
    
    The Dashboard will automatically update as you complete each section. Set the "Show Example Data" toggle to "NO" when ready to input your own data.
    """
    
    sheet.merge_cells('A27:I35')
    sheet['A27'] = guidance_text.strip()
    sheet['A27'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Final recommendation section
    sheet.merge_cells('A37:I37')
    sheet['A37'] = "FINAL RECOMMENDATION"
    sheet['A37'].font = Font(bold=True, color="FFFFFF")
    sheet['A37'].fill = GREEN_FILL
    sheet['A37'].alignment = Alignment(horizontal='center')
    
    sheet.merge_cells('A38:I41')
    sheet['A38'].fill = LIGHT_BLUE_FILL
    sheet['A38'].border = THIN_BORDER
    sheet['A38'].alignment = Alignment(wrap_text=True, vertical='top')

def setup_deal_structure(wb, sheet):
    """Set up the Deal Structure worksheet for analyzing transaction structure"""
    # Set column widths
    for col in range(1, 20):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 15
    
    # Title
    sheet.merge_cells('A1:I1')
    sheet['A1'] = "DEAL STRUCTURE ANALYSIS"
    sheet['A1'].font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['A1'].fill = GREEN_FILL
    
    # NMTC Structure Overview
    sheet.merge_cells('A3:I3')
    sheet['A3'] = "NMTC TRANSACTION STRUCTURE"
    sheet['A3'].font = Font(bold=True, color="FFFFFF")
    sheet['A3'].fill = GREEN_FILL
    sheet['A3'].alignment = Alignment(horizontal='center')
    
    structure_fields = [
        ("Transaction Structure Type:", "B4", "C4:E4"),
        ("Total Transaction Size ($):", "B5", "C5:E5"),
        ("NMTC Allocation Amount ($):", "B6", "C6:E6"),
        ("QLICI Loan Amount ($):", "B7", "C7:E7"),
        ("Tax Credit Equity Amount ($):", "B8", "C8:E8"),
        ("Source Leverage Loan Amount ($):", "B9", "C9:E9"),
        ("Leverage Ratio:", "B10", "C10:E10"),
        ("Expected Closing Date:", "B11", "C11:E11"),
        ("Expected Unwind/Exit Date:", "B12", "C12:E12")
    ]
    
    # Apply styling to the structure fields
    for label, label_cell, value_range in structure_fields:
        sheet[label_cell] = label
        sheet[label_cell].font = Font(bold=True)
        sheet[label_cell].alignment = Alignment(horizontal='right')
        sheet.merge_cells(value_range)
        value_cell = value_range.split(':')[0]
        sheet[value_cell].fill = LIGHT_BLUE_FILL
        sheet[value_cell].border = THIN_BORDER
    
    # CDE Information
    sheet.merge_cells('A14:I14')
    sheet['A14'] = "CDE INFORMATION"
    sheet['A14'].font = Font(bold=True, color="FFFFFF")
    sheet['A14'].fill = GREEN_FILL
    sheet['A14'].alignment = Alignment(horizontal='center')
    
    cde_fields = [
        ("CDE Name:", "B15", "C15:E15"),
        ("CDE Certification #:", "B16", "C16:E16"),
        ("CDE Controlling Entity:", "B17", "C17:E17"),
        ("Total NMTC Allocation:", "B18", "C18:E18"),
        ("Allocation Remaining:", "B19", "C19:E19"),
        ("Allocation Round Year:", "B20", "C20:E20"),
        ("Primary CDE Contact:", "B21", "C21:E21"),
        ("CDE Fee Structure:", "B22", "C22:E22")
    ]
    
    # Apply styling to the CDE fields
    for label, label_cell, value_range in cde_fields:
        sheet[label_cell] = label
        sheet[label_cell].font = Font(bold=True)
        sheet[label_cell].alignment = Alignment(horizontal='right')
        sheet.merge_cells(value_range)
        value_cell = value_range.split(':')[0]
        sheet[value_cell].fill = LIGHT_BLUE_FILL
        sheet[value_cell].border = THIN_BORDER
    
    # Tax Credit Investor Information
    sheet.merge_cells('A24:I24')
    sheet['A24'] = "TAX CREDIT INVESTOR INFORMATION"
    sheet['A24'].font = Font(bold=True, color="FFFFFF")
    sheet['A24'].fill = GREEN_FILL
    sheet['A24'].alignment = Alignment(horizontal='center')
    
    investor_fields = [
        ("Investor Name:", "B25", "C25:E25"),
        ("Primary Contact:", "B26", "C26:E26"),
        ("Tax Credit Price:", "B27", "C27:E27"),
        ("Equity Investment Amount:", "B28", "C28:E28"),
        ("Expected Yield:", "B29", "C29:E29"),
        ("Investor Requirements:", "B30", "C30:E30"),
        ("Investor Fee Structure:", "B31", "C31:E31")
    ]
    
    # Apply styling to the investor fields
    for label, label_cell, value_range in investor_fields:
        sheet[label_cell] = label
        sheet[label_cell].font = Font(bold=True)
        sheet[label_cell].alignment = Alignment(horizontal='right')
        sheet.merge_cells(value_range)
        value_cell = value_range.split(':')[0]
        sheet[value_cell].fill = LIGHT_BLUE_FILL
        sheet[value_cell].border = THIN_BORDER
    
    # Investment Fund & QLICI Structure
    sheet.merge_cells('A33:I33')
    sheet['A33'] = "INVESTMENT FUND & QLICI STRUCTURE"
    sheet['A33'].font = Font(bold=True, color="FFFFFF")
    sheet['A33'].fill = GREEN_FILL
    sheet['A33'].alignment = Alignment(horizontal='center')
    
    if_fields = [
        ("Investment Fund Name:", "B34", "C34:E34"),
        ("Investment Fund Manager:", "B35", "C35:E35"),
        ("QLICI Borrower Entity:", "B36", "C36:E36"),
        ("QLICI A Loan Amount:", "B37", "C37:E37"),
        ("QLICI A Loan Interest Rate:", "B38", "C38:E38"),
        ("QLICI A Loan Term:", "B39", "C39:E39"),
        ("QLICI B Loan (Sub) Amount:", "B40", "C40:E40"),
        ("Put/Call Options:", "B41", "C41:E41")
    ]
    
    # Apply styling to the investment fund fields
    for label, label_cell, value_range in if_fields:
        sheet[label_cell] = label
        sheet[label_cell].font = Font(bold=True)
        sheet[label_cell].alignment = Alignment(horizontal='right')
        sheet.merge_cells(value_range)
        value_cell = value_range.split(':')[0]
        sheet[value_cell].fill = LIGHT_BLUE_FILL
        sheet[value_cell].border = THIN_BORDER
    
    # Flow of Funds
    sheet.merge_cells('A43:I43')
    sheet['A43'] = "FLOW OF FUNDS SUMMARY"
    sheet['A43'].font = Font(bold=True, color="FFFFFF")
    sheet['A43'].fill = GREEN_FILL
    sheet['A43'].alignment = Alignment(horizontal='center')
    
    # Table headers for flow of funds
    flow_headers = ["Source", "Amount ($)", "Recipient", "Purpose", "Notes"]
    for col_idx, header in enumerate(flow_headers, start=1):
        col_letter = get_column_letter(col_idx*2-1)
        sheet[f'{col_letter}44'] = header
        sheet[f'{col_letter}44'].font = Font(bold=True)
        sheet[f'{col_letter}44'].fill = GREY_FILL
        sheet[f'{col_letter}44'].border = THIN_BORDER
        sheet[f'{col_letter}44'].alignment = Alignment(horizontal='center')
        
        if col_idx < len(flow_headers):
            sheet.merge_cells(start_row=44, start_column=col_idx*2-1, end_row=44, end_column=col_idx*2)
    
    # Add rows for flow of funds
    for row in range(45, 52):
        for col_idx in range(1, 10, 2):
            cell = sheet.cell(row=row, column=col_idx)
            cell.fill = LIGHT_BLUE_FILL
            cell.border = THIN_BORDER
            if col_idx < 9:
                sheet.merge_cells(start_row=row, start_column=col_idx, end_row=row, end_column=col_idx+1)
    
    # Source Leverage Lender Focus
    sheet.merge_cells('A54:I54')
    sheet['A54'] = "SOURCE LEVERAGE LENDER POSITION ASSESSMENT"
    sheet['A54'].font = Font(bold=True, color="FFFFFF")
    sheet['A54'].fill = GREEN_FILL
    sheet['A54'].alignment = Alignment(horizontal='center')
    
    lender_fields = [
        ("Collateral Position:", "B55", "C55:E55"),
        ("Security Interest Description:", "B56", "C56:E56"),
        ("Intercreditor Agreements:", "B57", "C57:E57"),
        ("Fee Stream to Lender:", "B58", "C58:E58"),
        ("Impact on Exit Strategy:", "B59", "C59:E59"),
        ("Forbearance Provisions:", "B60", "C60:E60")
    ]
    
    # Apply styling to the lender position fields
    for label, label_cell, value_range in lender_fields:
        sheet[label_cell] = label
        sheet[label_cell].font = Font(bold=True)
        sheet[label_cell].alignment = Alignment(horizontal='right')
        sheet.merge_cells(value_range)
        value_cell = value_range.split(':')[0]
        sheet[value_cell].fill = LIGHT_BLUE_FILL
        sheet[value_cell].border = THIN_BORDER
    
    # Structure Assessment
    sheet.merge_cells('A62:I62')
    sheet['A62'] = "DEAL STRUCTURE ASSESSMENT & RECOMMENDATIONS"
    sheet['A62'].font = Font(bold=True, color="FFFFFF")
    sheet['A62'].fill = GREEN_FILL
    sheet['A62'].alignment = Alignment(horizontal='center')
    
    assessment_fields = [
        ("Structure Complexity Rating:", "B63", "C63:E63"),
        ("Documentation Completeness:", "B64", "C64:E64"),
        ("Structure Risk Rating:", "B65", "C65:E65"),
        ("Overall Structure Assessment:", "B66", "C66:E66")
    ]
    
    for label, label_cell, value_range in assessment_fields:
        sheet[label_cell] = label
        sheet[label_cell].font = Font(bold=True)
        sheet[label_cell].alignment = Alignment(horizontal='right')
        sheet.merge_cells(value_range)
        value_cell = value_range.split(':')[0]
        sheet[value_cell].fill = LIGHT_BLUE_FILL
        sheet[value_cell].border = THIN_BORDER
    
    # Comments section
    sheet.merge_cells('A68:I68')
    sheet['A68'] = "ADDITIONAL NOTES & COMMENTS"
    sheet['A68'].font = Font(bold=True)
    sheet['A68'].fill = GREY_FILL
    sheet['A68'].alignment = Alignment(horizontal='center')
    
    sheet.merge_cells('A69:I72')
    sheet['A69'].fill = LIGHT_BLUE_FILL
    sheet['A69'].alignment = Alignment(wrap_text=True, vertical='top')
    sheet['A69'].border = THIN_BORDER

def setup_financial_analysis(wb, sheet):
    """Set up the Financial Analysis worksheet for assessing financial health and projections"""
    # Set column widths
    for col in range(1, 20):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 15
    
    # Title
    sheet.merge_cells('A1:I1')
    sheet['A1'] = "FINANCIAL ANALYSIS"
    sheet['A1'].font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['A1'].fill = GREEN_FILL
    
    # Historical Financial Analysis
    sheet.merge_cells('A3:I3')
    sheet['A3'] = "HISTORICAL FINANCIAL PERFORMANCE"
    sheet['A3'].font = Font(bold=True, color="FFFFFF")
    sheet['A3'].fill = GREEN_FILL
    sheet['A3'].alignment = Alignment(horizontal='center')
    
    # Headers for historical financial data
    period_headers = ["", "Year -3", "Year -2", "Year -1", "Current Year"]
    for col_idx, header in enumerate(period_headers):
        col_letter = get_column_letter(col_idx + 1)
        sheet[f'{col_letter}4'] = header
        sheet[f'{col_letter}4'].font = Font(bold=True)
        sheet[f'{col_letter}4'].fill = GREY_FILL
        sheet[f'{col_letter}4'].alignment = Alignment(horizontal='center')
        sheet[f'{col_letter}4'].border = THIN_BORDER
    
    # Key financial metrics
    historical_metrics = [
        "Total Revenue",
        "EBITDA",
        "Net Income",
        "Total Assets",
        "Total Liabilities",
        "Net Worth/Equity",
        "Current Ratio",
        "Debt-to-Equity Ratio",
        "Profit Margin (%)",
        "ROA (%)",
        "ROE (%)",
        "Debt Service Coverage"
    ]
    
    for row_idx, metric in enumerate(historical_metrics, start=5):
        sheet[f'A{row_idx}'] = metric
        sheet[f'A{row_idx}'].font = Font(bold=True)
        sheet[f'A{row_idx}'].alignment = Alignment(horizontal='left')
        
        for col_idx in range(2, 6):
            col_letter = get_column_letter(col_idx)
            sheet[f'{col_letter}{row_idx}'].fill = LIGHT_BLUE_FILL
            sheet[f'{col_letter}{row_idx}'].border = THIN_BORDER
            sheet[f'{col_letter}{row_idx}'].number_format = '#,##0.00_);(#,##0.00)'
    
    # Financial Projections
    sheet.merge_cells('A18:I18')
    sheet['A18'] = "FINANCIAL PROJECTIONS"
    sheet['A18'].font = Font(bold=True, color="FFFFFF")
    sheet['A18'].fill = GREEN_FILL
    sheet['A18'].alignment = Alignment(horizontal='center')
    
    # Headers for financial projections
    projection_headers = ["", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5", "Year 6", "Year 7"]
    for col_idx, header in enumerate(projection_headers):
        col_letter = get_column_letter(col_idx + 1)
        sheet[f'{col_letter}19'] = header
        sheet[f'{col_letter}19'].font = Font(bold=True)
        sheet[f'{col_letter}19'].fill = GREY_FILL
        sheet[f'{col_letter}19'].alignment = Alignment(horizontal='center')
        sheet[f'{col_letter}19'].border = THIN_BORDER
    
    # Projection metrics
    projection_metrics = [
        "Projected Revenue",
        "Projected EBITDA",
        "Projected Net Income",
        "Cash Flow Available for Debt Service",
        "Debt Service",
        "Projected DSCR",
        "Cumulative Cash Flow"
    ]
    
    for row_idx, metric in enumerate(projection_metrics, start=20):
        sheet[f'A{row_idx}'] = metric
        sheet[f'A{row_idx}'].font = Font(bold=True)
        sheet[f'A{row_idx}'].alignment = Alignment(horizontal='left')
        
        for col_idx in range(2, 9):
            col_letter = get_column_letter(col_idx)
            sheet[f'{col_letter}{row_idx}'].fill = LIGHT_BLUE_FILL
            sheet[f'{col_letter}{row_idx}'].border = THIN_BORDER
            sheet[f'{col_letter}{row_idx}'].number_format = '#,##0.00_);(#,##0.00)'
    
    # Source Leverage Loan Analysis
    sheet.merge_cells('A28:I28')
    sheet['A28'] = "SOURCE LEVERAGE LOAN ANALYSIS"
    sheet['A28'].font = Font(bold=True, color="FFFFFF")
    sheet['A28'].fill = GREEN_FILL
    sheet['A28'].alignment = Alignment(horizontal='center')
    
    loan_fields = [
        ("Loan Amount:", "B29", "C29:D29"),
        ("Interest Rate (%):", "B30", "C30:D30"),
        ("Term (Years):", "B31", "C31:D31"),
        ("Amortization Period:", "B32", "C32:D32"),
        ("Payment Frequency:", "B33", "C33:D33"),
        ("Annual Debt Service:", "B34", "C34:D34"),
        ("Loan-to-Value Ratio (%):", "B35", "C35:D35"),
        ("Collateral Description:", "B36", "C36:D36"),
        ("Loan Purpose:", "B37", "C37:D37")
    ]
    
    for label, label_cell, value_range in loan_fields:
        sheet[label_cell] = label
        sheet[label_cell].font = Font(bold=True)
        sheet[label_cell].alignment = Alignment(horizontal='right')
        sheet.merge_cells(value_range)
        value_cell = value_range.split(':')[0]
        sheet[value_cell].fill = LIGHT_BLUE_FILL
        sheet[value_cell].border = THIN_BORDER
    
    # Stress Testing
    sheet.merge_cells('A39:I39')
    sheet['A39'] = "STRESS TESTING SCENARIOS"
    sheet['A39'].font = Font(bold=True, color="FFFFFF")
    sheet['A39'].fill = GREEN_FILL
    sheet['A39'].alignment = Alignment(horizontal='center')
    
    # Stress test table headers
    stress_headers = ["Scenario", "Revenue Impact", "EBITDA Impact", "DSCR", "Break-Even Point", "Assessment"]
    
    for col_idx, header in enumerate(stress_headers):
        col_letter = get_column_letter(col_idx + 1)
        sheet[f'{col_letter}40'] = header
        sheet[f'{col_letter}40'].font = Font(bold=True)
        sheet[f'{col_letter}40'].fill = GREY_FILL
        sheet[f'{col_letter}40'].alignment = Alignment(horizontal='center')
        sheet[f'{col_letter}40'].border = THIN_BORDER
    
    # Stress test scenarios
    stress_scenarios = [
        "10% Revenue Decrease",
        "25% Revenue Decrease",
        "10% Cost Increase",
        "Interest Rate Increase (+2%)",
        "Combination Scenario"
    ]
    
    for row_idx, scenario in enumerate(stress_scenarios, start=41):
        sheet[f'A{row_idx}'] = scenario
        sheet[f'A{row_idx}'].font = Font(bold=True)
        
        for col_idx in range(1, 7):
            col_letter = get_column_letter(col_idx)
            sheet[f'{col_letter}{row_idx}'].fill = LIGHT_BLUE_FILL
            sheet[f'{col_letter}{row_idx}'].border = THIN_BORDER
    
    # Exit Strategy Analysis
    sheet.merge_cells('A47:I47')
    sheet['A47'] = "EXIT STRATEGY ANALYSIS"
    sheet['A47'].font = Font(bold=True, color="FFFFFF")
    sheet['A47'].fill = GREEN_FILL
    sheet['A47'].alignment = Alignment(horizontal='center')
    
    exit_fields = [
        ("Expected Exit Date:", "B48", "C48:D48"),
        ("Expected Exit Mechanism:", "B49", "C49:D49"),
        ("Projected Refinancing Terms:", "B50", "C50:D50"),
        ("Put/Call Options:", "B51", "C51:D51"),
        ("Exit Valuation Assumptions:", "B52", "C52:D52"),
        ("Exit Risk Assessment:", "B53", "C53:D53")
    ]
    
    for label, label_cell, value_range in exit_fields:
        sheet[label_cell] = label
        sheet[label_cell].font = Font(bold=True)
        sheet[label_cell].alignment = Alignment(horizontal='right')
        sheet.merge_cells(value_range)
        value_cell = value_range.split(':')[0]
        sheet[value_cell].fill = LIGHT_BLUE_FILL
        sheet[value_cell].border = THIN_BORDER
    
    # Financial Covenants
    sheet.merge_cells('F29:I29')
    sheet['F29'] = "FINANCIAL COVENANTS"
    sheet['F29'].font = Font(bold=True)
    sheet['F29'].fill = GREY_FILL
    sheet['F29'].alignment = Alignment(horizontal='center')
    
    covenant_fields = [
        ("Minimum DSCR:", "F30", "G30:I30"),
        ("Maximum LTV:", "F31", "G31:I31"),
        ("Minimum Working Capital:", "F32", "G32:I32"),
        ("Minimum Net Worth:", "F33", "G33:I33"),
        ("Maximum Debt-to-Equity:", "F34", "G34:I34"),
        ("Other Covenants:", "F35", "G35:I35")
    ]
    
    for label, label_cell, value_range in covenant_fields:
        sheet[label_cell] = label
        sheet[label_cell].font = Font(bold=True)
        sheet[label_cell].alignment = Alignment(horizontal='right')
        sheet.merge_cells(value_range)
        value_cell = value_range.split(':')[0]
        sheet[value_cell].fill = LIGHT_BLUE_FILL
        sheet[value_cell].border = THIN_BORDER
    
    # Overall Assessment
    sheet.merge_cells('A55:I55')
    sheet['A55'] = "FINANCIAL ANALYSIS ASSESSMENT & RECOMMENDATIONS"
    sheet['A55'].font = Font(bold=True, color="FFFFFF")
    sheet['A55'].fill = GREEN_FILL
    sheet['A55'].alignment = Alignment(horizontal='center')
    
    assessment_fields = [
        ("Historical Performance Assessment:", "B56", "C56:I56"),
        ("Projections Reasonableness:", "B57", "C57:I57"),
        ("Financial Strength Rating:", "B58", "C58:I58"),
        ("Debt Capacity Assessment:", "B59", "C59:I59"),
        ("Overall Financial Risk Rating:", "B60", "C60:I60")
    ]
    
    for label, label_cell, value_range in assessment_fields:
        sheet[label_cell] = label
        sheet[label_cell].font = Font(bold=True)
        sheet[label_cell].alignment = Alignment(horizontal='right')
        sheet.merge_cells(value_range)
        value_cell = value_range.split(':')[0]
        sheet[value_cell].fill = LIGHT_BLUE_FILL
        sheet[value_cell].border = THIN_BORDER
    
    # Comments section
    sheet.merge_cells('A62:I62')
    sheet['A62'] = "FINANCIAL ANALYSIS CONCLUSIONS & RECOMMENDATIONS"
    sheet['A62'].font = Font(bold=True)
    sheet['A62'].fill = GREY_FILL
    sheet['A62'].alignment = Alignment(horizontal='center')
    
    sheet.merge_cells('A63:I66')
    sheet['A63'].fill = LIGHT_BLUE_FILL
    sheet['A63'].alignment = Alignment(wrap_text=True, vertical='top')
    sheet['A63'].border = THIN_BORDER

def setup_reference_materials(wb, sheet):
    """Set up the Reference Materials worksheet for providing additional resources"""
    # Set column widths
    for col in range(1, 20):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 15
    
    # Title
    sheet.merge_cells('A1:I1')
    sheet['A1'] = "REFERENCE MATERIALS"
    sheet['A1'].font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['A1'].fill = GREEN_FILL
    
    # NMTC Program Overview
    sheet.merge_cells('A3:I3')
    sheet['A3'] = "NMTC PROGRAM OVERVIEW"
    sheet['A3'].font = Font(bold=True, color="FFFFFF")
    sheet['A3'].fill = GREEN_FILL
    sheet['A3'].alignment = Alignment(horizontal='center')
    
    nmtc_overview = """
    The New Markets Tax Credit (NMTC) program is a federal tax credit program that aims to encourage investment in low-income communities. 
    The program provides tax credits to investors who invest in qualified community development entities (CDEs), which in turn invest in qualified low-income community investments (QLICIs).
    
    KEY PROGRAM FEATURES:
    • Tax credits of up to 39% of investment
    • 7-year credit period
    • Investment must be made in a qualified CDE
    • CDE must invest in a QLICI
    • QLICI must be located in a low-income community
    """
    
    sheet.merge_cells('A4:I7')
    sheet['A4'] = nmtc_overview.strip()
    sheet['A4'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # NMTC Structure Diagram
    sheet.merge_cells('A9:I9')
    sheet['A9'] = "NMTC STRUCTURE DIAGRAM"
    sheet['A9'].font = Font(bold=True, color="FFFFFF")
    sheet['A9'].fill = GREEN_FILL
    sheet['A9'].alignment = Alignment(horizontal='center')
    
    # Try to insert an image if it exists, otherwise add a placeholder message
    try:
        if os.path.exists('nmtc_structure_diagram.png'):
            img = Image('nmtc_structure_diagram.png')
            img.width = 400
            img.height = 200
            sheet.add_image(img, 'A10')
        else:
            sheet.merge_cells('A10:E10')
            sheet['A10'] = "NMTC Structure Diagram (placeholder)"
            sheet['A10'].font = Font(bold=True)
            sheet['A10'].alignment = Alignment(horizontal='center')
            sheet.row_dimensions[10].height = 150
            
            sheet.merge_cells('A11:E14')
            sheet['A11'] = "Note: Insert your NMTC structure diagram here.\nTypical structure involves:\n1. Source Leverage Lender providing a loan to the Investment Fund\n2. Tax Credit Investor providing equity to the Investment Fund\n3. Investment Fund making a QEI in the CDE\n4. CDE making QLICI loans to the QALICB"
            sheet['A11'].font = Font(name='Arial', size=12)
            sheet['A11'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    except ImportError:
        # Handle case where openpyxl Image feature is not available
        sheet.merge_cells('A10:E10')
        sheet['A10'] = "NMTC Structure Diagram (placeholder)"
        sheet['A10'].font = Font(bold=True)
        sheet['A10'].alignment = Alignment(horizontal='center')
        
        sheet.merge_cells('A11:E14')
        sheet['A11'] = "Note: Insert your NMTC structure diagram here.\nTypical structure involves:\n1. Source Leverage Lender providing a loan to the Investment Fund\n2. Tax Credit Investor providing equity to the Investment Fund\n3. Investment Fund making a QEI in the CDE\n4. CDE making QLICI loans to the QALICB"
        sheet['A11'].font = Font(name='Arial', size=12)
        sheet['A11'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # NMTC Key Definitions
    sheet.merge_cells('A16:I16')
    sheet['A16'] = "NMTC KEY DEFINITIONS"
    sheet['A16'].font = Font(bold=True, color="FFFFFF")
    sheet['A16'].fill = GREEN_FILL
    sheet['A16'].alignment = Alignment(horizontal='center')
    
    # Set up definitions table
    sheet.merge_cells('A17:B17')
    sheet['A17'] = "Term"
    sheet['A17'].font = Font(bold=True)
    sheet['A17'].fill = GREY_FILL
    sheet['A17'].alignment = Alignment(horizontal='center')
    sheet['A17'].border = THIN_BORDER
    
    sheet.merge_cells('C17:E17')
    sheet['C17'] = "Definition"
    sheet['C17'].font = Font(bold=True)
    sheet['C17'].fill = GREY_FILL
    sheet['C17'].alignment = Alignment(horizontal='center')
    sheet['C17'].border = THIN_BORDER
    
    # Add key NMTC definitions
    definitions = [
        ("NMTC", "New Markets Tax Credit - A federal program that incentivizes community development and economic growth through tax credits that attract private investment to distressed communities."),
        ("CDE", "Community Development Entity - A domestic corporation or partnership that is an intermediary vehicle for the provision of loans, investments, or financial counseling in low-income communities."),
        ("CDFI", "Community Development Financial Institution - A specialized financial institution that works in markets that are underserved by traditional financial institutions."),
        ("QEI", "Qualified Equity Investment - An equity investment in a CDE that meets certain requirements specified in the NMTC Program."),
        ("QLICI", "Qualified Low-Income Community Investment - An investment by a CDE in, or loan to, a qualified active low-income community business."),
        ("QALICB", "Qualified Active Low-Income Community Business - A business that meets specific requirements and is located in a low-income community."),
        ("QCT", "Qualified Census Tract - A census tract with a poverty rate of at least 20% or median family income below 80% of the area median."),
        ("LIC", "Low-Income Community - Generally, a census tract with a poverty rate of at least 20% or median family income below 80% of the area median."),
        ("Leverage Lender", "Entity providing loan capital to the Investment Fund that combines with tax credit equity to fund the QEI."),
        ("Source Leverage Lender", "Original lender providing loan capital for the leverage loan structure."),
        ("Tax Credit Investor", "Entity that invests in the Investment Fund in exchange for the tax credits generated by the transaction."),
        ("Investment Fund", "Special purpose entity that receives the leverage loan and equity investment, and makes the QEI in the CDE."),
        ("Put/Call Option", "A mechanism for the investor to exit the transaction after the 7-year compliance period."),
        ("Recapture", "The IRS can recapture tax credits if certain conditions aren't met during the 7-year compliance period.")
    ]
    
    for i, (term, definition) in enumerate(definitions):
        row = 18 + i
        
        # Term column
        sheet.merge_cells(f'A{row}:B{row}')
        sheet[f'A{row}'] = term
        sheet[f'A{row}'].font = Font(bold=True)
        sheet[f'A{row}'].border = THIN_BORDER
        sheet[f'A{row}'].alignment = Alignment(horizontal='center')
        
        # Definition column
        sheet.merge_cells(f'C{row}:E{row}')
        sheet[f'C{row}'] = definition
        sheet[f'C{row}'].font = Font(name='Arial', size=12)
        sheet[f'C{row}'].border = THIN_BORDER
        sheet[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Set row height to accommodate wrapped text
        sheet.row_dimensions[row].height = 45
    
    # NMTC Compliance Requirements
    sheet.merge_cells('A30:I30')
    sheet['A30'] = "NMTC COMPLIANCE REQUIREMENTS"
    sheet['A30'].font = Font(bold=True, color="FFFFFF")
    sheet['A30'].fill = GREEN_FILL
    sheet['A30'].alignment = Alignment(horizontal='center')
    
    # Set up compliance requirements text
    sheet.merge_cells('A31:I40')
    compliance_text = """
    Key NMTC Compliance Requirements:
    
    1. 7-Year Compliance Period: The investor must maintain the QEI for a 7-year period to claim all tax credits.
    
    2. Substantially-All Test: At least 85% of QEI proceeds must be invested in QLICIs.
    
    3. Qualified Business Criteria: The QALICB must meet specific requirements, including:
       - At least 50% of gross income derived from active business in LIC
       - At least 40% of tangible property located within LIC
       - At least 40% of services performed in LIC
       - Less than 5% of assets in nonqualified financial property
       - No "sin businesses" (e.g., gambling, liquor stores, etc.)
    
    4. Recapture Events to Avoid:
       - CDE ceases to qualify as a CDE
       - The QEI fails the "substantially-all" requirement
       - The QEI is redeemed or otherwise cashed out
    
    5. Reporting Requirements:
       - Annual reporting to CDFI Fund
       - IRS Form 8874-A for investors
    """
    sheet['A31'] = compliance_text
    sheet['A31'].font = Font(name='Arial', size=12)
    sheet['A31'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Reference Checklist for Source Leverage Lenders
    sheet.merge_cells('A42:I42')
    sheet['A42'] = "SOURCE LEVERAGE LENDER CHECKLIST SUMMARY"
    sheet['A42'].font = Font(bold=True, color="FFFFFF")
    sheet['A42'].fill = GREEN_FILL
    sheet['A42'].alignment = Alignment(horizontal='center')
    
    # Set up lender checklist text
    sheet.merge_cells('A43:I55')
    lender_text = """
    Key Considerations for Source Leverage Lenders:
    
    1. Credit Analysis:
       - Analyze QALICB financials and business plan
       - Assess project feasibility and market demand
       - Evaluate project sources and uses
       - Analyze debt service coverage and loan-to-value ratios
    
    2. Security Position:
       - Analyze Investment Fund assets and pledges
       - Review forbearance agreements with CDE
       - Understand structural protections and limitations
       
    3. NMTC Structure Analysis:
       - Review flow of funds and fee structure
       - Analyze tax credit equity pricing and QEI amount
       - Verify eligibility of project and QALICB
       - Review CDE track record and allocation availability
       
    4. Exit Strategy:
       - Analyze unwinding options at end of compliance period
       - Review put/call structures and pricing
       - Assess refinancing options and timing
       
    5. Risk Mitigation:
       - Identify key project and compliance risks
       - Review guaranties and credit enhancements
       - Develop monitoring plan for construction and compliance
       - Establish reserves for debt service and other contingencies
    """
    sheet['A43'] = lender_text
    sheet['A43'].font = Font(name='Arial', size=12)
    sheet['A43'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # NMTC Resources
    sheet.merge_cells('A57:I57')
    sheet['A57'] = "NMTC RESOURCES"
    sheet['A57'].font = Font(bold=True, color="FFFFFF")
    sheet['A57'].fill = GREEN_FILL
    sheet['A57'].alignment = Alignment(horizontal='center')
    
    # Set up resources table
    resource_headers = ["Resource", "Website/Contact"]
    
    sheet.merge_cells('A58:C58')
    sheet['A58'] = resource_headers[0]
    sheet['A58'].font = Font(bold=True)
    sheet['A58'].fill = GREY_FILL
    sheet['A58'].alignment = Alignment(horizontal='center')
    sheet['A58'].border = THIN_BORDER
    
    sheet.merge_cells('D58:E58')
    sheet['D58'] = resource_headers[1]
    sheet['D58'].font = Font(bold=True)
    sheet['D58'].fill = GREY_FILL
    sheet['D58'].alignment = Alignment(horizontal='center')
    sheet['D58'].border = THIN_BORDER
    
    # Add resources
    resources = [
        ("CDFI Fund NMTC Program Page", "www.cdfifund.gov/programs-training/programs/new-markets-tax-credit"),
        ("Novogradac & Company", "www.novoco.com/resource-centers/new-markets-tax-credits"),
        ("New Markets Tax Credit Coalition", "www.nmtccoalition.org"),
        ("IRS NMTC Guidance", "www.irs.gov/businesses/new-markets-tax-credit"),
        ("Clarity Impact Finance", "www.clarityimpactfinance.com"),
        ("Community Development Financial Institutions Fund", "www.cdfifund.gov"),
        ("NMTC Mapping Tool", "www.cims.cdfifund.gov/preparation/?config=config_nmtc.xml")
    ]
    
    for i, (resource, website) in enumerate(resources):
        row = 59 + i
        
        # Resource column
        sheet.merge_cells(f'A{row}:C{row}')
        sheet[f'A{row}'] = resource
        sheet[f'A{row}'].font = Font(bold=True)
        sheet[f'A{row}'].border = THIN_BORDER
        sheet[f'A{row}'].alignment = Alignment(horizontal='center')
        
        # Website column
        sheet.merge_cells(f'D{row}:E{row}')
        sheet[f'D{row}'] = website
        sheet[f'D{row}'].font = Font(name='Arial', size=12)
        sheet[f'D{row}'].border = THIN_BORDER
        sheet[f'D{row}'].alignment = Alignment(horizontal='center')
    
    # Add Clarity Impact Finance branding
    sheet.merge_cells('A67:E67')
    sheet['A67'] = "Prepared by Clarity Impact Finance"
    sheet['A67'].font = Font(name='Arial', size=10, italic=True)
    sheet['A67'].alignment = Alignment(horizontal='center')
    
    sheet.merge_cells('A68:E68')
    sheet['A68'] = "www.clarityimpactfinance.com | contact@clarityimpactfinance.com"
    sheet['A68'].font = Font(name='Arial', size=10, italic=True)
    sheet['A68'].alignment = Alignment(horizontal='center')

def setup_structure_chart(wb, sheet):
    """Set up the NMTC Structure Chart tab"""
    
    # Title and instructions
    sheet.merge_cells('A1:I1')
    sheet['A1'] = "NMTC STRUCTURE CHART"
    sheet['A1'].font = Font(name='Arial', size=14, bold=True)
    sheet['A1'].alignment = Alignment(horizontal='center')
    
    # Add instructions
    sheet.merge_cells('A3:I3')
    sheet['A3'] = "Visual representation of the NMTC transaction structure"
    sheet['A3'].font = Font(name='Arial', size=10, italic=True)
    sheet['A3'].alignment = Alignment(horizontal='center')
    
    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        sheet.column_dimensions[col].width = 15
    
    # Structure chart explanation
    sheet.merge_cells('A5:I5')
    sheet['A5'] = "TYPICAL NMTC STRUCTURE COMPONENTS"
    sheet['A5'].font = Font(name='Arial', size=12, bold=True)
    sheet['A5'].alignment = Alignment(horizontal='center')
    
    components = [
        ("Investor", "Tax credit purchaser (typically a bank or other financial institution)"),
        ("Investment Fund", "Entity that receives equity from Investor and debt from Leverage Lender"),
        ("CDE(s)", "Community Development Entity that receives QEI from Investment Fund"),
        ("QALICB", "Qualified Active Low-Income Community Business (borrower)"),
        ("Leverage Lender(s)", "Provides loan to Investment Fund (often includes the sponsor)"),
        ("Source Leverage Lender", "Third-party lender providing funds to the leverage lender"),
        ("Flow of Funds", "Capital moves from Investor/Lenders → Investment Fund → CDE → QALICB")
    ]
    
    # Add component explanations
    for i, (component, explanation) in enumerate(components):
        row = 7 + i
        
        # Component column
        sheet.merge_cells(f'A{row}:B{row}')
        sheet[f'A{row}'] = component
        sheet[f'A{row}'].font = Font(bold=True)
        sheet[f'A{row}'].border = THIN_BORDER
        sheet[f'A{row}'].alignment = Alignment(horizontal='center')
        
        # Explanation column
        sheet.merge_cells(f'C{row}:E{row}')
        sheet[f'C{row}'] = explanation
        sheet[f'C{row}'].font = Font(name='Arial', size=12)
        sheet[f'C{row}'].border = THIN_BORDER
        sheet[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Set row height to accommodate wrapped text
        sheet.row_dimensions[row].height = 45
    
    # Add placeholder for structure chart
    sheet.merge_cells('B18:H30')
    sheet['B18'] = "STRUCTURE CHART PLACEHOLDER"
    sheet['B18'].font = Font(name='Arial', size=14, bold=True)
    sheet['B18'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Add a border around the chart placeholder
    border = Border(
        left=Side(style='medium'), 
        right=Side(style='medium'), 
        top=Side(style='medium'), 
        bottom=Side(style='medium')
    )
    
    for row in range(18, 31):
        for col in range(2, 9):
            sheet.cell(row=row, column=col).border = border
    
    # Add note at bottom
    sheet.merge_cells('A33:I33')
    sheet['A33'] = "Note: Customize this structure chart to reflect the specific transaction structure of your deal."
    sheet['A33'].font = Font(name='Arial', size=10, italic=True)
    sheet['A33'].alignment = Alignment(horizontal='center')

def setup_input_tab(wb, sheet):
    """Set up the Input tab for deal parameters and calculations"""
    
    # Column widths
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 20
    
    # Title
    sheet.merge_cells('A1:H1')
    sheet['A1'] = "NMTC DEAL PARAMETERS"
    sheet['A1'].font = Font(name='Arial', size=14, bold=True)
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet['A1'].fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    
    # Project Information section
    sheet.merge_cells('A3:H3')
    sheet['A3'] = "PROJECT INFORMATION"
    sheet['A3'].font = Font(name='Arial', size=12, bold=True)
    sheet['A3'].fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    # Field labels and input cells
    fields = [
        ("Project Name:", "B6", None),
        ("Closing Date:", "B7", None),
        ("Location:", "B8", None),
        ("Total Project Cost:", "B9", "_($* #,##0.00_);_($* (#,##0.00);_($* \"-\"??_);_(@_)"),
        ("Leverage Loan Amount:", "B10", "_($* #,##0.00_);_($* (#,##0.00);_($* \"-\"??_);_(@_)"),
        ("Tax Credit Investor:", "B11", None),
        ("Tax Credit Price (%):", "B12", "0.00%")
    ]
    
    for i, (label, cell, format) in enumerate(fields):
        sheet[f'A{i+5}'] = label
        sheet[f'A{i+5}'].font = Font(name='Arial', size=10, bold=True)
        sheet[f'A{i+5}'].alignment = Alignment(horizontal='right')
        
        # Format cells based on their expected content
        if format:
            sheet[cell].number_format = format
    
    # CDE Allocation section
    sheet.merge_cells('A15:H15')
    sheet['A15'] = "CDE ALLOCATION INPUT"
    sheet['A15'].font = Font(name='Arial', size=12, bold=True)
    sheet['A15'].fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    sheet['A17'] = "Number of CDEs:"
    sheet['A17'].font = Font(name='Arial', size=10, bold=True)
    sheet['A17'].alignment = Alignment(horizontal='right')
    
    # Create data validation for number of CDEs (1-5)
    dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5")
    dv.error = "Enter a number between 1 and 5"
    dv.errorTitle = "Invalid Input"
    dv.prompt = "Enter the number of CDEs (1-5)"
    dv.promptTitle = "Number of CDEs"
    sheet.add_data_validation(dv)
    dv.add('B17')
    
    # CDE Table headers
    table_headers = [("CDE Name", "A"), ("Allocation Amount", "B"), ("Fee (%)", "C"), ("Fee Amount", "D"), ("QLICI Amount", "E")]
    for header, col in table_headers:
        cell = f'{col}19'
        sheet[cell] = header
        sheet[cell].font = Font(name='Arial', size=10, bold=True)
        sheet[cell].fill = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid")
        sheet[cell].alignment = Alignment(horizontal='center')
    
    # Format the CDE allocation table rows
    for row in range(20, 25):  # Allow for up to 5 CDEs
        # Format allocation amount cells as currency
        sheet[f'B{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        
        # Format fee percentage cells as percentage
        sheet[f'C{row}'].number_format = '0.00%'
        
        # Add formulas for fee amount and QLICI amount
        sheet[f'D{row}'] = f'=B{row}*C{row}'
        sheet[f'D{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        
        sheet[f'E{row}'] = f'=B{row}-D{row}'
        sheet[f'E{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    # Total rows with formulas
    sheet['A25'] = "TOTAL"
    sheet['A25'].font = Font(name='Arial', size=10, bold=True)
    
    sheet['B25'] = "=SUM(B20:B24)"
    sheet['B25'].font = Font(name='Arial', size=10, bold=True)
    sheet['B25'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    sheet['D25'] = "=SUM(D20:D24)"
    sheet['D25'].font = Font(name='Arial', size=10, bold=True)
    sheet['D25'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    sheet['E25'] = "=SUM(E20:E24)"
    sheet['E25'].font = Font(name='Arial', size=10, bold=True)
    sheet['E25'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    # Sources and Uses section
    sheet.merge_cells('A30:H30')
    sheet['A30'] = "SOURCES & USES"
    sheet['A30'].font = Font(name='Arial', size=12, bold=True)
    sheet['A30'].fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    # Sources section
    sheet.merge_cells('A32:C32')
    sheet['A32'] = "SOURCES"
    sheet['A32'].font = Font(name='Arial', size=11, bold=True)
    sheet['A32'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    sheet['A32'].alignment = Alignment(horizontal='center')
    
    # Sources headers
    sources_headers = [("Source", "A"), ("Amount", "B"), ("% of Total", "C")]
    for header, col in sources_headers:
        cell = f'{col}33'
        sheet[cell] = header
        sheet[cell].font = Font(name='Arial', size=10, bold=True)
        sheet[cell].fill = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid")
        sheet[cell].alignment = Alignment(horizontal='center')
    
    # Sources rows with formulas
    sources = [
        ("Leverage Loan", "=B10", "=B34/B36"),
        ("Tax Credit Equity", "=B25*(B12)", "=B35/B36"),
        ("Total Sources", "=SUM(B34:B35)", "")
    ]
    
    for i, (source, formula, percent) in enumerate(sources, start=34):
        sheet[f'A{i}'] = source
        sheet[f'B{i}'] = formula
        sheet[f'B{i}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        
        if percent:
            sheet[f'C{i}'] = percent
            sheet[f'C{i}'].number_format = '0.00%'
        
        if source == "Total Sources":
            for col in ['A', 'B', 'C']:
                sheet[f'{col}{i}'].font = Font(name='Arial', size=10, bold=True)
    
    # Uses section
    sheet.merge_cells('A38:C38')
    sheet['A38'] = "USES"
    sheet['A38'].font = Font(name='Arial', size=11, bold=True)
    sheet['A38'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    sheet['A38'].alignment = Alignment(horizontal='center')
    
    # Uses headers
    uses_headers = [("Use", "A"), ("Amount", "B"), ("% of Total", "C")]
    for header, col in uses_headers:
        cell = f'{col}39'
        sheet[cell] = header
        sheet[cell].font = Font(name='Arial', size=10, bold=True)
        sheet[cell].fill = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid")
        sheet[cell].alignment = Alignment(horizontal='center')
    
    # Uses rows with formulas
    uses = [
        ("QLICI Loan to QALICB", "=E25", "=B40/B43"),
        ("CDE Fees", "=D25", "=B41/B43"),
        ("Other Closing Costs", "=0", "=B42/B43"),
        ("Total Uses", "=SUM(B40:B42)", "")
    ]
    
    for i, (use, formula, percent) in enumerate(uses, start=40):
        sheet[f'A{i}'] = use
        sheet[f'B{i}'] = formula
        sheet[f'B{i}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        
        if percent:
            sheet[f'C{i}'] = percent
            sheet[f'C{i}'].number_format = '0.00%'
        
        if use == "Total Uses":
            for col in ['A', 'B', 'C']:
                sheet[f'{col}{i}'].font = Font(name='Arial', size=10, bold=True)
    
    # Add validation check
    sheet.merge_cells('A46:D46')
    sheet['A46'] = "VALIDATION: SOURCES = USES"
    sheet['A46'].font = Font(name='Arial', size=11, bold=True)
    
    sheet.merge_cells('A47:D47')
    sheet['A47'] = "=IF(ABS(B36-B43)<0.01,\"BALANCED ✓\",\"ERROR: Sources and Uses do not match!\")"
    sheet['A47'].font = Font(name='Arial', size=10, bold=True)
    sheet['A47'].alignment = Alignment(horizontal='center')
    
    # Add note at bottom
    sheet.merge_cells('A50:H50')
    sheet['A50'] = "Note: Enter CDE allocation amounts and fees to automatically calculate sources and uses."
    sheet['A50'].font = Font(name='Arial', size=10, italic=True)
    sheet['A50'].alignment = Alignment(horizontal='center')

def add_formulas(wb):
    """Add Excel formulas to the workbook for calculations and dynamic updates"""
    dashboard = wb['Dashboard']
    
    # Update named ranges to use the newer method
    # Instead of using create_named_range, we'll define names directly
    from openpyxl.workbook.defined_name import DefinedName
    
    # Create defined names for risk assessments
    structure_risk = DefinedName('StructureRisk', attr_text="Dashboard!$H$8")
    financial_risk = DefinedName('FinancialRisk', attr_text="Dashboard!$H$9")
    
    # Add defined names to the workbook
    wb.defined_names.add(structure_risk)
    wb.defined_names.add(financial_risk)
    
    # Overall risk assessment formula - we need to make sure this cell is not merged
    dashboard['H10'] = '=IF(COUNTIF(H8:H9,"High")>0,"HIGH RISK",IF(COUNTIF(H8:H9,"Medium")>1,"MEDIUM RISK","LOW RISK"))'
    
    # Example data toggle formula
    dashboard['B4'] = 'YES'  # Default to showing examples

def write_to_cell(sheet, cell_address, value, number_format=None):
    """
    Write to a cell, safely handling merged cells by writing to the top-left cell
    
    Args:
        sheet: The worksheet to write to
        cell_address: The cell address (e.g., 'A1')
        value: The value to write
        number_format: Optional number format to apply to the cell
    """
    # Convert string values with currency symbols to numeric values
    if isinstance(value, str):
        # Remove currency symbols and commas for numeric conversion
        if value.startswith("$") and any(c.isdigit() for c in value):
            try:
                # Remove $ and commas, then convert to float
                cleaned_value = value.replace("$", "").replace(",", "")
                value = float(cleaned_value)
            except ValueError:
                # If conversion fails, keep original string value
                pass
        # Convert percentage strings to floats
        elif value.endswith("%") and any(c.isdigit() for c in value):
            try:
                # Remove % and convert to float decimal
                cleaned_value = value.replace("%", "")
                value = float(cleaned_value) / 100
            except ValueError:
                # If conversion fails, keep original string value
                pass
    
    # Check if the cell is part of a merged range
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        cell_col, cell_row = range_boundaries(cell_address + ":" + cell_address)[0:2]
        
        # If the cell is in a merged range, write to the top-left cell instead
        if (min_row <= cell_row <= max_row and min_col <= cell_col <= max_col):
            top_left = get_column_letter(min_col) + str(min_row)
            sheet[top_left] = value
            
            # Apply number format if provided
            if number_format and isinstance(value, (int, float)):
                sheet[top_left].number_format = number_format
            elif isinstance(value, (int, float)):
                # Default number formats based on value type
                if isinstance(value, int):
                    sheet[top_left].number_format = '#,##0'
                else:  # float
                    sheet[top_left].number_format = '#,##0.00'
            return
    
    # If not in a merged range, write directly
    sheet[cell_address] = value
    
    # Apply number format if provided
    if number_format and isinstance(value, (int, float)):
        sheet[cell_address].number_format = number_format
    elif isinstance(value, (int, float)):
        # Default number formats based on value type
        if isinstance(value, int):
            sheet[cell_address].number_format = '#,##0'
        else:  # float
            sheet[cell_address].number_format = '#,##0.00'

def add_example_data(wb, include_examples=True):
    """Add example data to the workbook if requested"""
    if not include_examples:
        return
    
    # Currency format
    currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    # Percentage format
    percentage_format = '0.00%'
    
    # Dashboard example data
    dashboard = wb['Dashboard']
    write_to_cell(dashboard, 'B8', "Midwest Community Health Center")  # Project Name
    write_to_cell(dashboard, 'B9', "Chicago, IL")  # Location
    write_to_cell(dashboard, 'B10', 25000000, currency_format)  # Total Project Cost
    write_to_cell(dashboard, 'B11', 15000000, currency_format)  # NMTC Allocation
    write_to_cell(dashboard, 'B12', 9000000, currency_format)  # Leverage Loan
    write_to_cell(dashboard, 'B13', "First National Bank")  # Tax Credit Investor
    write_to_cell(dashboard, 'B14', "Midwest Regional Bank")  # Source Leverage Lender
    write_to_cell(dashboard, 'C15', "January 15, 2024")  # Closing Date
    
    # Status section
    write_to_cell(dashboard, 'H7', "Complete")  # Deal Structure Review
    write_to_cell(dashboard, 'H8', "Medium")  # Financial Analysis
    write_to_cell(dashboard, 'H9', "High")  # Overall Recommendation
    
    # Deal Structure example data
    deal_structure = wb['Deal_Structure']
    write_to_cell(deal_structure, 'C4', "Leveraged Loan Structure")
    write_to_cell(deal_structure, 'C5', 15000000, currency_format)
    write_to_cell(deal_structure, 'C6', 5850000, currency_format)
    write_to_cell(deal_structure, 'C7', 15000000, currency_format)
    write_to_cell(deal_structure, 'C8', 2281500, currency_format)
    write_to_cell(deal_structure, 'C9', 12718500, currency_format)
    write_to_cell(deal_structure, 'C10', "2.17:1")
    write_to_cell(deal_structure, 'C11', "March 15, 2024")
    write_to_cell(deal_structure, 'C12', "April 30, 2031")
    
    write_to_cell(deal_structure, 'C15', "Midwest Regional CDE")
    write_to_cell(deal_structure, 'C16', "12ABC-123456")
    write_to_cell(deal_structure, 'C17', "Midwest Development Partners")
    write_to_cell(deal_structure, 'C18', 75000000, currency_format)
    write_to_cell(deal_structure, 'C19', 15000000, currency_format)
    write_to_cell(deal_structure, 'C20', 2023)
    
    write_to_cell(deal_structure, 'C25', "First National Bank")
    write_to_cell(deal_structure, 'C26', "Sarah Johnson")
    write_to_cell(deal_structure, 'C27', 0.82, percentage_format)
    write_to_cell(deal_structure, 'C28', 2281500, currency_format)
    
    # Financial Analysis example data
    financial = wb['Financial_Analysis']
    write_to_cell(financial, 'C4', "Positive - Meets Expectations")
    write_to_cell(financial, 'C5', "Healthcare services in LIC")
    write_to_cell(financial, 'C6', "15 years of stable operations")
    
    write_to_cell(financial, 'C14', 24500000, currency_format)
    write_to_cell(financial, 'C15', 25900000, currency_format)
    write_to_cell(financial, 'C16', 27400000, currency_format)
    write_to_cell(financial, 'C17', 0.057, percentage_format)
    write_to_cell(financial, 'C18', 0.065, percentage_format)
    
    write_to_cell(financial, 'C24', 15800000, currency_format)
    write_to_cell(financial, 'C25', 16200000, currency_format)
    write_to_cell(financial, 'C26', 16900000, currency_format)
    write_to_cell(financial, 'C27', 0.028, percentage_format)
    write_to_cell(financial, 'C28', 0.041, percentage_format)
    
    write_to_cell(financial, 'C34', 8700000, currency_format)
    write_to_cell(financial, 'C35', 9700000, currency_format)
    write_to_cell(financial, 'C36', 10500000, currency_format)
    write_to_cell(financial, 'C37', 0.115, percentage_format)
    write_to_cell(financial, 'C38', 0.082, percentage_format)
    
    write_to_cell(financial, 'C43', 12718500, currency_format)
    write_to_cell(financial, 'C44', "Interest-only for 7 years")
    write_to_cell(financial, 'C45', 1.25)
    write_to_cell(financial, 'C46', 1.35)
    write_to_cell(financial, 'C47', 1.48)
    
    # Structure Chart example data
    structure_chart = wb['Structure_Chart']
    for i, component in enumerate(["First National Bank", "Midwest Investment Fund, LLC", "Midwest Regional CDE", 
                                   "Community Healthcare QALICB, Inc.", "Community Healthcare, Inc.", 
                                   "Midwest Regional Bank", "See Reference Materials Tab for Full Flow of Funds"]):
        write_to_cell(structure_chart, f'A{i+7}', component)
    
    # Input Tab example data
    input_tab = wb['Input']
    write_to_cell(input_tab, 'B6', "Midwest Community Health Center")
    write_to_cell(input_tab, 'B7', "1/15/2024")
    write_to_cell(input_tab, 'B8', "Chicago, IL")
    write_to_cell(input_tab, 'B9', 25000000, currency_format)
    write_to_cell(input_tab, 'B10', 9000000, currency_format)
    write_to_cell(input_tab, 'B11', "First National Bank")
    write_to_cell(input_tab, 'B12', 0.39, percentage_format)
    
    # CDE allocation examples
    write_to_cell(input_tab, 'B17', 2)
    
    write_to_cell(input_tab, 'A20', "Midwest Regional CDE")
    write_to_cell(input_tab, 'B20', 10000000, currency_format)
    write_to_cell(input_tab, 'C20', 0.03, percentage_format)
    
    write_to_cell(input_tab, 'A21', "Urban Development CDE")
    write_to_cell(input_tab, 'B21', 5000000, currency_format)
    write_to_cell(input_tab, 'C21', 0.0325, percentage_format)

def add_validation_and_formulas(wb):
    """Add data validation and formulas to the workbook"""
    # Add data validation for dropdown fields
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Dashboard
    dashboard = wb['Dashboard']
    
    # Add dropdown for showing example data
    dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
    dv.error = "Please select YES or NO"
    dv.errorTitle = "Invalid Input"
    dv.prompt = "Select YES to include example data or NO to clear examples"
    dv.promptTitle = "Example Data"
    dashboard.add_data_validation(dv)
    dv.add('B4')
    
    # Add dropdown for status fields
    status_dv = DataValidation(type="list", formula1='"Complete,In Progress,Not Started,High,Medium,Low"', allow_blank=True)
    status_dv.error = "Please select a valid status"
    status_dv.errorTitle = "Invalid Input"
    status_dv.prompt = "Select status"
    status_dv.promptTitle = "Status"
    dashboard.add_data_validation(status_dv)
    
    # Apply to status cells - use a list not a set
    status_cells = ['H7', 'H8', 'H9']
    for cell in status_cells:
        status_dv.add(cell)
    
    # Deal Structure
    deal_structure = wb['Deal_Structure']
    
    # Financial Analysis
    financial = wb['Financial_Analysis']
    
    # Add formulas for dashboard
    dashboard['H10'] = '=IF(OR(H8="High",H9="High"),"HIGH",IF(COUNTIF(H8:H9,"Medium")>0,"MEDIUM","LOW"))'
    
    # Input tab formulas and validation
    input_tab = wb['Input']
    
    # Create data validation for CDE inputs
    cde_count_dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5")
    cde_count_dv.error = "Enter a number between 1 and 5"
    cde_count_dv.errorTitle = "Invalid Input"
    cde_count_dv.prompt = "Enter the number of CDEs (1-5)"
    cde_count_dv.promptTitle = "Number of CDEs"
    input_tab.add_data_validation(cde_count_dv)
    cde_count_dv.add('B17')
    
    # Date validation (optional)
    date_dv = DataValidation(type="date")
    date_dv.error = "Please enter a valid date"
    date_dv.errorTitle = "Invalid Date"
    date_dv.prompt = "Enter date in MM/DD/YYYY format"
    date_dv.promptTitle = "Date Input"
    input_tab.add_data_validation(date_dv)
    date_dv.add('B7')  # Closing date field
    
    # Percentage validation
    pct_dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="1")
    pct_dv.error = "Please enter a percentage between 0% and 100%"
    pct_dv.errorTitle = "Invalid Percentage"
    pct_dv.prompt = "Enter percentage (e.g., 0.39 for 39%)"
    pct_dv.promptTitle = "Percentage Input"
    input_tab.add_data_validation(pct_dv)
    pct_dv.add('B12')  # Tax Credit Price
    
    # Add validation for CDE fees
    for row in range(20, 25):
        pct_dv.add(f'C{row}')  # CDE fee percentages
        
    # Ensure validation for numbers
    for row in range(20, 25):
        # Ensure values are numeric in these cells
        input_tab[f'B{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        input_tab[f'C{row}'].number_format = '0.00%'
        input_tab[f'D{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        input_tab[f'E{row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    # Make sure dollar amounts are formatted correctly - use a list not a set
    monetary_cells = ['B9', 'B10']
    for cell in monetary_cells:
        input_tab[cell].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        
    # Overall risk assessment formula
    dashboard['H10'].font = Font(name='Arial', size=10, bold=True)
    
    # Add Clarity Impact Finance branding to all sheets
    # Company colors - green and orange
    green_fill = PatternFill(start_color="00A651", end_color="00A651", fill_type="solid")
    orange_fill = PatternFill(start_color="F7941D", end_color="F7941D", fill_type="solid")
    
    # Use a list for sheet names instead of wb.sheetnames which may be a set
    sheet_names = list(wb.sheetnames)
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        # Find the last row with content plus a few rows for spacing
        max_row = sheet.max_row + 3
        
        # Add branding at the bottom of each sheet
        sheet.merge_cells(f'A{max_row}:F{max_row}')
        cell = sheet[f'A{max_row}']
        cell.value = "Prepared by Clarity Impact Finance"
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Add contact info
        sheet.merge_cells(f'A{max_row+1}:F{max_row+1}')
        contact_cell = sheet[f'A{max_row+1}']
        contact_cell.value = "contact@clarityimpactfinance.com"
        contact_cell.font = Font(name='Arial', size=9)
        contact_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Add a stylish green line above branding
        for col in range(1, 7):  # Columns A-F
            border_cell = sheet.cell(row=max_row-1, column=col)
            border_cell.border = Border(bottom=Side(style='medium', color='00A651'))
    
# Main execution block
if __name__ == "__main__":
    try:
        import traceback
        # Create NMTC Leverage Lender Checklist workbook
        print("Creating NMTC Leverage Lender Underwriting Checklist...")
        wb = create_workbook()
        print("Workbook created successfully")
        
        try:
            # Add data validation and formulas
            print("Adding data validation and formulas...")
            add_validation_and_formulas(wb)
            print("Data validation and formulas added successfully")
        except Exception as validation_error:
            print(f"Error in add_validation_and_formulas: {str(validation_error)}")
            traceback.print_exc()
            raise
        
        try:
            # Add example data (defaults to YES)
            print("Adding example data...")
            add_example_data(wb, include_examples=True)
            print("Example data added successfully")
        except Exception as example_error:
            print(f"Error in add_example_data: {str(example_error)}")
            traceback.print_exc()
            raise
        
        # Save the file
        filename = "NMTC_Leverage_Lender_Underwriting_Checklist.xlsx"
        print(f"Saving to {filename}...")
        wb.save(filename)
        
        # Get the full path to the file
        file_path = os.path.abspath(filename)
        print(f"Excel file created successfully: {file_path}")
        
    except Exception as e:
        print(f"Error creating Excel file: {str(e)}")
        traceback.print_exc()
        sys.exit(1)
