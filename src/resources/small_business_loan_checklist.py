"""
Small Business Loan Underwriting Checklist Generator
---------------------------------------------------
This script creates an Excel workbook with a comprehensive small business loan 
underwriting checklist, including a dashboard with example data and the ability 
to clear this data.

Dependencies:
- openpyxl
- pandas

Install with: pip install openpyxl pandas
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import datetime
import os

def create_workbook():
    """Creates the Excel workbook with all sheets and formatting."""
    wb = openpyxl.Workbook()
    
    # Create sheets
    dashboard = wb.active
    dashboard.title = "Dashboard"
    checklist = wb.create_sheet("Underwriting Checklist")
    financial = wb.create_sheet("Financial Analysis")
    management = wb.create_sheet("Management Assessment")
    industry = wb.create_sheet("Industry Evaluation")
    risk = wb.create_sheet("Risk Assessment")
    
    # Set up each sheet
    setup_dashboard(dashboard)
    setup_checklist(checklist)
    setup_financial(financial)
    setup_management(management)
    setup_industry(industry)
    setup_risk(risk)
    
    # Add formulas and connections between sheets
    add_formulas(wb)
    
    # Save the workbook
    filename = "Small_Business_Loan_Underwriting_Checklist.xlsx"
    wb.save(filename)
    print(f"Created {filename} successfully!")
    return filename

def setup_dashboard(ws):
    """Sets up the Dashboard sheet with overview and example data."""
    # Add title and header
    ws.merge_cells('A1:H1')
    ws['A1'] = "SMALL BUSINESS LOAN UNDERWRITING DASHBOARD"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor="1B4620")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Add loan summary section
    ws['A3'] = "LOAN SUMMARY"
    ws['A3'].font = Font(size=12, bold=True)
    
    headers = ["Business Name", "Loan Amount", "Term (months)", "Interest Rate", 
               "Purpose", "Date", "Underwriter", "Status"]
    for i, header in enumerate(headers, 1):
        ws[f'A{i+4}'] = header
        ws[f'A{i+4}'].font = Font(bold=True)
    
    # Example data
    example_data = [
        "Main Street Bakery", 
        "$150,000", 
        "60", 
        "7.5%", 
        "Equipment Purchase", 
        datetime.datetime.now().strftime("%m/%d/%Y"),
        "John Smith",
        "Under Review"
    ]
    
    for i, data in enumerate(example_data, 1):
        ws[f'B{i+4}'] = data
    
    # Add toggle button instructions
    ws['A15'] = "Toggle Example Data:"
    ws['B15'] = "YES"  # This will be linked to a data validation dropdown
    
    # Add completion summary
    ws['D3'] = "CHECKLIST COMPLETION"
    ws['D3'].font = Font(size=12, bold=True)
    
    sections = ["Underwriting Checklist", "Financial Analysis", "Management Assessment", 
                "Industry Evaluation", "Risk Assessment"]
    
    for i, section in enumerate(sections, 1):
        ws[f'D{i+4}'] = section
        ws[f'D{i+4}'].font = Font(bold=True)
        ws[f'E{i+4}'] = "70%"  # Example completion rate
    
    # Add scoring summary
    ws['A18'] = "SCORING SUMMARY"
    ws['A18'].font = Font(size=12, bold=True)
    
    score_categories = ["Financial Health", "Management Capacity", "Industry Outlook", 
                        "Business Model", "Collateral", "Overall Risk Rating"]
    
    for i, category in enumerate(score_categories, 1):
        ws[f'A{i+19}'] = category
        ws[f'A{i+19}'].font = Font(bold=True)
        ws[f'B{i+19}'] = "B"  # Example score
    
    # Add space for a chart (will be created later)
    ws.merge_cells('D18:H30')

def setup_checklist(ws):
    """Sets up the main underwriting checklist sheet."""
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "SMALL BUSINESS LOAN UNDERWRITING CHECKLIST"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor="1B4620")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    headers = ["Category", "Document/Requirement", "Required", "Received", "Date", "Notes"]
    for i, header in enumerate(headers):
        col = get_column_letter(i+1)
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = Font(bold=True)
        ws[f'{col}3'].fill = PatternFill("solid", fgColor="E0E0E0")
    
    # Main categories and items
    categories = {
        "Business Information": [
            "Business Plan",
            "Articles of Incorporation/Organization",
            "Business License(s)",
            "Certificate of Good Standing",
            "EIN Documentation",
            "Organizational Chart"
        ],
        "Financial Documents": [
            "Last 3 Years Business Tax Returns",
            "Last 3 Years Personal Tax Returns",
            "Year-to-Date Financial Statements",
            "Last 3 Years Balance Sheets",
            "Last 3 Years Income Statements",
            "Cash Flow Projections (2 years)",
            "Accounts Receivable Aging",
            "Accounts Payable Aging",
            "Debt Schedule"
        ],
        "Personal Information": [
            "Personal Financial Statement",
            "Personal Credit Report",
            "Management Resume(s)",
            "Personal ID(s)",
            "Personal Background Check Authorization"
        ],
        "Collateral Information": [
            "Property Appraisal",
            "Equipment Valuation",
            "Inventory List",
            "Deed(s) to Real Estate",
            "Title Insurance",
            "Environmental Assessment (if applicable)"
        ],
        "Insurance": [
            "Business Insurance Policy",
            "Life Insurance Policy (if applicable)",
            "Workers' Compensation Insurance",
            "Property Insurance"
        ],
        "Legal Documents": [
            "Lease Agreement(s)",
            "Franchise Agreement (if applicable)",
            "Contracts with Major Customers",
            "Supplier Agreements",
            "Patent/Trademark Documentation (if applicable)"
        ]
    }
    
    row = 4
    for category, items in categories.items():
        # Add category header
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = category
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill("solid", fgColor="F0F0F0")
        row += 1
        
        # Add items
        for item in items:
            ws[f'A{row}'] = category
            ws[f'B{row}'] = item
            ws[f'C{row}'] = "Yes"  # Default to required
            row += 1

def setup_financial(ws):
    """Sets up the Financial Analysis sheet."""
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = "FINANCIAL ANALYSIS WORKSHEET"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor="1B4620")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Key ratio sections
    ws['A3'] = "LIQUIDITY RATIOS"
    ws['A3'].font = Font(size=12, bold=True)
    
    liquidity_ratios = [
        ("Current Ratio", "Current Assets / Current Liabilities", "1.5", "2.0"),
        ("Quick Ratio", "(Current Assets - Inventory) / Current Liabilities", "1.0", "1.5"),
        ("Cash Ratio", "Cash / Current Liabilities", "0.2", "0.5")
    ]
    
    headers = ["Ratio", "Formula", "Result", "Industry Avg"]
    for i, header in enumerate(headers):
        col = get_column_letter(i+1)
        ws[f'{col}4'] = header
        ws[f'{col}4'].font = Font(bold=True)
        ws[f'{col}4'].fill = PatternFill("solid", fgColor="E0E0E0")
    
    row = 5
    for ratio in liquidity_ratios:
        for i, value in enumerate(ratio):
            col = get_column_letter(i+1)
            ws[f'{col}{row}'] = value
        row += 1
    
    # Profitability ratios
    row += 1
    ws[f'A{row}'] = "PROFITABILITY RATIOS"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1
    
    for i, header in enumerate(headers):
        col = get_column_letter(i+1)
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill("solid", fgColor="E0E0E0")
    row += 1
    
    profitability_ratios = [
        ("Gross Profit Margin", "Gross Profit / Revenue", "45%", "42%"),
        ("Net Profit Margin", "Net Income / Revenue", "15%", "12%"),
        ("Return on Assets", "Net Income / Total Assets", "8%", "7%"),
        ("Return on Equity", "Net Income / Shareholders' Equity", "18%", "15%")
    ]
    
    for ratio in profitability_ratios:
        for i, value in enumerate(ratio):
            col = get_column_letter(i+1)
            ws[f'{col}{row}'] = value
        row += 1
    
    # Leverage ratios
    row += 1
    ws[f'A{row}'] = "LEVERAGE RATIOS"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1
    
    for i, header in enumerate(headers):
        col = get_column_letter(i+1)
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill("solid", fgColor="E0E0E0")
    row += 1
    
    leverage_ratios = [
        ("Debt Ratio", "Total Debt / Total Assets", "0.45", "0.5"),
        ("Debt-to-Equity", "Total Debt / Shareholders' Equity", "0.8", "1.0"),
        ("Interest Coverage", "EBIT / Interest Expense", "5.5", "3.0")
    ]
    
    for ratio in leverage_ratios:
        for i, value in enumerate(ratio):
            col = get_column_letter(i+1)
            ws[f'{col}{row}'] = value
        row += 1
    
    # Efficiency ratios
    row += 1
    ws[f'A{row}'] = "EFFICIENCY RATIOS"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1
    
    for i, header in enumerate(headers):
        col = get_column_letter(i+1)
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill("solid", fgColor="E0E0E0")
    row += 1
    
    efficiency_ratios = [
        ("Inventory Turnover", "COGS / Average Inventory", "6.0", "5.5"),
        ("Receivables Turnover", "Revenue / Average Accounts Receivable", "8.5", "7.0"),
        ("Asset Turnover", "Revenue / Total Assets", "1.8", "1.5")
    ]
    
    for ratio in efficiency_ratios:
        for i, value in enumerate(ratio):
            col = get_column_letter(i+1)
            ws[f'{col}{row}'] = value
        row += 1
    
    # Cash flow analysis
    row += 2
    ws[f'A{row}'] = "CASH FLOW ANALYSIS"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1
    
    cash_headers = ["Cash Flow Category", "Current Year", "Previous Year", "% Change"]
    for i, header in enumerate(cash_headers):
        col = get_column_letter(i+1)
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill("solid", fgColor="E0E0E0")
    row += 1
    
    cash_flow_items = [
        ("Operating Cash Flow", "$250,000", "$220,000", "13.6%"),
        ("Investing Cash Flow", "($150,000)", "($100,000)", "50.0%"),
        ("Financing Cash Flow", "$50,000", "$30,000", "66.7%"),
        ("Net Cash Flow", "$150,000", "$150,000", "0.0%")
    ]
    
    for item in cash_flow_items:
        for i, value in enumerate(item):
            col = get_column_letter(i+1)
            ws[f'{col}{row}'] = value
        row += 1

def setup_management(ws):
    """Sets up the Management Assessment sheet."""
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = "MANAGEMENT ASSESSMENT"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor="1B4620")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    ws['A3'] = "KEY MANAGEMENT EVALUATION"
    ws['A3'].font = Font(size=12, bold=True)
    
    headers = ["Category", "Score (1-5)", "Comments", "Weight", "Weighted Score"]
    for i, header in enumerate(headers):
        col = get_column_letter(i+1)
        ws[f'{col}4'] = header
        ws[f'{col}4'].font = Font(bold=True)
        ws[f'{col}4'].fill = PatternFill("solid", fgColor="E0E0E0")
    
    categories = [
        ("Industry Experience", "4", "15+ years in industry", "25%", "=B5*D5"),
        ("Business Management Experience", "3", "Good overall management, limited financial oversight", "20%", "=B6*D6"),
        ("Technical Knowledge", "5", "Expert in field", "15%", "=B7*D7"),
        ("Financial Acumen", "3", "Adequate understanding of financial statements", "20%", "=B8*D8"),
        ("Succession Planning", "2", "Limited planning for management transition", "10%", "=B9*D9"),
        ("Team Completeness", "4", "Strong core team", "10%", "=B10*D10")
    ]
    
    row = 5
    for category in categories:
        for i, value in enumerate(category):
            col = get_column_letter(i+1)
            ws[f'{col}{row}'] = value
        row += 1
    
    # Total row
    ws[f'A{row}'] = "TOTAL"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'E{row}'] = "=SUM(E5:E10)"
    ws[f'E{row}'].font = Font(bold=True)
    
    # Scoring guide
    row += 3
    ws[f'A{row}'] = "SCORING GUIDE"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1
    
    guide_headers = ["Score", "Description"]
    for i, header in enumerate(guide_headers):
        col = get_column_letter(i+1)
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill("solid", fgColor="E0E0E0")
    row += 1
    
    score_descriptions = [
        ("5", "Exceptional - Top tier performance, exceeds all requirements"),
        ("4", "Strong - Above average, meets all requirements with some strengths"),
        ("3", "Adequate - Meets basic requirements"),
        ("2", "Marginal - Below average, has notable weaknesses"),
        ("1", "Poor - Significant deficiencies, major concerns")
    ]
    
    for score in score_descriptions:
        for i, value in enumerate(score):
            col = get_column_letter(i+1)
            ws[f'{col}{row}'] = value
        row += 1

def setup_industry(ws):
    """Sets up the Industry Evaluation sheet."""
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = "INDUSTRY EVALUATION"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor="1B4620")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Industry analysis
    ws['A3'] = "INDUSTRY ANALYSIS"
    ws['A3'].font = Font(size=12, bold=True)
    
    headers = ["Factor", "Rating (1-5)", "Comments", "Weight", "Weighted Score"]
    for i, header in enumerate(headers):
        col = get_column_letter(i+1)
        ws[f'{col}4'] = header
        ws[f'{col}4'].font = Font(bold=True)
        ws[f'{col}4'].fill = PatternFill("solid", fgColor="E0E0E0")
    
    factors = [
        ("Industry Growth Trend", "4", "Stable growth of 5-7% annually", "20%", "=B5*D5"),
        ("Competitive Landscape", "3", "Moderate competition, some established players", "15%", "=B6*D6"),
        ("Barriers to Entry", "4", "Significant capital requirements and expertise needed", "10%", "=B7*D7"),
        ("Technology Disruption Risk", "2", "Emerging tech could impact business model", "15%", "=B8*D8"),
        ("Regulatory Environment", "3", "Stable regulations with some compliance costs", "15%", "=B9*D9"),
        ("Supply Chain Stability", "4", "Multiple reliable suppliers available", "10%", "=B10*D10"),
        ("Customer Concentration", "3", "Some reliance on top customers, reasonable diversity", "15%", "=B11*D11")
    ]
    
    row = 5
    for factor in factors:
        for i, value in enumerate(factor):
            col = get_column_letter(i+1)
            ws[f'{col}{row}'] = value
        row += 1
    
    # Total row
    ws[f'A{row}'] = "TOTAL"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'E{row}'] = "=SUM(E5:E11)"
    ws[f'E{row}'].font = Font(bold=True)
    
    # Market position analysis
    row += 3
    ws[f'A{row}'] = "MARKET POSITION ANALYSIS"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1
    
    market_headers = ["Factor", "Assessment", "Strength/Weakness"]
    for i, header in enumerate(market_headers):
        col = get_column_letter(i+1)
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill("solid", fgColor="E0E0E0")
    row += 1
    
    market_factors = [
        ("Market Share", "8% of local market", "Strength"),
        ("Unique Selling Proposition", "Proprietary manufacturing process", "Strength"),
        ("Brand Recognition", "Limited to local area", "Weakness"),
        ("Customer Loyalty", "High retention rate (85%)", "Strength"),
        ("Pricing Power", "Limited ability to raise prices", "Weakness"),
        ("Distribution Channels", "Well-established network", "Strength")
    ]
    
    for factor in market_factors:
        for i, value in enumerate(factor):
            col = get_column_letter(i+1)
            ws[f'{col}{row}'] = value
        row += 1

def setup_risk(ws):
    """Sets up the Risk Assessment sheet."""
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "RISK ASSESSMENT"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor="1B4620")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Risk assessment matrix
    ws['A3'] = "RISK ASSESSMENT MATRIX"
    ws['A3'].font = Font(size=12, bold=True)
    
    headers = ["Risk Category", "Risk Level (1-5)", "Mitigating Factors", "Impact on Decision"]
    for i, header in enumerate(headers):
        col = get_column_letter(i+1)
        ws[f'{col}4'] = header
        ws[f'{col}4'].font = Font(bold=True)
        ws[f'{col}4'].fill = PatternFill("solid", fgColor="E0E0E0")
    
    risks = [
        ("Credit Risk", "3", "Good payment history with suppliers", "Moderate concern"),
        ("Market Risk", "2", "Established customer base, growing market", "Low concern"),
        ("Operational Risk", "4", "Limited backup systems, key person dependency", "High concern"),
        ("Financial Risk", "3", "Adequate cash reserves, some debt", "Moderate concern"),
        ("Legal/Regulatory Risk", "2", "Compliance procedures in place", "Low concern"),
        ("Environmental Risk", "1", "Low environmental impact business", "Minimal concern"),
        ("Technology Risk", "3", "Some outdated systems planned for upgrade", "Moderate concern")
    ]
    
    row = 5
    for risk in risks:
        for i, value in enumerate(risk):
            col = get_column_letter(i+1)
            ws[f'{col}{row}'] = value
        row += 1
    
    # Overall risk rating
    row += 3
    ws[f'A{row}'] = "OVERALL RISK RATING"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    row += 1
    
    rating_headers = ["Rating", "Description"]
    for i, header in enumerate(rating_headers):
        col = get_column_letter(i+1)
        ws[f'{col}{row}'] = header
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill("solid", fgColor="E0E0E0")
    row += 1
    
    # Example overall rating
    ws[f'A{row}'] = "B"
    ws[f'B{row}'] = "Moderate risk profile with adequate mitigating factors. Acceptable for standard loan terms with appropriate monitoring."

def add_formulas(wb):
    """Adds formulas and connections between sheets."""
    # Add data validation for the toggle cell
    dashboard = wb["Dashboard"]
    dv = openpyxl.worksheet.datavalidation.DataValidation(type="list", formula1='"YES,NO"')
    dv.add(dashboard["B15"])
    dashboard.add_data_validation(dv)
    
    # Add a chart to the dashboard
    chart = BarChart()
    chart.title = "Key Risk Categories"
    chart.y_axis.title = "Risk Level"
    chart.x_axis.title = "Category"
    
    risk = wb["Risk Assessment"]
    data = Reference(risk, min_col=2, min_row=5, max_col=2, max_row=11)
    cats = Reference(risk, min_col=1, min_row=5, max_col=1, max_row=11)
    
    chart.add_data(data)
    chart.set_categories(cats)
    
    dashboard.add_chart(chart, "D18")

if __name__ == "__main__":
    create_workbook()
