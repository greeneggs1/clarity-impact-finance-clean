#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate Excel Template for Cash Flow Analysis

This script creates a formatted Excel template that can be used with the
Small Business Cash Flow Analysis tool.

Author: Clarity Impact Finance
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import datetime

def create_excel_template(output_file='small_business_financials_template.xlsx'):
    """
    Create a formatted Excel template for financial data input.
    
    Parameters:
    ----------
    output_file : str
        Path for the output Excel file
    """
    # Create a new workbook
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create sheets for income statement and balance sheet
    ws_income = wb.create_sheet("Income Statement")
    ws_balance = wb.create_sheet("Balance Sheet")
    ws_instructions = wb.create_sheet("Instructions")
    
    # Current year for generating sample years
    current_year = datetime.datetime.now().year
    years = [str(current_year - 2), str(current_year - 1), str(current_year)]
    
    # Format income statement sheet
    format_income_statement(ws_income, years)
    
    # Format balance sheet
    format_balance_sheet(ws_balance, years)
    
    # Add instructions
    add_instructions(ws_instructions)
    
    # Auto-adjust column widths
    for ws in [ws_income, ws_balance, ws_instructions]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    try:
        wb.save(output_file)
        print(f"Successfully created template at {output_file}")
        return True
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        return False


def format_income_statement(ws, years):
    """Format the income statement worksheet."""
    # Add title
    ws.merge_cells('A1:E1')
    ws['A1'] = "Small Business Income Statement"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add subtitle with instructions
    ws.merge_cells('A2:E2')
    ws['A2'] = "Enter annual data for all completed fiscal years"
    ws['A2'].font = Font(italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Add header row
    header_row = ['Account'] + years
    for i, header in enumerate(header_row, start=1):
        cell = ws.cell(row=4, column=i, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Add income statement accounts
    accounts = [
        # Revenue section
        {'name': 'Revenue', 'section': True},
        {'name': 'Gross Sales', 'indent': 1},
        {'name': 'Returns and Allowances', 'indent': 1},
        {'name': 'Net Sales', 'bold': True, 'indent': 1},
        {'name': 'Other Revenue', 'indent': 1},
        {'name': 'Total Revenue', 'bold': True},
        {'name': ''},
        
        # Cost of Goods Sold section
        {'name': 'Cost of Goods Sold', 'section': True},
        {'name': 'Beginning Inventory', 'indent': 1},
        {'name': 'Purchases', 'indent': 1},
        {'name': 'Labor', 'indent': 1},
        {'name': 'Materials', 'indent': 1},
        {'name': 'Other COGS Costs', 'indent': 1},
        {'name': 'Ending Inventory', 'indent': 1},
        {'name': 'Total Cost of Goods Sold', 'bold': True},
        {'name': ''},
        
        # Gross Profit
        {'name': 'Gross Profit', 'bold': True},
        {'name': ''},
        
        # Operating Expenses
        {'name': 'Operating Expenses', 'section': True},
        {'name': 'Salaries and Wages', 'indent': 1},
        {'name': 'Payroll Taxes', 'indent': 1},
        {'name': 'Employee Benefits', 'indent': 1},
        {'name': 'Rent', 'indent': 1},
        {'name': 'Utilities', 'indent': 1},
        {'name': 'Insurance', 'indent': 1},
        {'name': 'Depreciation', 'indent': 1},
        {'name': 'Amortization', 'indent': 1},
        {'name': 'Office Supplies', 'indent': 1},
        {'name': 'Repairs and Maintenance', 'indent': 1},
        {'name': 'Marketing and Advertising', 'indent': 1},
        {'name': 'Professional Fees', 'indent': 1},
        {'name': 'Travel and Entertainment', 'indent': 1},
        {'name': 'Other Operating Expenses', 'indent': 1},
        {'name': 'Total Operating Expenses', 'bold': True},
        {'name': ''},
        
        # Operating Income
        {'name': 'Operating Income', 'bold': True},
        {'name': ''},
        
        # Other Income and Expenses
        {'name': 'Other Income and Expenses', 'section': True},
        {'name': 'Interest Income', 'indent': 1},
        {'name': 'Interest Expense', 'indent': 1},
        {'name': 'Gain/Loss on Sale of Assets', 'indent': 1},
        {'name': 'Other Income/Expenses', 'indent': 1},
        {'name': 'Total Other Income and Expenses', 'bold': True},
        {'name': ''},
        
        # Income Before Taxes
        {'name': 'Income Before Taxes', 'bold': True},
        {'name': 'Income Taxes', 'bold': True},
        {'name': 'Net Income', 'bold': True}
    ]
    
    # Add accounts to the worksheet
    for i, account in enumerate(accounts, start=5):
        # Account name
        cell = ws.cell(row=i, column=1, value=account.get('name', ''))
        
        # Apply formatting
        if account.get('bold', False):
            cell.font = Font(bold=True)
        
        if account.get('section', False):
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Add indentation
        indent = account.get('indent', 0)
        if indent > 0:
            cell.alignment = Alignment(indent=indent)
        
        # Add sample formulas for totals (in the last year column)
        if account.get('name') == 'Net Sales':
            ws.cell(row=i, column=len(years)+1, value="=B6-B7")
        elif account.get('name') == 'Total Revenue':
            ws.cell(row=i, column=len(years)+1, value="=B9+B10")
        elif account.get('name') == 'Total Cost of Goods Sold':
            ws.cell(row=i, column=len(years)+1, value="=B13+B14+B15+B16+B17-B18")
        elif account.get('name') == 'Gross Profit':
            ws.cell(row=i, column=len(years)+1, value="=B11-B20")
        elif account.get('name') == 'Total Operating Expenses':
            ws.cell(row=i, column=len(years)+1, value="=SUM(B24:B37)")
        elif account.get('name') == 'Operating Income':
            ws.cell(row=i, column=len(years)+1, value="=B22-B39")
        elif account.get('name') == 'Total Other Income and Expenses':
            ws.cell(row=i, column=len(years)+1, value="=B43-B44+B45+B46")
        elif account.get('name') == 'Income Before Taxes':
            ws.cell(row=i, column=len(years)+1, value="=B41+B48")
        elif account.get('name') == 'Net Income':
            ws.cell(row=i, column=len(years)+1, value="=B50-B51")


def format_balance_sheet(ws, years):
    """Format the balance sheet worksheet."""
    # Add title
    ws.merge_cells('A1:E1')
    ws['A1'] = "Small Business Balance Sheet"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add subtitle with instructions
    ws.merge_cells('A2:E2')
    ws['A2'] = "Enter data as of the last day of each fiscal year"
    ws['A2'].font = Font(italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Add header row
    header_row = ['Account'] + years
    for i, header in enumerate(header_row, start=1):
        cell = ws.cell(row=4, column=i, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Add balance sheet accounts
    accounts = [
        # Assets
        {'name': 'ASSETS', 'section': True},
        {'name': 'Current Assets', 'bold': True},
        {'name': 'Cash and Cash Equivalents', 'indent': 1},
        {'name': 'Accounts Receivable', 'indent': 1},
        {'name': 'Inventory', 'indent': 1},
        {'name': 'Prepaid Expenses', 'indent': 1},
        {'name': 'Other Current Assets', 'indent': 1},
        {'name': 'Total Current Assets', 'bold': True},
        {'name': ''},
        
        {'name': 'Non-Current Assets', 'bold': True},
        {'name': 'Property, Plant & Equipment', 'indent': 1},
        {'name': 'Accumulated Depreciation', 'indent': 1},
        {'name': 'Net PP&E', 'bold': True, 'indent': 1},
        {'name': 'Intangible Assets', 'indent': 1},
        {'name': 'Accumulated Amortization', 'indent': 1},
        {'name': 'Net Intangible Assets', 'bold': True, 'indent': 1},
        {'name': 'Long-term Investments', 'indent': 1},
        {'name': 'Other Long-term Assets', 'indent': 1},
        {'name': 'Total Non-Current Assets', 'bold': True},
        {'name': ''},
        
        {'name': 'TOTAL ASSETS', 'bold': True},
        {'name': ''},
        
        # Liabilities and Equity
        {'name': 'LIABILITIES AND EQUITY', 'section': True},
        {'name': 'Current Liabilities', 'bold': True},
        {'name': 'Accounts Payable', 'indent': 1},
        {'name': 'Short-term Debt', 'indent': 1},
        {'name': 'Current Portion of Long-term Debt', 'indent': 1},
        {'name': 'Accrued Expenses', 'indent': 1},
        {'name': 'Deferred Revenue', 'indent': 1},
        {'name': 'Other Current Liabilities', 'indent': 1},
        {'name': 'Total Current Liabilities', 'bold': True},
        {'name': ''},
        
        {'name': 'Non-Current Liabilities', 'bold': True},
        {'name': 'Long-term Debt', 'indent': 1},
        {'name': 'Deferred Tax Liabilities', 'indent': 1},
        {'name': 'Other Long-term Liabilities', 'indent': 1},
        {'name': 'Total Non-Current Liabilities', 'bold': True},
        {'name': ''},
        
        {'name': 'TOTAL LIABILITIES', 'bold': True},
        {'name': ''},
        
        {'name': 'Equity', 'bold': True},
        {'name': 'Capital Stock', 'indent': 1},
        {'name': 'Additional Paid-in Capital', 'indent': 1},
        {'name': 'Retained Earnings', 'indent': 1},
        {'name': 'Other Equity', 'indent': 1},
        {'name': 'Total Equity', 'bold': True},
        {'name': ''},
        
        {'name': 'TOTAL LIABILITIES AND EQUITY', 'bold': True}
    ]
    
    # Add accounts to the worksheet
    for i, account in enumerate(accounts, start=5):
        # Account name
        cell = ws.cell(row=i, column=1, value=account.get('name', ''))
        
        # Apply formatting
        if account.get('bold', False):
            cell.font = Font(bold=True)
        
        if account.get('section', False):
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Add indentation
        indent = account.get('indent', 0)
        if indent > 0:
            cell.alignment = Alignment(indent=indent)
        
        # Add sample formulas for totals (in the last year column)
        if account.get('name') == 'Total Current Assets':
            ws.cell(row=i, column=len(years)+1, value="=SUM(B7:B11)")
        elif account.get('name') == 'Net PP&E':
            ws.cell(row=i, column=len(years)+1, value="=B14+B15")
        elif account.get('name') == 'Net Intangible Assets':
            ws.cell(row=i, column=len(years)+1, value="=B17+B18")
        elif account.get('name') == 'Total Non-Current Assets':
            ws.cell(row=i, column=len(years)+1, value="=B16+B19+B20+B21")
        elif account.get('name') == 'TOTAL ASSETS':
            ws.cell(row=i, column=len(years)+1, value="=B12+B23")
        elif account.get('name') == 'Total Current Liabilities':
            ws.cell(row=i, column=len(years)+1, value="=SUM(B28:B33)")
        elif account.get('name') == 'Total Non-Current Liabilities':
            ws.cell(row=i, column=len(years)+1, value="=SUM(B36:B38)")
        elif account.get('name') == 'TOTAL LIABILITIES':
            ws.cell(row=i, column=len(years)+1, value="=B34+B40")
        elif account.get('name') == 'Total Equity':
            ws.cell(row=i, column=len(years)+1, value="=SUM(B44:B47)")
        elif account.get('name') == 'TOTAL LIABILITIES AND EQUITY':
            ws.cell(row=i, column=len(years)+1, value="=B42+B49")


def add_instructions(ws):
    """Add instructions to the worksheet."""
    # Add title
    ws.merge_cells('A1:E1')
    ws['A1'] = "Instructions for Financial Data Input"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add instructions
    instructions = [
        {"title": "General Instructions", "content": [
            "This template is designed to be used with the Small Business Cash Flow Analysis tool.",
            "Enter financial data for at least two fiscal years, with the most recent year in the rightmost column.",
            "Use consistent accounting methods across all years for accurate trend analysis.",
            "All amounts should be entered as positive numbers, except where indicated (e.g., Accumulated Depreciation)."
        ]},
        {"title": "Income Statement Instructions", "content": [
            "Enter annual data for each completed fiscal year.",
            "Revenue and expenses should be entered for the entire fiscal year.",
            "Cost of Goods Sold should include all direct costs related to product creation.",
            "Depreciation and Amortization must be entered separately for accurate EBITDA calculation.",
            "Ensure Net Income is calculated correctly as Income Before Taxes minus Income Taxes."
        ]},
        {"title": "Balance Sheet Instructions", "content": [
            "Enter data as of the last day of each fiscal year.",
            "Accounts Receivable should be entered net of allowance for doubtful accounts.",
            "Accumulated Depreciation and Amortization should be entered as negative numbers.",
            "Current Portion of Long-term Debt refers to debt due within the next 12 months.",
            "Ensure Total Assets equals Total Liabilities and Equity for each year."
        ]},
        {"title": "Required Accounts", "content": [
            "The following accounts must be filled in for proper analysis:",
            "- Revenue",
            "- Cost of Goods Sold",
            "- Operating Expenses",
            "- Depreciation",
            "- Amortization",
            "- Interest Expense",
            "- Net Income",
            "- Cash",
            "- Accounts Receivable",
            "- Inventory",
            "- Total Current Assets",
            "- Total Assets",
            "- Accounts Payable",
            "- Current Portion of Long-term Debt",
            "- Total Current Liabilities",
            "- Long-term Debt",
            "- Total Liabilities",
            "- Equity"
        ]}
    ]
    
    row = 3
    for section in instructions:
        # Add section title
        ws.cell(row=row, column=1, value=section["title"])
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        # Add section content
        for line in section["content"]:
            ws.cell(row=row, column=1, value=line)
            row += 1
        
        # Add space between sections
        row += 1
    
    # Add footer
    row += 2
    ws.cell(row=row, column=1, value="For questions or assistance, contact Clarity Impact Finance:")
    row += 1
    ws.cell(row=row, column=1, value="info@clarityimpactfinance.com")


if __name__ == "__main__":
    create_excel_template() 