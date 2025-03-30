#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Small Business Cash Flow Analysis Tool

This script provides a comprehensive cash flow analysis framework for small business lending,
including:
- Uniform Credit Analysis (UCA) Cash Flow
- EBITDA Analysis
- Debt Service Coverage Ratio
- Various financial ratios and metrics

Features:
- Import financial data from Excel or CSV
- Perform complete cash flow analysis
- Export results to Excel with formatted worksheets
- Visualize cash flow trends and key metrics

Author: Clarity Impact Finance
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
import os
import datetime


class SmallBusinessCashFlowAnalyzer:
    """
    A comprehensive cash flow analysis tool for small business lending.
    """
    
    def __init__(self):
        """Initialize the analyzer with empty data structures."""
        self.income_statement = None
        self.balance_sheet = None
        self.years = []
        self.uca_cash_flow = None
        self.ebitda_analysis = None
        self.dscr_analysis = None
        self.ratios = None
        
    def load_from_excel(self, file_path, income_sheet='Income Statement', 
                        balance_sheet='Balance Sheet'):
        """
        Load financial data from an Excel file with specific sheets.
        
        Parameters:
        ----------
        file_path : str
            Path to the Excel file
        income_sheet : str
            Name of the sheet containing income statement data
        balance_sheet : str
            Name of the sheet containing balance sheet data
        """
        try:
            # Load income statement
            self.income_statement = pd.read_excel(file_path, sheet_name=income_sheet, index_col=0)
            # Load balance sheet
            self.balance_sheet = pd.read_excel(file_path, sheet_name=balance_sheet, index_col=0)
            
            # Extract years from columns
            self.years = list(self.income_statement.columns)
            
            print(f"Successfully loaded data from {file_path}")
            print(f"Years analyzed: {', '.join(map(str, self.years))}")
            
            return True
        except Exception as e:
            print(f"Error loading data: {e}")
            return False
    
    def create_sample_data(self, years=3):
        """
        Create sample financial data for demonstration purposes.
        
        Parameters:
        ----------
        years : int
            Number of years to generate data for
        """
        # Generate year labels (current year and past years)
        current_year = datetime.datetime.now().year
        self.years = [str(current_year - i) for i in range(years)][::-1]
        
        # Create sample income statement
        self.income_statement = pd.DataFrame({
            self.years[0]: {
                'Revenue': 500000,
                'Cost of Goods Sold': 300000,
                'Gross Profit': 200000,
                'Operating Expenses': 120000,
                'Depreciation': 15000,
                'Amortization': 5000,
                'EBIT': 60000,
                'Interest Expense': 8000,
                'Income Before Taxes': 52000,
                'Income Taxes': 15600,
                'Net Income': 36400
            }
        })
        
        # Add some growth for subsequent years
        for i in range(1, years):
            growth_factor = 1.1 + (i * 0.05)  # Increasing growth
            self.income_statement[self.years[i]] = self.income_statement[self.years[0]] * growth_factor
        
        # Create sample balance sheet
        self.balance_sheet = pd.DataFrame({
            self.years[0]: {
                'Cash': 45000,
                'Accounts Receivable': 60000,
                'Inventory': 75000,
                'Other Current Assets': 10000,
                'Total Current Assets': 190000,
                'Property, Plant & Equipment': 250000,
                'Accumulated Depreciation': -65000,
                'Net PP&E': 185000,
                'Intangible Assets': 25000,
                'Other Long-term Assets': 15000,
                'Total Assets': 415000,
                'Accounts Payable': 35000,
                'Short-term Debt': 15000,
                'Current Portion of Long-term Debt': 12000,
                'Other Current Liabilities': 18000,
                'Total Current Liabilities': 80000,
                'Long-term Debt': 120000,
                'Other Long-term Liabilities': 20000,
                'Total Liabilities': 220000,
                'Equity': 195000,
                'Total Liabilities & Equity': 415000
            }
        })
        
        # Add some changes for subsequent years
        for i in range(1, years):
            asset_growth = 1.08 + (i * 0.04)
            liability_growth = 1.05 + (i * 0.03)
            
            new_balance = {}
            for account, value in self.balance_sheet[self.years[0]].items():
                if 'Asset' in account or account in ['Cash', 'Inventory', 'Receivable', 'PP&E']:
                    new_balance[account] = value * asset_growth
                elif 'Liabilit' in account or 'Debt' in account or 'Payable' in account:
                    new_balance[account] = value * liability_growth
                else:
                    # For equity and other calculated fields, just use a placeholder
                    new_balance[account] = value * asset_growth
            
            # Fix the balance sheet to balance
            total_assets = (new_balance['Total Current Assets'] 
                           + new_balance['Net PP&E'] 
                           + new_balance['Intangible Assets'] 
                           + new_balance['Other Long-term Assets'])
            new_balance['Total Assets'] = total_assets
            
            total_liabilities = new_balance['Total Liabilities']
            new_balance['Equity'] = total_assets - total_liabilities
            new_balance['Total Liabilities & Equity'] = total_liabilities + new_balance['Equity']
            
            self.balance_sheet[self.years[i]] = pd.Series(new_balance)
        
        print("Created sample financial data for demonstration")
        return True
    
    def calculate_uca_cash_flow(self):
        """
        Calculate Uniform Credit Analysis (UCA) Cash Flow.
        
        The UCA Cash Flow method is a standardized approach to calculate
        cash flow from operations, with adjustments for working capital changes.
        """
        if self.income_statement is None or self.balance_sheet is None:
            print("Error: Financial data not loaded")
            return False
        
        data = {}
        for year in self.years:
            # Get values from income statement
            try:
                net_income = self.income_statement.loc['Net Income', year]
                depreciation = self.income_statement.loc['Depreciation', year]
                amortization = self.income_statement.loc['Amortization', year]
                
                # Calculate changes in balance sheet accounts if we have previous year data
                year_idx = self.years.index(year)
                
                if year_idx > 0:
                    prev_year = self.years[year_idx - 1]
                    
                    # Working capital changes
                    change_ar = (self.balance_sheet.loc['Accounts Receivable', prev_year] - 
                                 self.balance_sheet.loc['Accounts Receivable', year])
                    
                    change_inventory = (self.balance_sheet.loc['Inventory', prev_year] - 
                                       self.balance_sheet.loc['Inventory', year])
                    
                    change_ap = (self.balance_sheet.loc['Accounts Payable', year] - 
                                self.balance_sheet.loc['Accounts Payable', prev_year])
                    
                    # Other changes may vary by business
                    change_other_current = (
                        (self.balance_sheet.loc['Other Current Assets', prev_year] - 
                         self.balance_sheet.loc['Other Current Assets', year]) +
                        (self.balance_sheet.loc['Other Current Liabilities', year] - 
                         self.balance_sheet.loc['Other Current Liabilities', prev_year])
                    )
                else:
                    # For the first year, we can't calculate changes
                    change_ar = 0
                    change_inventory = 0
                    change_ap = 0
                    change_other_current = 0
                
                # UCA Cash Flow calculation
                uca_cash_flow = (
                    net_income +
                    depreciation +
                    amortization +
                    change_ar +  # Decrease in AR is positive
                    change_inventory +  # Decrease in inventory is positive
                    change_ap +  # Increase in AP is positive
                    change_other_current
                )
                
                # Store results
                data[year] = {
                    'Net Income': net_income,
                    'Depreciation': depreciation,
                    'Amortization': amortization,
                    'Change in Accounts Receivable': change_ar,
                    'Change in Inventory': change_inventory,
                    'Change in Accounts Payable': change_ap,
                    'Change in Other Current Items': change_other_current,
                    'UCA Cash Flow': uca_cash_flow
                }
                
            except KeyError as e:
                print(f"Missing key in financial data: {e}")
                return False
            except Exception as e:
                print(f"Error calculating UCA Cash Flow: {e}")
                return False
        
        # Create DataFrame from results
        self.uca_cash_flow = pd.DataFrame(data)
        return True
    
    def calculate_ebitda(self):
        """
        Calculate EBITDA (Earnings Before Interest, Taxes, Depreciation and Amortization)
        and related metrics.
        """
        if self.income_statement is None:
            print("Error: Income statement data not loaded")
            return False
        
        data = {}
        for year in self.years:
            try:
                # Direct EBITDA calculation from income statement
                revenue = self.income_statement.loc['Revenue', year]
                cogs = self.income_statement.loc['Cost of Goods Sold', year]
                op_expenses = self.income_statement.loc['Operating Expenses', year]
                depreciation = self.income_statement.loc['Depreciation', year]
                amortization = self.income_statement.loc['Amortization', year]
                interest = self.income_statement.loc['Interest Expense', year]
                taxes = self.income_statement.loc['Income Taxes', year]
                
                # Calculate EBITDA components
                gross_profit = revenue - cogs
                operating_income = gross_profit - op_expenses
                ebit = operating_income  # Same as EBIT in our structure
                ebitda = ebit + depreciation + amortization
                
                # Store results
                data[year] = {
                    'Revenue': revenue,
                    'Gross Profit': gross_profit,
                    'Gross Margin (%)': (gross_profit / revenue) * 100 if revenue > 0 else 0,
                    'Operating Income': operating_income,
                    'Operating Margin (%)': (operating_income / revenue) * 100 if revenue > 0 else 0,
                    'EBIT': ebit,
                    'EBIT Margin (%)': (ebit / revenue) * 100 if revenue > 0 else 0,
                    'EBITDA': ebitda,
                    'EBITDA Margin (%)': (ebitda / revenue) * 100 if revenue > 0 else 0,
                    'Interest Expense': interest,
                    'Income Taxes': taxes
                }
                
            except KeyError as e:
                print(f"Missing key in income statement: {e}")
                return False
            except Exception as e:
                print(f"Error calculating EBITDA: {e}")
                return False
        
        # Create DataFrame from results
        self.ebitda_analysis = pd.DataFrame(data)
        return True
    
    def calculate_debt_service_coverage(self):
        """
        Calculate Debt Service Coverage Ratio (DSCR) and related debt metrics.
        """
        if self.income_statement is None or self.balance_sheet is None:
            print("Error: Financial data not loaded")
            return False
        
        if self.ebitda_analysis is None:
            success = self.calculate_ebitda()
            if not success:
                return False
        
        data = {}
        for year in self.years:
            try:
                # Get necessary values
                ebitda = self.ebitda_analysis.loc['EBITDA', year]
                interest = self.income_statement.loc['Interest Expense', year]
                
                # Get debt service components
                current_portion_ltd = self.balance_sheet.loc['Current Portion of Long-term Debt', year]
                
                # Calculate total debt service
                principal_payments = current_portion_ltd  # This is simplified
                total_debt_service = interest + principal_payments
                
                # Calculate coverage ratios
                times_interest_earned = ebitda / interest if interest > 0 else float('inf')
                dscr = ebitda / total_debt_service if total_debt_service > 0 else float('inf')
                
                # Get debt levels
                short_term_debt = self.balance_sheet.loc['Short-term Debt', year]
                long_term_debt = self.balance_sheet.loc['Long-term Debt', year]
                total_debt = short_term_debt + long_term_debt + current_portion_ltd
                
                # Get total assets for debt ratio
                total_assets = self.balance_sheet.loc['Total Assets', year]
                
                # Store results
                data[year] = {
                    'EBITDA': ebitda,
                    'Interest Expense': interest,
                    'Principal Payments': principal_payments,
                    'Total Debt Service': total_debt_service,
                    'Times Interest Earned': times_interest_earned,
                    'Debt Service Coverage Ratio (DSCR)': dscr,
                    'Short-term Debt': short_term_debt,
                    'Long-term Debt': long_term_debt,
                    'Total Debt': total_debt,
                    'Debt to Assets Ratio': total_debt / total_assets if total_assets > 0 else float('inf')
                }
                
            except KeyError as e:
                print(f"Missing key in financial data: {e}")
                return False
            except Exception as e:
                print(f"Error calculating Debt Service Coverage: {e}")
                return False
        
        # Create DataFrame from results
        self.dscr_analysis = pd.DataFrame(data)
        return True
    
    def calculate_financial_ratios(self):
        """
        Calculate key financial ratios for small business lending analysis.
        """
        if self.income_statement is None or self.balance_sheet is None:
            print("Error: Financial data not loaded")
            return False
        
        data = {}
        for year in self.years:
            try:
                # Profitability Ratios
                revenue = self.income_statement.loc['Revenue', year]
                net_income = self.income_statement.loc['Net Income', year]
                total_assets = self.balance_sheet.loc['Total Assets', year]
                equity = self.balance_sheet.loc['Equity', year]
                
                return_on_assets = (net_income / total_assets) * 100 if total_assets > 0 else 0
                return_on_equity = (net_income / equity) * 100 if equity > 0 else 0
                net_profit_margin = (net_income / revenue) * 100 if revenue > 0 else 0
                
                # Liquidity Ratios
                current_assets = self.balance_sheet.loc['Total Current Assets', year]
                current_liabilities = self.balance_sheet.loc['Total Current Liabilities', year]
                inventory = self.balance_sheet.loc['Inventory', year]
                
                current_ratio = current_assets / current_liabilities if current_liabilities > 0 else float('inf')
                quick_ratio = (current_assets - inventory) / current_liabilities if current_liabilities > 0 else float('inf')
                
                # Efficiency Ratios
                accounts_receivable = self.balance_sheet.loc['Accounts Receivable', year]
                accounts_payable = self.balance_sheet.loc['Accounts Payable', year]
                cogs = self.income_statement.loc['Cost of Goods Sold', year]
                
                receivables_turnover = revenue / accounts_receivable if accounts_receivable > 0 else float('inf')
                days_receivables = 365 / receivables_turnover if receivables_turnover > 0 else float('inf')
                
                inventory_turnover = cogs / inventory if inventory > 0 else float('inf')
                days_inventory = 365 / inventory_turnover if inventory_turnover > 0 else float('inf')
                
                payables_turnover = cogs / accounts_payable if accounts_payable > 0 else float('inf')
                days_payables = 365 / payables_turnover if payables_turnover > 0 else float('inf')
                
                # Cash Conversion Cycle
                cash_conversion_cycle = days_receivables + days_inventory - days_payables
                
                # Store results
                data[year] = {
                    # Profitability
                    'Return on Assets (%)': return_on_assets,
                    'Return on Equity (%)': return_on_equity,
                    'Net Profit Margin (%)': net_profit_margin,
                    
                    # Liquidity
                    'Current Ratio': current_ratio,
                    'Quick Ratio': quick_ratio,
                    
                    # Efficiency
                    'Receivables Turnover': receivables_turnover,
                    'Days Receivables Outstanding': days_receivables,
                    'Inventory Turnover': inventory_turnover,
                    'Days Inventory Outstanding': days_inventory,
                    'Payables Turnover': payables_turnover,
                    'Days Payables Outstanding': days_payables,
                    'Cash Conversion Cycle (Days)': cash_conversion_cycle
                }
                
            except KeyError as e:
                print(f"Missing key in financial data: {e}")
                return False
            except Exception as e:
                print(f"Error calculating Financial Ratios: {e}")
                return False
        
        # Create DataFrame from results
        self.ratios = pd.DataFrame(data)
        return True
    
    def run_full_analysis(self):
        """
        Run all analysis components.
        """
        success = True
        success = success and self.calculate_uca_cash_flow()
        success = success and self.calculate_ebitda()
        success = success and self.calculate_debt_service_coverage()
        success = success and self.calculate_financial_ratios()
        
        return success
    
    def format_excel_sheet(self, ws, df, title):
        """
        Format an Excel worksheet with a DataFrame.
        
        Parameters:
        ----------
        ws : openpyxl.worksheet.worksheet.Worksheet
            Worksheet to format
        df : pandas.DataFrame
            DataFrame to write to the worksheet
        title : str
            Title for the worksheet
        """
        # Add title
        ws.merge_cells('A1:F1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add DataFrame starting at row 3
        rows = dataframe_to_rows(df, index=True, header=True)
        for r_idx, row in enumerate(rows, start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Format the header row
                if r_idx == 3:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                    
                # Format the index column
                if c_idx == 1 and r_idx > 3:
                    cell.font = Font(bold=True)
                    
                # Format numbers
                if isinstance(value, (int, float)) and r_idx > 3 and c_idx > 1:
                    if 'Ratio' in str(df.index[r_idx-4]) or 'Margin' in str(df.index[r_idx-4]) or '%' in str(df.index[r_idx-4]):
                        # Format as percentage with 2 decimal places
                        ws.cell(row=r_idx, column=c_idx).number_format = '0.00%'
                    else:
                        # Format as number with commas and 2 decimal places
                        ws.cell(row=r_idx, column=c_idx).number_format = '#,##0.00'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def add_chart(self, ws, data_range, title, categories_range, position):
        """
        Add a line chart to the worksheet.
        
        Parameters:
        ----------
        ws : openpyxl.worksheet.worksheet.Worksheet
            Worksheet to add chart to
        data_range : str
            Excel range for the data (e.g., 'B4:D10')
        title : str
            Title for the chart
        categories_range : str
            Excel range for the categories (e.g., 'A4:A10')
        position : str
            Cell position for the top-left corner of the chart
        """
        chart = LineChart()
        chart.title = title
        chart.style = 2
        chart.height = 10
        chart.width = 20
        
        data = Reference(ws, min_col=2, min_row=3, max_col=len(self.years)+1, max_row=ws.max_row)
        categories = Reference(ws, min_col=1, min_row=4, max_row=ws.max_row)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        ws.add_chart(chart, position)
    
    def export_to_excel(self, output_file='small_business_cash_flow_analysis.xlsx'):
        """
        Export all analysis results to a formatted Excel file.
        
        Parameters:
        ----------
        output_file : str
            Path for the output Excel file
        """
        if (self.uca_cash_flow is None or self.ebitda_analysis is None or 
            self.dscr_analysis is None or self.ratios is None):
            print("Error: Analysis not complete. Run full analysis first.")
            return False
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Add sheets for each analysis component
        ws_summary = wb.create_sheet("Summary")
        ws_uca = wb.create_sheet("UCA Cash Flow")
        ws_ebitda = wb.create_sheet("EBITDA Analysis")
        ws_dscr = wb.create_sheet("Debt Service Coverage")
        ws_ratios = wb.create_sheet("Financial Ratios")
        
        # Format each sheet
        self.format_excel_sheet(ws_uca, self.uca_cash_flow, "Uniform Credit Analysis (UCA) Cash Flow")
        self.format_excel_sheet(ws_ebitda, self.ebitda_analysis, "EBITDA Analysis")
        self.format_excel_sheet(ws_dscr, self.dscr_analysis, "Debt Service Coverage Analysis")
        self.format_excel_sheet(ws_ratios, self.ratios, "Financial Ratios Analysis")
        
        # Add some charts
        self.add_chart(ws_ebitda, 'B9:D9', 'EBITDA Trend', 'A4:A9', 'A20')
        self.add_chart(ws_dscr, 'B6:D6', 'Debt Service Coverage Ratio Trend', 'A4:A6', 'A15')
        
        # Create summary sheet
        ws_summary.merge_cells('A1:F1')
        ws_summary['A1'] = "Small Business Cash Flow Analysis Summary"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary['A1'].alignment = Alignment(horizontal='center')
        
        # Add key metrics to summary
        key_metrics = [
            ('EBITDA', self.ebitda_analysis.loc['EBITDA']),
            ('EBITDA Margin (%)', self.ebitda_analysis.loc['EBITDA Margin (%)']),
            ('UCA Cash Flow', self.uca_cash_flow.loc['UCA Cash Flow']),
            ('Debt Service Coverage Ratio', self.dscr_analysis.loc['Debt Service Coverage Ratio (DSCR)']),
            ('Current Ratio', self.ratios.loc['Current Ratio']),
            ('Return on Equity (%)', self.ratios.loc['Return on Equity (%)'])
        ]
        
        ws_summary['A3'] = "Key Financial Metrics"
        ws_summary['A3'].font = Font(bold=True, size=12)
        
        for i, (metric, values) in enumerate(key_metrics, start=4):
            ws_summary[f'A{i}'] = metric
            ws_summary[f'A{i}'].font = Font(bold=True)
            
            for j, year in enumerate(self.years, start=0):
                ws_summary.cell(row=3, column=j+2, value=year).font = Font(bold=True)
                ws_summary.cell(row=i, column=j+2, value=values[year])
                
                # Format cells
                if "%" in metric or "Margin" in metric:
                    ws_summary.cell(row=i, column=j+2).number_format = '0.00%'
                else:
                    ws_summary.cell(row=i, column=j+2).number_format = '#,##0.00'
        
        # Add analysis summary and recommendations
        ws_summary['A12'] = "Analysis Summary"
        ws_summary['A12'].font = Font(bold=True, size=12)
        
        last_year = self.years[-1]
        
        # Generate some insights
        dscr = self.dscr_analysis.loc['Debt Service Coverage Ratio (DSCR)', last_year]
        current_ratio = self.ratios.loc['Current Ratio', last_year]
        roe = self.ratios.loc['Return on Equity (%)', last_year]
        
        dscr_comment = (
            "Strong debt service capacity" if dscr >= 1.5 else
            "Adequate debt service capacity" if dscr >= 1.25 else
            "Marginal debt service capacity" if dscr >= 1.1 else
            "Weak debt service capacity - high risk"
        )
        
        liquidity_comment = (
            "Strong liquidity position" if current_ratio >= 2.0 else
            "Adequate liquidity" if current_ratio >= 1.5 else
            "Tight liquidity - monitor closely" if current_ratio >= 1.0 else
            "Liquidity concern - potential cash flow issues"
        )
        
        profitability_comment = (
            "Excellent profitability" if roe >= 20 else
            "Good profitability" if roe >= 15 else
            "Average profitability" if roe >= 10 else
            "Below average profitability - review business model"
        )
        
        ws_summary['A13'] = f"Debt Service: {dscr_comment} (DSCR: {dscr:.2f})"
        ws_summary['A14'] = f"Liquidity: {liquidity_comment} (Current Ratio: {current_ratio:.2f})"
        ws_summary['A15'] = f"Profitability: {profitability_comment} (ROE: {roe:.2f}%)"
        
        # Overall recommendation
        ws_summary['A17'] = "Lending Recommendation"
        ws_summary['A17'].font = Font(bold=True, size=12)
        
        if dscr >= 1.25 and current_ratio >= 1.5 and roe >= 12:
            recommendation = "Strong candidate for lending - low risk profile"
        elif dscr >= 1.1 and current_ratio >= 1.0 and roe >= 8:
            recommendation = "Acceptable candidate for lending - moderate risk profile"
        else:
            recommendation = "Caution advised - higher risk profile, consider additional collateral or guarantees"
        
        ws_summary['A18'] = recommendation
        
        # Add header with company info
        ws_summary['A20'] = "Report Generated:"
        ws_summary['B20'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        
        # Auto-adjust column widths
        for ws in [ws_summary, ws_uca, ws_ebitda, ws_dscr, ws_ratios]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        try:
            wb.save(output_file)
            print(f"Successfully exported analysis to {output_file}")
            return True
        except Exception as e:
            print(f"Error saving Excel file: {e}")
            return False


def main():
    """
    Main function to demonstrate the cash flow analyzer.
    """
    print("\n========== Small Business Cash Flow Analysis Tool ==========\n")
    print("This tool provides comprehensive cash flow analysis for small business lending.")
    print("It includes UCA Cash Flow, EBITDA Analysis, and Debt Service Coverage.\n")
    
    analyzer = SmallBusinessCashFlowAnalyzer()
    
    # Create sample data for demonstration
    print("Creating sample financial data for demonstration...")
    analyzer.create_sample_data(years=3)
    
    # Run full analysis
    print("\nRunning full cash flow analysis...")
    success = analyzer.run_full_analysis()
    
    if success:
        print("\nAnalysis complete! Key results:")
        
        # Show some key results
        last_year = analyzer.years[-1]
        print(f"\nResults for {last_year}:")
        print(f"EBITDA: ${analyzer.ebitda_analysis.loc['EBITDA', last_year]:,.2f}")
        print(f"UCA Cash Flow: ${analyzer.uca_cash_flow.loc['UCA Cash Flow', last_year]:,.2f}")
        print(f"Debt Service Coverage Ratio: {analyzer.dscr_analysis.loc['Debt Service Coverage Ratio (DSCR)', last_year]:.2f}")
        
        # Export to Excel
        print("\nExporting results to Excel...")
        output_file = "small_business_cash_flow_analysis.xlsx"
        analyzer.export_to_excel(output_file)
        
        print(f"\nAnalysis complete! Results exported to {output_file}")
        print("\nThis Excel file can be used for small business lending decisions.")
        print("It includes formatted worksheets for UCA Cash Flow, EBITDA Analysis,")
        print("Debt Service Coverage, and Financial Ratios.")
    else:
        print("\nAnalysis failed. Please check the error messages above.")


if __name__ == "__main__":
    main() 