#!/usr/bin/env python3
"""
Small Business Loan Underwriting Checklist Generator

This script generates a comprehensive Excel-based checklist for small business loan underwriting,
including financial analysis, management assessment, and industry evaluation.

Requirements:
- openpyxl: pip install openpyxl
- pandas: pip install pandas

Usage:
python small_business_loan_underwriting_checklist.py
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_underwriting_checklist(output_file='Small_Business_Loan_Underwriting_Checklist.xlsx'):
    """
    Creates a comprehensive small business loan underwriting checklist in Excel format.
    
    Args:
        output_file (str): The name of the output Excel file
    """
    # Create a Pandas ExcelWriter object
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    
    # Create the main sections as DataFrames
    
    # 1. Borrower Information
    borrower_info = pd.DataFrame({
        'Item': [
            'Legal Business Name',
            'DBA (if applicable)',
            'Business Address',
            'Business Phone',
            'Business Email',
            'Website',
            'Tax ID / EIN',
            'Date Established',
            'Business Structure (LLC, Corp, etc.)',
            'Industry (NAICS Code)',
            'Number of Employees',
            'Annual Revenue (most recent year)',
            'Loan Amount Requested',
            'Loan Purpose',
            'Requested Term',
            'Primary Contact Name',
            'Primary Contact Title',
            'Primary Contact Phone',
            'Primary Contact Email'
        ],
        'Information': [''] * 19,
        'Notes': [''] * 19
    })
    
    # 2. Financial Analysis
    financial_analysis = pd.DataFrame({
        'Item': [
            '-- Required Financial Documents --',
            'Business Tax Returns (3 years)',
            'Personal Tax Returns (3 years)',
            'Business Financial Statements (Income Statement)',
            'Business Financial Statements (Balance Sheet)',
            'Business Financial Statements (Cash Flow Statement)',
            'Interim Financial Statements (YTD)',
            'Business Debt Schedule',
            'Accounts Receivable Aging',
            'Accounts Payable Aging',
            'Bank Statements (12 months)',
            'Business Plan/Projections',
            '-- Financial Ratios & Analysis --',
            'Debt Service Coverage Ratio (DSCR)',
            'EBITDA Calculation',
            'Current Ratio',
            'Quick Ratio',
            'Debt-to-Equity Ratio',
            'Profit Margin',
            'Inventory Turnover',
            'Receivables Turnover',
            'Return on Assets (ROA)',
            'Working Capital',
            'Revenue Trend Analysis',
            'Expense Trend Analysis',
            'Cash Flow Analysis',
            'Break-even Analysis',
            'Global Cash Flow Analysis (Business + Personal)'
        ],
        'Status': [''] * 28,
        'Value/Results': [''] * 28,
        'Notes': [''] * 28
    })
    
    # 3. Management & Ownership Assessment
    management_assessment = pd.DataFrame({
        'Item': [
            '-- Owner/Guarantor Information --',
            'Personal Financial Statement (PFS)',
            'Resume or Management Experience Summary',
            'Credit Report/Score',
            'ID Verification',
            'Ownership Verification (Operating Agreement/Bylaws)',
            'Background Check',
            'Bankruptcy History',
            'Legal History (Judgments, Liens, etc.)',
            'References (Industry, Business, Personal)',
            '-- Management Assessment --',
            'Experience in Industry (years)',
            'Management Structure',
            'Succession Plan',
            'Key Person Risk Assessment',
            'Leadership Team Qualifications'
        ],
        'Status': [''] * 16,
        'Results/Findings': [''] * 16,
        'Notes': [''] * 16
    })
    
    # 4. Industry & Market Analysis
    industry_analysis = pd.DataFrame({
        'Item': [
            'Industry Risk Assessment',
            'Market Size and Growth',
            'Competitive Analysis',
            'Local Market Conditions',
            'Industry Trends',
            'Seasonal Factors',
            'Regulatory Environment',
            'Technology Disruption Risk',
            'Supply Chain Analysis',
            'Customer Concentration Analysis',
            'Vendor Concentration Analysis',
            'Economic Cycle Sensitivity'
        ],
        'Assessment': [''] * 12,
        'Risk Level': [''] * 12,
        'Notes': [''] * 12
    })
    
    # 5. Collateral Analysis
    collateral_analysis = pd.DataFrame({
        'Item': [
            'Real Estate (Address/Description)',
            'Real Estate Appraisal',
            'Equipment (Description)',
            'Equipment Valuation',
            'Inventory Valuation',
            'Accounts Receivable',
            'Cash/Investments',
            'Other Business Assets',
            'Personal Assets Pledged',
            'Lien Searches (UCC)',
            'Title Reports/Insurance',
            'Environmental Assessment (if applicable)',
            '-- Collateral Coverage --',
            'Total Collateral Value',
            'Loan-to-Value Ratio',
            'Collateral Coverage Ratio'
        ],
        'Status': [''] * 16,
        'Value': [''] * 16,
        'Notes': [''] * 16
    })
    
    # 6. Risk Assessment & Mitigants
    risk_assessment = pd.DataFrame({
        'Risk Category': [
            'Credit Risk',
            'Market Risk',
            'Operational Risk',
            'Financial Risk',
            'Industry Risk',
            'Management Risk',
            'Legal/Regulatory Risk',
            'Collateral Risk',
            'External Factors Risk',
            'Cash Flow Risk'
        ],
        'Risk Description': [''] * 10,
        'Risk Level (Low/Medium/High)': [''] * 10,
        'Mitigating Factors': [''] * 10,
        'Action Items': [''] * 10
    })
    
    # 7. Loan Structure & Terms
    loan_structure = pd.DataFrame({
        'Item': [
            'Loan Amount',
            'Loan Type',
            'Interest Rate',
            'Term (months)',
            'Payment Frequency',
            'Payment Amount',
            'Collateral',
            'Guarantors',
            'Financial Covenants',
            'Reporting Requirements',
            'Prepayment Penalties',
            'Disbursement Structure',
            'Fees',
            'Insurance Requirements',
            'Special Conditions'
        ],
        'Details': [''] * 15,
        'Notes': [''] * 15
    })
    
    # 8. Compliance & Documentation
    compliance_docs = pd.DataFrame({
        'Item': [
            'BSA/AML Compliance Check',
            'OFAC Check',
            'Patriot Act Compliance',
            'Corporate Documents (Articles, Bylaws, etc.)',
            'Business License/Permits',
            'Insurance Verification',
            'Environmental Certification (if needed)',
            'Flood Zone Certification (if needed)',
            'Leases (if applicable)',
            'Franchise Agreements (if applicable)',
            'Certificate of Good Standing',
            'IRS Form 4506-T',
            'Loan Application Completeness',
            'Disclosures Provided',
            'Signature Authority Verification'
        ],
        'Status': [''] * 15,
        'Date Completed': [''] * 15,
        'Notes': [''] * 15
    })
    
    # 9. Final Decision & Approvals
    final_decision = pd.DataFrame({
        'Item': [
            'Underwriter Recommendation',
            'Loan Committee Decision',
            'Approved Loan Amount',
            'Approved Interest Rate',
            'Approved Term',
            'Special Conditions',
            'Required Guarantors',
            'Required Collateral',
            'Decision Date',
            'Decision Rationale',
            'Approval Level/Authority',
            'Exceptions to Policy',
            'Expiration of Approval',
            'Next Steps',
            'Final Signatures Required'
        ],
        'Details': [''] * 15,
        'Date': [''] * 15,
        'Notes': [''] * 15
    })
    
    # Write each DataFrame to a separate sheet
    borrower_info.to_excel(writer, sheet_name='1. Borrower Information', index=False)
    financial_analysis.to_excel(writer, sheet_name='2. Financial Analysis', index=False)
    management_assessment.to_excel(writer, sheet_name='3. Management Assessment', index=False)
    industry_analysis.to_excel(writer, sheet_name='4. Industry Analysis', index=False)
    collateral_analysis.to_excel(writer, sheet_name='5. Collateral Analysis', index=False)
    risk_assessment.to_excel(writer, sheet_name='6. Risk Assessment', index=False)
    loan_structure.to_excel(writer, sheet_name='7. Loan Structure', index=False)
    compliance_docs.to_excel(writer, sheet_name='8. Compliance & Docs', index=False)
    final_decision.to_excel(writer, sheet_name='9. Final Decision', index=False)
    
    # Create a summary sheet
    loan_summary = pd.DataFrame({
        'Information': [
            'Borrower Name',
            'Loan Amount',
            'Loan Purpose',
            'Loan Term',
            'Interest Rate',
            'Debt Service Coverage Ratio',
            'Loan-to-Value Ratio',
            'Credit Score',
            'Industry',
            'Years in Business',
            'Annual Revenue',
            'Underwriter',
            'Date of Analysis',
            'Status',
            'Primary Risk Factors',
            'Primary Mitigating Factors',
            'Recommendation'
        ],
        'Details': [''] * 17
    })
    
    loan_summary.to_excel(writer, sheet_name='Loan Summary', index=False)
    
    # Get the workbook to apply formatting
    workbook = writer.book
    
    # Apply formatting to all sheets
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Format header row
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Apply formatting to header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Format the section headers in the financial analysis
        if sheet_name == '2. Financial Analysis':
            for row_idx, value in enumerate(financial_analysis['Item']):
                if value.startswith('--'):
                    cell = worksheet.cell(row=row_idx+2, column=1)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    # Apply to the entire row
                    for col_idx in range(1, 5):
                        worksheet.cell(row=row_idx+2, column=col_idx).fill = PatternFill(
                            start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Format the section headers in the management assessment
        if sheet_name == '3. Management Assessment':
            for row_idx, value in enumerate(management_assessment['Item']):
                if value.startswith('--'):
                    cell = worksheet.cell(row=row_idx+2, column=1)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    # Apply to the entire row
                    for col_idx in range(1, 5):
                        worksheet.cell(row=row_idx+2, column=col_idx).fill = PatternFill(
                            start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Format the section headers in the collateral analysis
        if sheet_name == '5. Collateral Analysis':
            for row_idx, value in enumerate(collateral_analysis['Item']):
                if value.startswith('--'):
                    cell = worksheet.cell(row=row_idx+2, column=1)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    # Apply to the entire row
                    for col_idx in range(1, 5):
                        worksheet.cell(row=row_idx+2, column=col_idx).fill = PatternFill(
                            start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Set column widths
        column_widths = {
            'A': 40,  # Item column
            'B': 25,  # Status/Information column
            'C': 25,  # Value/Results column
            'D': 40,  # Notes column
            'E': 30   # Extra column if present
        }
        
        for col, width in column_widths.items():
            if get_column_letter(worksheet.max_column) >= col:
                worksheet.column_dimensions[col].width = width
        
        # Add data validation for Risk Level columns
        if sheet_name in ['4. Industry Analysis', '6. Risk Assessment']:
            risk_column = 3 if sheet_name == '4. Industry Analysis' else 3
            risk_dv = DataValidation(type="list", formula1='"Low,Medium,High"', allow_blank=True)
            risk_dv.error = "Please select from the list"
            risk_dv.errorTitle = "Invalid Entry"
            risk_dv.prompt = "Select risk level"
            risk_dv.promptTitle = "Risk Level"
            
            # Add the data validation to the worksheet
            worksheet.add_data_validation(risk_dv)
            
            # Apply to cells
            last_row = worksheet.max_row
            risk_range = f"{get_column_letter(risk_column)}2:{get_column_letter(risk_column)}{last_row}"
            risk_dv.add(risk_range)
        
        # Add data validation for Status columns
        if sheet_name in ['2. Financial Analysis', '3. Management Assessment', '5. Collateral Analysis', '8. Compliance & Docs']:
            status_column = 2
            status_dv = DataValidation(type="list", formula1='"Pending,In Progress,Completed,N/A"', allow_blank=True)
            status_dv.error = "Please select from the list"
            status_dv.errorTitle = "Invalid Entry"
            status_dv.prompt = "Select status"
            status_dv.promptTitle = "Status"
            
            # Add the data validation to the worksheet
            worksheet.add_data_validation(status_dv)
            
            # Apply to cells
            last_row = worksheet.max_row
            status_range = f"{get_column_letter(status_column)}2:{get_column_letter(status_column)}{last_row}"
            status_dv.add(status_range)
    
    # Create an Instructions sheet
    instructions_sheet = workbook.create_sheet("Instructions", 0)
    
    # Add instructions
    instructions = [
        ["SMALL BUSINESS LOAN UNDERWRITING CHECKLIST", ""],
        ["", ""],
        ["Instructions:", ""],
        ["1. This checklist is designed to guide lenders through the small business loan underwriting process.", ""],
        ["2. Each sheet represents a different phase of the underwriting process.", ""],
        ["3. Use the 'Loan Summary' sheet to capture the key details and final recommendation.", ""],
        ["4. Complete each section thoroughly before making a final lending decision.", ""],
        ["5. Some fields include dropdown options - click on cells to see available selections.", ""],
        ["6. Add notes to document your analysis and findings.", ""],
        ["7. All sheets are customizable - add or remove items as needed for your specific lending requirements.", ""],
        ["", ""],
        ["Section Overview:", ""],
        ["- Borrower Information: Basic details about the business and primary contact", ""],
        ["- Financial Analysis: Documentation and ratio analysis of the business financials", ""],
        ["- Management Assessment: Evaluation of the owners, guarantors, and management team", ""],
        ["- Industry Analysis: Assessment of industry risks and market conditions", ""],
        ["- Collateral Analysis: Evaluation of available collateral and coverage ratios", ""],
        ["- Risk Assessment: Identification of key risks and mitigating factors", ""],
        ["- Loan Structure: Details of the proposed loan terms and conditions", ""],
        ["- Compliance & Documentation: Checklist of required regulatory compliance and documentation", ""],
        ["- Final Decision: Record of the approval process and conditions", ""],
        ["- Loan Summary: Executive summary of key loan information and decision", ""]
    ]
    
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = instructions_sheet.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=16)
            elif row_idx == 3 or row_idx == 12:
                cell.font = Font(bold=True)
    
    # Set column widths
    instructions_sheet.column_dimensions['A'].width = 60
    instructions_sheet.column_dimensions['B'].width = 40
    
    # Add a title to the workbook
    instructions_sheet.merge_cells('A1:B1')
    instructions_sheet.cell(1, 1).alignment = Alignment(horizontal='center')
    
    # Save the Excel file
    writer.close()
    
    print(f"Successfully created {output_file}")
    print("The file includes 10 sheets with comprehensive small business loan underwriting sections.")

if __name__ == "__main__":
    create_underwriting_checklist() 